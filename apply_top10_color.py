"""
第1条：对最近十个交易日行业/概念板块涨幅排名前十中出现次数超过2次的板块涂色
"""
import openpyxl
from collections import Counter
from openpyxl.styles import PatternFill

wb = openpyxl.load_workbook('data/板块轮动Top10_v2_行业概念分开.xlsx')

COLOR_MAP = {
    3: 'FFEB9C',
    4: 'F4B942',
    5: 'F4A0A0',
}

def get_fill(count):
    if count >= 6:
        return PatternFill(start_color='C00000', end_color='C00000', fill_type='solid')
    return PatternFill(start_color=COLOR_MAP.get(count, 'FFEB9C'),
                       end_color=COLOR_MAP.get(count, 'FFEB9C'), fill_type='solid')

for sheet_name in ['行业板块', '概念板块']:
    ws = wb[sheet_name]

    counter = Counter()
    cell_map = {}  # {(row, col): board_name}

    for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=False), start=2):
        if row[0].value is None:
            break
        for col_idx in range(1, 11):
            cell = row[col_idx]
            if cell.value:
                board_name = str(cell.value).split(' ')[0].strip()
                counter[board_name] += 1
                cell_map[(row_idx, col_idx)] = board_name

    hot_boards = {b for b, cnt in counter.items() if cnt > 2}

    for (row_idx, col_idx), board_name in cell_map.items():
        if board_name in hot_boards:
            count = counter[board_name]
            ws.cell(row=row_idx, column=col_idx).fill = get_fill(count)

    print(f'{sheet_name}: 涂色板块数={len(hot_boards)}')
    for board, cnt in sorted(counter.items(), key=lambda x: -x[1]):
        if cnt > 2:
            print(f'  {board}: {cnt}次')

wb.save('data/板块轮动Top10_v2_行业概念分开.xlsx')
print('\n涂色完成，已保存')
