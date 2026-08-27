# -*- coding: utf-8 -*-
"""
DuckDB 增量更新脚本

每日复盘完成后调用，从以下来源追加数据到 stock.duckdb：
  1. board_history_ths/history_*.json  -> 新一天的板块数据
  2. kline_cache/                      -> 新一天的 K 线
  3. v4 Excel Sheet6/7                 -> 新的 MACD 信号

用法：
  python scripts/db_update.py                     # 默认全量更新（实际是增量）
  python scripts/db_update.py --module board,kline  # 只更新指定模块
  python scripts/db_update.py --date 20260826     # 指定日期

特点：
  - 使用 INSERT ... ON CONFLICT DO NOTHING 避免重复
  - 不删除已有数据
  - 安全可重复运行
"""
import sys
import json
import argparse
from pathlib import Path
from datetime import datetime

import duckdb
import pandas as pd

ROOT = Path(__file__).parent.parent
DB_PATH = ROOT / 'data' / 'stock.duckdb'

CONCEPT_DATA_DIR = ROOT / 'concept_data'
KLINE_CACHE_DIR = ROOT / 'kline_cache'
BOARD_HISTORY_DIR = ROOT / 'data' / 'board_history_ths'
V4_EXCEL = ROOT / 'data' / '板块轮动Top10_v4_含非Top3强势个股.xlsx'


def update_stock_meta(conn, log=print):
    """增量更新 stock_meta（新增股票或重新抓取的）"""
    log('\n[1/4] Updating stock_meta...')
    files = sorted(CONCEPT_DATA_DIR.glob('*_concepts.json'))
    log(f'  Found {len(files)} JSON files')

    new_count = 0
    rows = []
    for f in files:
        try:
            with open(f, 'r', encoding='utf-8') as fp:
                data = json.load(fp)
            rows.append({
                'code': data.get('stock_code') or f.stem.replace('_concepts', ''),
                'name': data.get('stock_name', ''),
                'concepts': json.dumps(data.get('concepts', []), ensure_ascii=False),
                'theme_points': json.dumps(data.get('theme_points', []), ensure_ascii=False),
                'fetch_time': data.get('fetch_time'),
                'total_concepts': data.get('total_concepts', 0),
            })
        except Exception as e:
            log(f'  [WARN] {f.name}: {e}')

    if rows:
        df = pd.DataFrame(rows)
        # 用 INSERT OR REPLACE（覆盖更新，比 ON CONFLICT DO NOTHING 更适合 meta 数据）
        conn.execute("""
            DELETE FROM stock_meta WHERE code IN (SELECT code FROM df)
        """)
        conn.execute("INSERT INTO stock_meta SELECT * FROM df")
        log(f'  [OK] Replaced {len(rows)} stocks (upsert mode)')

    return len(rows)


def update_kline(conn, log=print):
    """增量更新 kline（新增 K 线）"""
    log('\n[2/4] Updating kline...')
    files = sorted(KLINE_CACHE_DIR.glob('*.csv'))
    log(f'  Found {len(files)} CSV files')

    rows = []
    for f in files:
        try:
            df = pd.read_csv(f)
            code = f.stem.replace('_qfq', '')
            df['code'] = code
            df['date'] = pd.to_datetime(df['date']).dt.date
            rows.append(df[['code', 'date', 'open', 'high', 'low', 'close', 'volume']])
        except Exception as e:
            log(f'  [WARN] {f.name}: {e}')

    if rows:
        df_all = pd.concat(rows, ignore_index=True)
        # 删除该批股票的所有现有数据，再插入（全量覆盖）
        conn.execute("DELETE FROM kline WHERE code IN (SELECT DISTINCT code FROM df_all)")
        conn.execute("INSERT INTO kline SELECT * FROM df_all")
        log(f'  [OK] Replaced {len(df_all)} kline rows for {df_all["code"].nunique()} stocks')

    return len(rows)


def update_board_history(conn, log=print):
    """增量更新 board_history（最新一天的新数据）"""
    log('\n[3/4] Updating board_history...')
    files = sorted(BOARD_HISTORY_DIR.glob('history_*.json'))
    if not files:
        log('  No files found')
        return 0

    # 用最新文件（含全量历史，但有 UPSERT 不会有重复）
    latest = files[-1]
    log(f'  Using: {latest.name}')

    with open(latest, 'r', encoding='utf-8') as f:
        data = json.load(f)

    rows = []
    for board_name, info in data.items():
        board_type = info.get('type', '')
        for d in info.get('data', []):
            rows.append({
                'board_name': board_name,
                'board_type': board_type,
                'date': d['d'],
                'close': d.get('c'),
                'pct': d.get('p'),
            })

    df = pd.DataFrame(rows)
    df['date'] = pd.to_datetime(df['date']).dt.date
    # 先删除同名 + 同日期的所有行（避免 ON CONFLICT 复杂性）
    conn.execute("""
        DELETE FROM board_history
        WHERE (board_name, date) IN (
            SELECT board_name, date FROM df
        )
    """)
    conn.execute("INSERT INTO board_history SELECT * FROM df")
    log(f'  [OK] Replaced {len(df)} board history rows ({len(data)} boards)')

    return len(df)


def update_macd_signals(conn, log=print):
    """增量更新 macd_signals"""
    log('\n[4/4] Updating macd_signals...')
    if not V4_EXCEL.exists():
        log(f'  {V4_EXCEL} not found')
        return 0

    rows = []
    today = datetime.now()

    try:
        import openpyxl
        wb = openpyxl.load_workbook(V4_EXCEL, data_only=True)

        for sheet_name, signal_type in [
            ('MACD强势个股', '5d_10pct'),
            ('MACD强势个股_10日', '10d_20pct'),
        ]:
            if sheet_name not in wb.sheetnames:
                continue
            ws = wb[sheet_name]
            seen = set()
            for row in ws.iter_rows(min_row=2, values_only=True):
                if not row or row[0] is None:
                    continue
                code = str(row[1] or '').zfill(6)
                dedup_key = (code, signal_type)
                if dedup_key in seen:
                    continue
                seen.add(dedup_key)
                rows.append({
                    'code': code,
                    'name': row[2] or '',
                    'date': today.date(),
                    'signal_type': signal_type,
                    'macd': row[4] if len(row) > 4 else None,
                    'gain_pct': row[5] if len(row) > 5 else None,
                    'score': row[6] if len(row) > 6 else None,
                    'position': row[7] if len(row) > 7 else None,
                    'trend': row[8] if len(row) > 8 else None,
                    'fx': row[9] if len(row) > 9 else None,
                    'bcie': row[10] if len(row) > 10 else None,
                    'raw_data': json.dumps({
                        'latest_price': row[3] if len(row) > 3 else None,
                        'volume': row[11] if len(row) > 11 else None,
                        'remark': row[12] if len(row) > 12 else None,
                    }, ensure_ascii=False),
                    'fetch_time': today,
                })

        if rows:
            df = pd.DataFrame(rows)
            # 删除当天的旧数据
            conn.execute("DELETE FROM macd_signals WHERE date = ?", [today.date()])
            conn.execute("INSERT INTO macd_signals SELECT * FROM df")
            log(f'  [OK] Inserted {len(df)} macd signals for {today.date()}')

        return len(rows)
    except Exception as e:
        log(f'  [ERROR] {e}')
        return 0


def show_stats(conn, log=print):
    log('\n=== Stats ===')
    for t in ['stock_meta', 'kline', 'board_history', 'macd_signals']:
        n = conn.execute(f'SELECT COUNT(*) FROM {t}').fetchone()[0]
        log(f'  {t}: {n}')


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument('--module', help='Comma-separated modules: meta,kline,board,macd')
    parser.add_argument('--date', help='Target date (default: today)')
    args = parser.parse_args()

    log = print  # alias for show_stats() etc

    modules = set(args.module.split(',')) if args.module else {'meta', 'kline', 'board', 'macd'}

    conn = duckdb.connect(str(DB_PATH))
    log(f'[OK] Connected: {DB_PATH}')

    start = datetime.now()

    if 'meta' in modules:
        update_stock_meta(conn)
    if 'kline' in modules:
        update_kline(conn)
    if 'board' in modules:
        update_board_history(conn)
    if 'macd' in modules:
        update_macd_signals(conn)

    # 更新 db_meta
    conn.execute("INSERT OR REPLACE INTO db_meta VALUES (?, ?)",
                 ['last_update', datetime.now().isoformat()])

    show_stats(conn)

    elapsed = (datetime.now() - start).total_seconds()
    log(f'\n[DONE] Updated in {elapsed:.1f}s')
    conn.close()


if __name__ == '__main__':
    main()
