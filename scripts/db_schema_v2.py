# -*- coding: utf-8 -*-
"""
新增 Sheet4/5 数据表 + 迁移历史数据
DB 先行架构 v1.0

新增表:
  - top3_stocks:     每日Top3强势个股
  - non_top3_stocks: 非Top3板块强势个股

来源:
  - 从 v4 Excel 的 Sheet4/Sheet5 历史数据迁移
  - 每日 wencai 查询增量写入
"""
import sys
from pathlib import Path
import json
import time
import openpyxl
import duckdb
import pandas as pd
import requests

ROOT = Path(__file__).parent.parent
DB_PATH = ROOT / 'data' / 'stock.duckdb'
V4 = ROOT / 'data' / '板块轮动Top10_v4_含非Top3强势个股.xlsx'


# ========== 1. 创建新表 ==========
def create_tables():
    conn = duckdb.connect(str(DB_PATH))

    conn.execute("""
        CREATE SEQUENCE IF NOT EXISTS top3_id;
    """)

    conn.execute("""
        CREATE TABLE IF NOT EXISTS top3_stocks (
            date        DATE,
            board_name  VARCHAR,
            board_type  VARCHAR,
            board_pct   DOUBLE,
            rank_       SMALLINT,
            stock_code  VARCHAR,
            stock_name  VARCHAR,
            stock_pct   DOUBLE,
            fetch_time  TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            PRIMARY KEY (date, board_name, rank_, stock_code)
        )
    """)

    conn.execute("""
        CREATE TABLE IF NOT EXISTS non_top3_stocks (
            date        DATE,
            stock_code  VARCHAR,
            stock_name  VARCHAR,
            board_name  VARCHAR,
            board_pct   DOUBLE,
            stock_pct   DOUBLE,
            fetch_time  TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            PRIMARY KEY (date, stock_code)
        )
    """)

    conn.close()
    print('[OK] Tables created')


# ========== 2. 从 v4 Excel 迁移历史数据 ==========
def migrate_from_v4():
    conn = duckdb.connect(str(DB_PATH))

    # --- Sheet4: 每日Top3强势个股 ---
    wb = openpyxl.load_workbook(str(V4), data_only=True)
    ws4 = wb['每日Top3强势个股']

    rows4 = []
    current_date = None
    for row in ws4.iter_rows(min_row=2, values_only=True):
        date, board, board_pct, btype, stock, stock_pct = row[:6]
        if date:
            current_date = date
        if not current_date or current_date is None:
            continue
        if stock and stock not in ('—', '无', '', None):
            rows4.append({
                'date': current_date,
                'board_name': board,
                'board_type': btype,
                'board_pct': float(str(board_pct).replace('%', '').replace('+', '')) if board_pct else 0,
                'rank_': 0,  # rank filled below
                'stock_code': '',
                'stock_name': stock,
                'stock_pct': float(str(stock_pct).replace('%', '').replace('+', '')) if stock_pct else 0,
            })

    # assign rank per date
    if rows4:
        df4 = pd.DataFrame(rows4)
        df4['rank_'] = df4.groupby('date').cumcount() + 1
        df4['fetch_time'] = pd.Timestamp.now()
        conn.execute("DELETE FROM top3_stocks")
        conn.execute("INSERT INTO top3_stocks SELECT * FROM df4")
        print(f'[OK] top3_stocks: {len(df4)} rows migrated')
    else:
        print('[SKIP] top3_stocks: no data in v4')

    # --- Sheet5: 非Top3板块强势个股 ---
    ws5 = wb['非Top3板块强势个股']
    rows5 = []
    current_date = None
    for row in ws5.iter_rows(min_row=2, values_only=True):
        date, code, name, board, board_pct, stock_pct = row[:6]
        if date:
            current_date = date
        if not current_date or current_date is None:
            continue
        if name and name not in ('—', '无', '', None) and not str(name).startswith('无'):
            rows5.append({
                'date': current_date,
                'stock_code': str(code).strip() if code else '',
                'stock_name': name,
                'board_name': board or '',
                'board_pct': float(str(board_pct).replace('%', '').replace('+', '')) if board_pct else 0,
                'stock_pct': float(str(stock_pct).replace('%', '').replace('+', '')) if stock_pct else 0,
            })

    if rows5:
        df5 = pd.DataFrame(rows5)
        df5['fetch_time'] = pd.Timestamp.now()
        conn.execute("DELETE FROM non_top3_stocks")
        conn.execute("INSERT INTO non_top3_stocks SELECT * FROM df5")
        print(f'[OK] non_top3_stocks: {len(df5)} rows migrated')
    else:
        print('[SKIP] non_top3_stocks: no data in v4')

    conn.close()


# ========== 3. 查询缺失日期 ==========
def get_missing_dates(table, date_col, n_days=10):
    """从 akshare 获取最近 n_days 个交易日，返回缺失的日期"""
    import akshare as ak
    now = pd.Timestamp.now()
    # 获取最近 n_days 交易日
    df = ak.tool_trade_date_hist_sina()
    df['date'] = pd.to_datetime(df['trade_date']).dt.date
    today = now.date()
    # 今天之前的 n_days 个交易日
    past = df[df['date'] <= today].tail(n_days)
    dates = sorted(past['date'].tolist(), reverse=True)
    print(f'  最近 {n_days} 交易日: {[str(d) for d in dates]}')

    conn = duckdb.connect(str(DB_PATH))
    existing = conn.execute(f"SELECT DISTINCT {date_col}::DATE FROM {table}").df()
    existing_dates = set(pd.to_datetime(existing.iloc[:, 0]).dt.date.tolist())
    conn.close()

    missing = [d for d in dates if d not in existing_dates]
    print(f'  已有: {len(existing_dates)} 天, 缺失: {len(missing)} 天')
    return missing


def main():
    print('=== DB Schema v2: Sheet4/5 Tables ===\n')

    print('[1/2] Creating tables...')
    create_tables()

    print('\n[2/2] Migrating from v4 Excel...')
    migrate_from_v4()

    print('\n[3] Verifying...')
    conn = duckdb.connect(str(DB_PATH))
    for tbl, date_col in [('top3_stocks', 'date'), ('non_top3_stocks', 'date')]:
        cnt = conn.execute(f"SELECT COUNT(*) FROM {tbl}").fetchone()[0]
        dates = conn.execute(f"SELECT {date_col}, COUNT(*) FROM {tbl} GROUP BY {date_col} ORDER BY {date_col} DESC").fetchall()
        print(f'  {tbl}: {cnt} rows, {len(dates)} dates')
        for d, n in dates[:5]:
            print(f'    {d}: {n} rows')
    conn.close()

    print('\n[DONE] Schema v2 ready')


if __name__ == '__main__':
    main()
