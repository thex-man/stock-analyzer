# -*- coding: utf-8 -*-
"""从备份 v4 Excel 导入历史 MACD 信号（MACD强势个股_v2 / _10日_v2，16天截面）到 macd_signals"""
import sys
from pathlib import Path
import openpyxl
import duckdb
import pandas as pd

BAK = Path('data/板块轮动Top10_v4_含非Top3强势个股.bak.20260827_004301.xlsx')
DB = Path('data/stock.duckdb')

wb = openpyxl.load_workbook(str(BAK), data_only=True)

conn = duckdb.connect(str(DB))

def import_sheet(sheet_name, signal_type):
    ws = wb[sheet_name]
    # 找表头行
    header_row = None
    for i, row in enumerate(ws.iter_rows(min_row=1, max_row=3, values_only=True), 1):
        if row[0] == '信号日':
            header_row = i
            headers = list(row)
            break
    if header_row is None:
        print(f'[SKIP] {sheet_name}: no header')
        return

    print(f'[{sheet_name}] header row {header_row}: {[h for h in headers if h]}')
    rows = []
    for row in ws.iter_rows(min_row=header_row + 1, values_only=True):
        if not row or not row[0]:
            continue
        d = str(row[0])[:10]
        code = str(row[2]).strip() if row[2] else ''
        name = str(row[3]).strip() if row[3] else ''
        if not code or not name:
            continue
        def f(v):
            try:
                return float(v)
            except (TypeError, ValueError):
                return None
        rows.append({
            'date': d, 'code': code, 'name': name,
            'macd': f(row[5]), 'gain_pct': f(row[headers.index([h for h in headers if h and '涨幅' in str(h)][0])] if any('涨幅' in str(h or '') for h in headers) else None),
            'score': None, 'position': None, 'trend': None, 'fx': None, 'bcie': None,
            'signal_type': signal_type,
            'raw_data': '{}',
            'fetch_time': pd.Timestamp.now(),
        })
    df = pd.DataFrame(rows)
    # 只导入 DB 缺的日期
    for d in sorted(df['date'].unique()):
        exists = conn.execute(
            "SELECT COUNT(*) FROM macd_signals WHERE date::VARCHAR=? AND signal_type=?",
            [d, signal_type]).fetchone()[0]
        if exists > 0:
            print(f'  {d}: already {exists} rows, skip')
            continue
        sub = df[df['date'] == d]
        conn.execute("""
            INSERT INTO macd_signals (code, name, date, signal_type, macd, gain_pct,
                score, position, trend, fx, bcie, raw_data, fetch_time)
            SELECT code, name, date::DATE, signal_type, macd, gain_pct,
                score, position, trend, fx, bcie, raw_data, fetch_time FROM sub
        """)
        print(f'  {d}: imported {len(sub)} rows ({signal_type})')
    conn.commit()

import_sheet('MACD强势个股_v2', '5d_10pct')
import_sheet('MACD强势个股_10日_v2', '10d_20pct')

df = conn.execute("SELECT date, signal_type, COUNT(*) FROM macd_signals GROUP BY date, signal_type ORDER BY date DESC, signal_type").df()
print(df.to_string())
conn.close()
