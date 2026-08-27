# -*- coding: utf-8 -*-
"""
补 MACD 历史信号的缠论字段（缺口2）
从备份 v4 的 _v2 sheets 重新解析：缠论分数/位置/趋势/分型/背驰
"""
from pathlib import Path
import openpyxl
import duckdb
import pandas as pd

BAK = Path('data/板块轮动Top10_v4_含非Top3强势个股.bak.20260827_004301.xlsx')
DB = Path('data/stock.duckdb')

wb = openpyxl.load_workbook(str(BAK), data_only=True)
conn = duckdb.connect(str(DB))

def f(v):
    try:
        if v is None or v in ('—', ''):
            return None
        return float(v)
    except (ValueError, TypeError):
        return None

for sheet_name, stype in [('MACD强势个股_v2', '5d_10pct'), ('MACD强势个股_10日_v2', '10d_20pct')]:
    ws = wb[sheet_name]
    headers = [str(h) if h else '' for h in next(ws.iter_rows(min_row=1, max_row=1, values_only=True))]
    idx = {h: i for i, h in enumerate(headers)}
    print(f'[{sheet_name}] 列: {[h for h in headers if h]}')
    i_score = idx.get('缠论分数')
    i_pos = idx.get('位置')
    i_trend = idx.get('趋势')
    i_fx = idx.get('分型')
    i_bcie = idx.get('背驰')
    if i_score is None:
        print(f'  无缠论分数列，跳过')
        continue

    updated = 0
    has_val = 0
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or not row[0]:
            continue
        d = str(row[0])[:10]
        code = str(row[2]).strip() if row[2] else ''
        if not code or len(code) != 6:
            continue
        score = f(row[i_score]) if i_score is not None else None
        pos = str(row[i_pos]) if i_pos is not None and row[i_pos] else ''
        trend = str(row[i_trend]) if i_trend is not None and row[i_trend] else ''
        fx = str(row[i_fx]) if i_fx is not None and row[i_fx] else ''
        bcie = str(row[i_bcie]) if i_bcie is not None and row[i_bcie] else ''
        if score is not None:
            has_val += 1
        # 只更新当前为 NULL 的历史行
        conn.execute("""
            UPDATE macd_signals SET score=?, position=?, trend=?, fx=?, bcie=?
            WHERE date::VARCHAR=? AND code=? AND signal_type=? AND score IS NULL
        """, [score, pos, trend, fx, bcie, d, code, stype])
        updated += 1

    conn.commit()
    print(f'  处理 {updated} 行，其中有缠论分数值 {has_val} 行')

# 验证
r = conn.execute("""
    SELECT SUM(CASE WHEN score IS NULL THEN 1 ELSE 0 END) null_score, COUNT(*) total
    FROM macd_signals WHERE date < '2026-08-26'
""").fetchone()
print(f'\n0826 前: score 仍 NULL {r[0]}/{r[1]}')
conn.close()
