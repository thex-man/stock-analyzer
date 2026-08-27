# -*- coding: utf-8 -*-
"""
每日同步：v4 Excel → DuckDB（DB 先行架构 v1.0）
=================================================
每日复盘最后一步。把当天 wencai 抓取的结果从 v4 Excel 写入 DB：
  - Sheet4 每日Top3强势个股   → top3_stocks
  - Sheet5 非Top3板块强势个股 → non_top3_stocks
  - Sheet6 MACD强势个股        → macd_signals (5d_10pct)
  - Sheet7 MACD强势个股_10日   → macd_signals (10d_20pct)

幂等：同日数据先 DELETE 再 INSERT。
用法: python scripts/db_sync_today.py
"""
import sys
from pathlib import Path
from datetime import datetime, time as dtime
import openpyxl
import duckdb
import pandas as pd

ROOT = Path(__file__).parent.parent
DB_PATH = ROOT / 'data' / 'stock.duckdb'
V4 = ROOT / 'data' / '板块轮动Top10_v4_含非Top3强势个股.xlsx'


def trade_date():
    """9:15 前 = 上一交易日"""
    now = datetime.now()
    if now.time() < dtime(9, 15):
        return (now - pd.Timedelta(days=1)).date()
    return now.date()


def f(v):
    try:
        if v is None or v in ('—', '', '无'):
            return None
        return float(str(v).replace('%', '').replace('+', ''))
    except (ValueError, TypeError):
        return None


def main():
    td = trade_date()
    print(f'[SYNC] Trade date: {td}')

    wb = openpyxl.load_workbook(str(V4), data_only=True)
    conn = duckdb.connect(str(DB_PATH))
    ts = pd.Timestamp.now()

    # ---- Sheet4 → top3_stocks ----
    if '每日Top3强势个股' in wb.sheetnames:
        ws = wb['每日Top3强势个股']
        rows = []
        cur_date = None
        for row in ws.iter_rows(min_row=2, values_only=True):
            d = str(row[0])[:10] if row[0] else cur_date
            if d:
                cur_date = d
            stock = row[4]
            if not cur_date or not stock or str(stock).strip() in ('—', '无', ''):
                continue
            parts = str(stock).split()
            scode = parts[1].strip() if len(parts) >= 2 else ''
            sname = parts[0].strip() if parts else ''
            rows.append((cur_date, row[1] or '', row[3] or '行业', f(row[2]) or 0, 1, scode, sname, f(row[5]) or 0, ts))
        if rows:
            df = pd.DataFrame(rows, columns=['date', 'board_name', 'board_type', 'board_pct', 'rank_', 'stock_code', 'stock_name', 'stock_pct', 'fetch_time'])
            for d in df['date'].unique():
                conn.execute("DELETE FROM top3_stocks WHERE date::VARCHAR=?", [d])
            conn.execute("""INSERT INTO top3_stocks (date, board_name, board_type, board_pct, rank_, stock_code, stock_name, stock_pct, fetch_time)
                            SELECT date::DATE, board_name, board_type, board_pct, rank_, stock_code, stock_name, stock_pct, fetch_time FROM df""")
            print(f'  top3_stocks: {len(df)} rows')
        else:
            print('  top3_stocks: no data (skip)')

    # ---- Sheet5 → non_top3_stocks ----
    if '非Top3板块强势个股' in wb.sheetnames:
        ws = wb['非Top3板块强势个股']
        rows = []
        cur_date = None
        for row in ws.iter_rows(min_row=2, values_only=True):
            d = str(row[0])[:10] if row[0] else cur_date
            if d:
                cur_date = d
            name = row[2]
            if not cur_date or not name or str(name).startswith('无') or str(name).strip() in ('—', ''):
                continue
            rows.append((cur_date, str(row[1]).strip() if row[1] else '', str(name).strip(),
                         row[3] or '', f(row[4]) or 0, f(row[5]) or 0, ts))
        if rows:
            df = pd.DataFrame(rows, columns=['date', 'stock_code', 'stock_name', 'board_name', 'board_pct', 'stock_pct', 'fetch_time'])
            for d in df['date'].unique():
                conn.execute("DELETE FROM non_top3_stocks WHERE date::VARCHAR=?", [d])
            conn.execute("""INSERT INTO non_top3_stocks (date, stock_code, stock_name, board_name, board_pct, stock_pct, fetch_time)
                            SELECT date::DATE, stock_code, stock_name, board_name, board_pct, stock_pct, fetch_time FROM df""")
            print(f'  non_top3_stocks: {len(df)} rows')
            # 概念为空的用规范化匹配回填
            empty = df[(df['board_name'].isna()) | (df['board_name'] == '')]
            if len(empty) > 0:
                print(f'  [INFO] {len(empty)} 行概念为空，稍后可用 db_fix_board_name.py 回填')
        else:
            print('  non_top3_stocks: no data (skip)')

    # ---- Sheet6/7 → macd_signals ----
    for sheet, stype in [('MACD强势个股', '5d_10pct'), ('MACD强势个股_10日', '10d_20pct')]:
        if sheet not in wb.sheetnames:
            continue
        ws = wb[sheet]
        headers = [str(h) if h else '' for h in next(ws.iter_rows(min_row=1, max_row=1, values_only=True))]
        idx = {h: i for i, h in enumerate(headers)}
        gain_col = next((h for h in headers if '涨幅' in h), None)
        rows = []
        for rank_i, row in enumerate(ws.iter_rows(min_row=2, values_only=True), 1):
            code = str(row[idx.get('代码', 1)] or '').strip()
            name = str(row[idx.get('名称', 2)] or '').strip()
            if not code or not name or len(code) != 6:
                continue
            rows.append((code, name, str(td), stype,
                         f(row[idx.get('MACD', 4)]),
                         f(row[idx.get(gain_col, 5)]) if gain_col else None,
                         f(row[idx.get('缠论分数', 6)]),
                         str(row[idx.get('位置', 7)] or ''), str(row[idx.get('趋势', 8)] or ''),
                         str(row[idx.get('分型', 9)] or ''), str(row[idx.get('背驰', 10)] or ''),
                         '{}', ts))
        if rows:
            df = pd.DataFrame(rows, columns=['code', 'name', 'date', 'signal_type', 'macd', 'gain_pct', 'score', 'position', 'trend', 'fx', 'bcie', 'raw_data', 'fetch_time'])
            df = df.drop_duplicates(subset=['code', 'date', 'signal_type'], keep='first')
            conn.execute("DELETE FROM macd_signals WHERE date::VARCHAR=? AND signal_type=?", [str(td), stype])
            conn.execute("""INSERT INTO macd_signals (code, name, date, signal_type, macd, gain_pct, score, position, trend, fx, bcie, raw_data, fetch_time)
                            SELECT code, name, date::DATE, signal_type, macd, gain_pct, score, position, trend, fx, bcie, raw_data, fetch_time FROM df""")
            print(f'  macd_signals[{stype}]: {len(df)} rows for {td}')

    conn.commit()

    # ---- 统计 ----
    print('\n=== Stats ===')
    for tbl in ['top3_stocks', 'non_top3_stocks', 'macd_signals']:
        n = conn.execute(f"SELECT COUNT(*) FROM {tbl}").fetchone()[0]
        last = conn.execute(f"SELECT MAX(date) FROM {tbl}").fetchone()[0]
        print(f'  {tbl}: {n} rows (latest: {last})')
    conn.close()
    print('\n[DONE]')


if __name__ == '__main__':
    main()
