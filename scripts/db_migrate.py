# -*- coding: utf-8 -*-
"""
DuckDB 数据迁移脚本

从以下来源导入到 stock.duckdb：
  1. concept_data/*.json          -> stock_meta（5688 只股票）
  2. kline_cache/*.csv            -> kline（1403 只股票）
  3. data/board_history_ths/*.json -> board_history（465 个板块）
  4. Excel Sheet6/7               -> macd_signals（MACD/缠论）

用法：
  python scripts/db_migrate.py                      # 全量迁移（首次）
  python scripts/db_migrate.py --skip kline,meta    # 跳过某些模块
  python scripts/db_migrate.py --only meta,kline    # 只迁移某些模块
"""
import sys
import os
import json
import glob
import argparse
import duckdb
import pandas as pd
from pathlib import Path
from datetime import datetime

# 路径
ROOT = Path(__file__).parent.parent
DB_PATH = ROOT / 'data' / 'stock.duckdb'
CONCEPT_DATA_DIR = ROOT / 'concept_data'
KLINE_CACHE_DIR = ROOT / 'kline_cache'
BOARD_HISTORY_DIR = ROOT / 'data' / 'board_history_ths'
V4_EXCEL = ROOT / 'data' / '板块轮动Top10_v4_含非Top3强势个股.xlsx'


def migrate_stock_meta(conn, log=print):
    """导入股票元数据 + 概念（5688 个 JSON）"""
    log('\n[1/4] Migrating stock_meta (concept_data/*.json)...')
    files = sorted(CONCEPT_DATA_DIR.glob('*_concepts.json'))
    log(f'  Found {len(files)} files')

    rows = []
    errors = []
    for i, f in enumerate(files):
        try:
            with open(f, 'r', encoding='utf-8') as fp:
                data = json.load(fp)
            code = data.get('stock_code') or f.stem.replace('_concepts', '')
            rows.append({
                'code': code,
                'name': data.get('stock_name', ''),
                'concepts': json.dumps(data.get('concepts', []), ensure_ascii=False),
                'theme_points': json.dumps(data.get('theme_points', []), ensure_ascii=False),
                'fetch_time': data.get('fetch_time'),
                'total_concepts': data.get('total_concepts', 0),
            })
        except Exception as e:
            errors.append((f.name, str(e)))

        if (i + 1) % 1000 == 0:
            log(f'  ... processed {i + 1}/{len(files)}')

    if rows:
        df = pd.DataFrame(rows)
        # 删除已有数据（重复 ID 用新数据覆盖）
        conn.execute("DELETE FROM stock_meta WHERE code IN (SELECT code FROM df)")
        conn.execute("INSERT INTO stock_meta SELECT * FROM df")
        log(f'  [OK] Inserted {len(rows)} stocks ({len(errors)} errors)')

    return len(rows), len(errors)


def migrate_kline(conn, log=print):
    """导入 K 线（1403 个 CSV）"""
    log('\n[2/4] Migrating kline (kline_cache/*.csv)...')
    files = sorted(KLINE_CACHE_DIR.glob('*.csv'))
    log(f'  Found {len(files)} files')

    rows = []
    errors = []
    for i, f in enumerate(files):
        try:
            df = pd.read_csv(f)
            # code 从文件名提取（如 300001_qfq.csv -> 300001）
            code = f.stem.replace('_qfq', '')
            df['code'] = code
            df['date'] = pd.to_datetime(df['date']).dt.date
            rows.append(df[['code', 'date', 'open', 'high', 'low', 'close', 'volume']])

        except Exception as e:
            errors.append((f.name, str(e)))

        if (i + 1) % 500 == 0:
            log(f'  ... processed {i + 1}/{len(files)}')

    if rows:
        df_all = pd.concat(rows, ignore_index=True)
        conn.execute("DELETE FROM kline WHERE code IN (SELECT DISTINCT code FROM df_all)")
        conn.execute("INSERT INTO kline SELECT * FROM df_all")
        log(f'  [OK] Inserted {len(df_all)} kline rows ({len(errors)} errors)')

    return len(df_all) if rows else 0, len(errors)


def migrate_board_history(conn, log=print):
    """导入板块历史（465 个板块）"""
    log('\n[3/4] Migrating board_history (board_history_ths/history_*.json)...')
    files = sorted(BOARD_HISTORY_DIR.glob('history_*.json'))
    if not files:
        log('  No files found')
        return 0, 0

    latest = files[-1]
    log(f'  Latest: {latest.name}')

    with open(latest, 'r', encoding='utf-8') as f:
        data = json.load(f)

    rows = []
    for board_name, info in data.items():
        board_type = info.get('type', '')
        for d in info.get('data', []):
            rows.append({
                'board_name': board_name,
                'board_type': board_type,
                'date': d['d'],
                'close': d.get('c'),
                'pct': d.get('p'),
            })

    df = pd.DataFrame(rows)
    df['date'] = pd.to_datetime(df['date']).dt.date
    conn.execute("DELETE FROM board_history")
    conn.execute("INSERT INTO board_history SELECT * FROM df")
    log(f'  [OK] Inserted {len(df)} rows ({len(data)} boards)')

    return len(df), 0


def migrate_macd_signals(conn, log=print):
    """导入 MACD/缠论信号（从 v4 Excel Sheet6/7）"""
    log('\n[4/4] Migrating macd_signals (v4 Excel)...')
    if not V4_EXCEL.exists():
        log(f'  {V4_EXCEL} not found, skipping')
        return 0, 0

    rows = []
    today = datetime.now()

    try:
        import openpyxl
        wb = openpyxl.load_workbook(V4_EXCEL, data_only=True)

        for sheet_name, signal_type in [
            ('MACD强势个股', '5d_10pct'),
            ('MACD强势个股_10日', '10d_20pct'),
        ]:
            if sheet_name not in wb.sheetnames:
                continue
            ws = wb[sheet_name]
            seen = set()  # 去重（代码在 Sheet 里出现多次）
            for row in ws.iter_rows(min_row=2, values_only=True):
                if not row or row[0] is None:
                    continue
                code = str(row[1] or '').zfill(6)
                # 去重：同一代码同一类型只保留一次（取第一个出现的）
                dedup_key = (code, signal_type)
                if dedup_key in seen:
                    continue
                seen.add(dedup_key)
                rows.append({
                    'code': str(row[1] or '').zfill(6),
                    'name': row[2] or '',
                    'date': today.date(),  # 当前交易日
                    'signal_type': signal_type,
                    'macd': row[4] if len(row) > 4 else None,
                    'gain_pct': row[5] if len(row) > 5 else None,
                    'score': row[6] if len(row) > 6 else None,
                    'position': row[7] if len(row) > 7 else None,
                    'trend': row[8] if len(row) > 8 else None,
                    'fx': row[9] if len(row) > 9 else None,
                    'bcie': row[10] if len(row) > 10 else None,
                    'raw_data': json.dumps({
                        'latest_price': row[3] if len(row) > 3 else None,
                        'volume': row[11] if len(row) > 11 else None,
                        'remark': row[12] if len(row) > 12 else None,
                    }, ensure_ascii=False),
                    'fetch_time': today,
                })

        if rows:
            df = pd.DataFrame(rows)
            conn.execute("DELETE FROM macd_signals WHERE date = ?", [today.date()])
            conn.execute("INSERT INTO macd_signals SELECT * FROM df")
            log(f'  [OK] Inserted {len(df)} macd signals')
        return len(rows), 0
    except Exception as e:
        log(f'  [ERROR] {e}')
        return 0, 1


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument('--only', help='只导入某些模块（逗号分隔）')
    parser.add_argument('--skip', help='跳过某些模块（逗号分隔）')
    args = parser.parse_args()

    only = set(args.only.split(',')) if args.only else None
    skip = set(args.skip.split(',')) if args.skip else set()

    def should_run(name):
        if only and name not in only:
            return False
        if skip and name in skip:
            return False
        return True

    conn = duckdb.connect(str(DB_PATH))
    print(f'[OK] Connected: {DB_PATH}')

    start = datetime.now()

    if should_run('meta'):
        migrate_stock_meta(conn)
    if should_run('kline'):
        migrate_kline(conn)
    if should_run('board'):
        migrate_board_history(conn)
    if should_run('macd'):
        migrate_macd_signals(conn)

    # 更新 db_meta
    conn.execute("INSERT OR REPLACE INTO db_meta VALUES (?, ?)",
                 ['last_migrate', datetime.now().isoformat()])

    # 最终统计
    print('\n=== Final stats ===')
    for t in ['stock_meta', 'kline', 'board_history', 'macd_signals']:
        n = conn.execute(f'SELECT COUNT(*) FROM {t}').fetchone()[0]
        print(f'  {t}: {n} rows')

    elapsed = (datetime.now() - start).total_seconds()
    print(f'\n[DONE] Migrated in {elapsed:.1f}s')
    conn.close()


if __name__ == '__main__':
    main()
