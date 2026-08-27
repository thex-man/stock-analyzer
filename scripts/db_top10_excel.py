# -*- coding: utf-8 -*-
"""
DB 版 Sheet1/2/3 生成器：从 DuckDB 读取板块历史，生成与 board_top10_excel.py 输出一致的 Excel。

用法：
  python scripts/db_top10_excel.py                    # 默认输出到 data/板块轮动Top10_v2_db.xlsx
  python scripts/db_top10_excel.py --out custom.xlsx  # 自定义输出路径
  python scripts/db_top10_excel.py --compare          # 与原版输出对比
"""
import sys
import argparse
from pathlib import Path
from collections import Counter
from datetime import datetime, date, time as dtime

import duckdb
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

ROOT = Path(__file__).parent.parent
DB_PATH = ROOT / 'data' / 'stock.duckdb'
DEFAULT_OUT = ROOT / 'data' / '板块轮动Top10_v2_db.xlsx'

# 配色（与原版一致）
PALETTE = [
    'FF6B6B', 'FF9F43', 'FECA57', '48DBFB', '1DD1A1',
    'A55EEA', 'FD79A8', '00CEC9', '6C5CE7', 'FDCB6E',
    '74B9FF', '55EFC4', 'FF7675', 'D63031', 'E17055',
    'FAB1A0', '81ECEC', 'B8E994', 'FFA07A', 'FFD93D',
]


def get_last_10_dates(conn, board_type: str = '行业') -> list:
    """取最近 10 个交易日"""
    df = conn.execute("""
        SELECT DISTINCT date FROM board_history
        WHERE board_type = ?
        ORDER BY date DESC LIMIT 10
    """, [board_type]).df()
    # 转换为 date 对象（避免 pandas Timestamp 传给 DuckDB DATE 列出错）
    return [d.date() if hasattr(d, 'date') else d for d in df['date']]


def get_top10_by_date(conn, board_type: str, target_date) -> pd.DataFrame:
    """取某日某类型板块 Top 10（按 pct 降序）"""
    df = conn.execute("""
        SELECT board_name, pct
        FROM board_history
        WHERE board_type = ? AND date = ?
        ORDER BY pct DESC
        LIMIT 10
    """, [board_type, target_date]).df()
    # 确保 pct 是 float
    df['pct'] = df['pct'].astype(float)
    return df


def build_daily_top10(conn, dates: list) -> dict:
    """构建 daily_top10 = {date: {行业: [(pct, name)...], 概念: [...]}}"""
    daily = {}
    for d in dates:
        d_date = d.date() if hasattr(d, 'date') else d
        hy = get_top10_by_date(conn, '行业', d_date)
        gn = get_top10_by_date(conn, '概念', d_date)
        daily[d] = {
            '行业': [(float(row['pct']), str(row['board_name'])) for _, row in hy.iterrows()],
            '概念': [(float(row['pct']), str(row['board_name'])) for _, row in gn.iterrows()],
        }
    return daily


def get_board_counts(daily_top10: dict, board_type: str) -> Counter:
    counts = Counter()
    for date_data in daily_top10.values():
        for pct, name in date_data[board_type]:
            counts[name] += 1
    return counts


def gen_excel(conn, daily_top10: dict, out_path: Path):
    """生成 Excel（与原版结构一致）"""
    hy_counts = get_board_counts(daily_top10, '行业')
    gn_counts = get_board_counts(daily_top10, '概念')

    repeat_hy = {n: c for n, c in hy_counts.items() if c > 2}
    repeat_gn = {n: c for n, c in gn_counts.items() if c > 2}
    all_repeat_names = set(repeat_hy.keys()) | set(repeat_gn.keys())

    # 分配颜色
    name_to_color = {}
    for i, name in enumerate(sorted(all_repeat_names)):
        name_to_color[name] = PALETTE[i % len(PALETTE)]

    wb = Workbook()
    ws1 = wb.active
    ws1.title = '行业板块'
    ws2 = wb.create_sheet('概念板块')
    ws3 = wb.create_sheet('色卡图例')

    thin = Side(style='thin', color='CCCCCC')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    def write_top10_sheet(ws, board_type):
        # 表头
        ws.cell(row=1, column=1, value='日期')
        for i in range(10):
            ws.cell(row=1, column=i + 2, value=f'第{i + 1}名')
        ws.column_dimensions['A'].width = 14
        for i in range(10):
            ws.column_dimensions[get_column_letter(i + 2)].width = 18

        # 数据
        for row_i, (date_obj, items_dict) in enumerate(daily_top10.items(), 2):
            date_str = date_obj.strftime('%Y-%m-%d')
            c = ws.cell(row=row_i, column=1, value=date_str)
            c.font = Font(bold=True)
            c.alignment = Alignment(horizontal='center')
            c.border = border
            items = items_dict.get(board_type, [])
            for i, (pct, name) in enumerate(items):
                c = ws.cell(row=row_i, column=i + 2, value=f'{name} {pct:+.2f}%')
                c.alignment = Alignment(horizontal='center')
                c.border = border
                if name in name_to_color:
                    c.fill = PatternFill(
                        start_color=name_to_color[name],
                        end_color=name_to_color[name],
                        fill_type='solid',
                    )
                    c.font = Font(bold=True, color='000000')

    write_top10_sheet(ws1, '行业')
    write_top10_sheet(ws2, '概念')

    # 色卡图例
    ws3.column_dimensions['A'].width = 18
    ws3.column_dimensions['B'].width = 10
    ws3.column_dimensions['C'].width = 18
    ws3.column_dimensions['D'].width = 10
    ws3.column_dimensions['E'].width = 32

    headers = ['行业板块', '上榜次数', '概念板块', '上榜次数', '★ 灰色=出现1-2次，不涂色']
    for i, h in enumerate(headers, 1):
        c = ws3.cell(row=1, column=i, value=h)
        c.font = Font(bold=True, color='FFFFFF')
        c.fill = PatternFill(start_color='2F5496', end_color='2F5496', fill_type='solid')
        c.alignment = Alignment(horizontal='center')
        c.border = border

    max_len = max(len(repeat_hy), len(repeat_gn), 1)
    sorted_hy = sorted(repeat_hy.items(), key=lambda x: -x[1])
    sorted_gn = sorted(repeat_gn.items(), key=lambda x: -x[1])
    for i in range(max_len):
        row_i = i + 2
        if i < len(sorted_hy):
            name, count = sorted_hy[i]
            c = ws3.cell(row=row_i, column=1, value=name)
            c.fill = PatternFill(
                start_color=name_to_color[name],
                end_color=name_to_color[name],
                fill_type='solid',
            )
            c.font = Font(bold=True)
            c.border = border
            c2 = ws3.cell(row=row_i, column=2, value=f'{count}次')
            c2.alignment = Alignment(horizontal='center')
            c2.border = border
        if i < len(sorted_gn):
            name, count = sorted_gn[i]
            c = ws3.cell(row=row_i, column=3, value=name)
            c.fill = PatternFill(
                start_color=name_to_color[name],
                end_color=name_to_color[name],
                fill_type='solid',
            )
            c.font = Font(bold=True)
            c.border = border
            c2 = ws3.cell(row=row_i, column=4, value=f'{count}次')
            c2.alignment = Alignment(horizontal='center')
            c2.border = border

    wb.save(str(out_path))
    print(f'[OK] Saved: {out_path}')


def compare_with_original(conn, daily_top10: dict, original_path: Path):
    """与原版 v2 Excel 对比"""
    if not original_path.exists():
        print(f'[WARN] Original not found: {original_path}, skip compare')
        return

    print(f'\n=== Compare with {original_path.name} ===')
    import openpyxl
    wb = openpyxl.load_workbook(original_path, data_only=True)

    for sheet_name, board_type in [('行业板块', '行业'), ('概念板块', '概念')]:
        if sheet_name not in wb.sheetnames:
            print(f'  [SKIP] {sheet_name} not in original')
            continue
        ws = wb[sheet_name]
        # 取每个日期的第一名对比
        diffs = 0
        for date_obj, items_dict in daily_top10.items():
            date_str = date_obj.strftime('%Y-%m-%d')
            items = items_dict.get(board_type, [])
            # 找原表里的对应日期
            for row in ws.iter_rows(min_row=2, values_only=True):
                if row[0] == date_str:
                    # 取原表第 1 名
                    if row[1] is None:
                        continue
                    # 解析 "name +pct%"
                    try:
                        orig_name = row[1].rsplit(' ', 1)[0]
                        orig_pct = float(row[1].rsplit(' ', 1)[1].rstrip('%'))
                    except (ValueError, IndexError):
                        continue
                    db_pct, db_name = items[0] if items else (None, None)
                    if db_name is None:
                        continue
                    if abs(orig_pct - db_pct) > 0.01 or orig_name != db_name:
                        diffs += 1
                        print(f'  [DIFF] {date_str} {board_type} #1: orig={orig_name} {orig_pct:+.2f}% vs db={db_name} {db_pct:+.2f}%')
                    break
        if diffs == 0:
            print(f'  [OK] {sheet_name}: all match')


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument('--out', default=str(DEFAULT_OUT), help='Output Excel path')
    parser.add_argument('--compare', action='store_true', help='Compare with original')
    parser.add_argument('--compare-orig', default=str(ROOT / 'data' / '板块轮动Top10_v2_行业概念分开.xlsx'))
    args = parser.parse_args()

    print(f'[OK] Connecting to {DB_PATH}')
    conn = duckdb.connect(str(DB_PATH), read_only=True)

    # 取最近 10 个交易日（行业和概念取并集）
    hy_dates = get_last_10_dates(conn, '行业')
    gn_dates = get_last_10_dates(conn, '概念')
    all_dates = sorted(set(hy_dates) | set(gn_dates), reverse=True)[:10]
    print(f'[OK] Last 10 dates: {all_dates[0]} to {all_dates[-1]}')

    daily_top10 = build_daily_top10(conn, all_dates)
    print(f'[OK] Built daily Top10: {len(daily_top10)} days')

    gen_excel(conn, daily_top10, Path(args.out))

    if args.compare:
        compare_with_original(conn, daily_top10, Path(args.compare_orig))

    conn.close()
    print('\n[DONE]')


if __name__ == '__main__':
    main()
