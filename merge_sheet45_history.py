# -*- coding: utf-8 -*-
"""合并备份中的历史Sheet4/5数据到当前v4"""
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

bak_path = 'data/板块轮动Top10_v4_含非Top3强势个股.bak.20260825_194537.xlsx'
cur_path = 'data/板块轮动Top10_v4_含非Top3强势个股.xlsx'

wb_bak = openpyxl.load_workbook(bak_path, data_only=True)
wb_cur = openpyxl.load_workbook(cur_path)

# ============ Sheet4: 每日Top3强势个股 ============
print('处理 Sheet4...')
ws4_src = wb_bak['每日Top3强势个股']
ws4_dst = wb_cur['每日Top3强势个股']

# 读取备份全部数据行（排除表头）
src_rows_s4 = []
for row in ws4_src.iter_rows(values_only=True):
    if row[0] and '2026' in str(row[0]) and str(row[0]) != '日期':
        src_rows_s4.append(row)

print(f'  备份 Sheet4: {len(src_rows_s4)} 行历史数据')

# 找当前v4最后一行
max_row_s4 = ws4_dst.max_row
print(f'  当前 Sheet4: {max_row_s4} 行')

# 追加历史数据（但跳过 2026-08-25，因为今天已经写入了）
today_date = '2026-08-25'
count_added_s4 = 0
for row_data in src_rows_s4:
    date_val = str(row_data[0])[:10]
    if date_val == today_date:
        continue  # 跳过今日（已存在）
    max_row_s4 += 1
    for col_i, val in enumerate(row_data, 1):
        c = ws4_dst.cell(row=max_row_s4, column=col_i, value=val)
    count_added_s4 += 1

print(f'  新增 {count_added_s4} 行历史数据到 Sheet4')

# ============ Sheet5: 非Top3板块强势个股 ============
print('处理 Sheet5...')
ws5_src = wb_bak['非Top3板块强势个股']
ws5_dst = wb_cur['非Top3板块强势个股']

src_rows_s5 = []
for row in ws5_src.iter_rows(values_only=True):
    if row[0] and '2026' in str(row[0]) and str(row[0]) != '日期':
        src_rows_s5.append(row)

print(f'  备份 Sheet5: {len(src_rows_s5)} 行历史数据')

max_row_s5 = ws5_dst.max_row
print(f'  当前 Sheet5: {max_row_s5} 行')

count_added_s5 = 0
for row_data in src_rows_s5:
    date_val = str(row_data[0])[:10]
    if date_val == today_date:
        continue
    max_row_s5 += 1
    for col_i, val in enumerate(row_data, 1):
        c = ws5_dst.cell(row=max_row_s5, column=col_i, value=val)
    count_added_s5 += 1

print(f'  新增 {count_added_s5} 行历史数据到 Sheet5')

wb_bak.close()
wb_cur.save(cur_path)
print(f'\n已保存: {cur_path}')
print(f'Sheet4: 08-11~08-24 + 今日(08-25) 合计 {count_added_s4 + ws4_dst.max_row - 1} 条（含表头)')
print(f'Sheet5: 08-11~08-24 + 今日(08-25) 合计 {count_added_s5 + ws5_dst.max_row - 1} 条（含表头）')
