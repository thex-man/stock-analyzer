# -*- coding: utf-8 -*-
"""
Sheet6: 创业板 MACD>0 且近5日涨幅>10% 的股票
用缠论结构打分排序
"""
import baostock as bs
import pandas as pd
import numpy as np
import os, json, time

# ===== 1. 从 concept_data 获取创业板股票列表 =====
concept_dir = r'D:\stock\tool\stock\concept_data'
files = sorted(os.listdir(concept_dir))
cy_files = [f for f in files if f.startswith('30') and f.endswith('_concepts.json')]
cy_codes = [f.replace('_concepts.json', '') for f in cy_files]
print(f'创业板股票: {len(cy_codes)} 只')

end_date = '2026-08-21'
# MACD需要至少34根K线(约2个月)
start_date = '2026-06-01'

def get_kline(code6, bs_session):
    """获取单只股票K线，返回DataFrame或None"""
    try:
        rs = bs_session.query_history_k_data_plus(
            'sz.' + code6,
            'date,open,high,low,close,volume,pctChg',
            start_date=start_date, end_date=end_date,
            frequency='d', adjustflag='2')
        data = []
        while rs.error_code == '0' and rs.next():
            data.append(rs.get_row_data())
        if len(data) < 34:
            return None
        df = pd.DataFrame(data, columns=['date','open','high','low','close','volume','pctChg'])
        df[['open','high','low','close','volume','pctChg']] = df[['open','high','low','close','volume','pctChg']].astype(float)
        return df
    except Exception:
        return None

def calc_macd(closes):
    """计算MACD指标"""
    ema12 = pd.Series(closes).ewm(span=12, adjust=False).mean()
    ema26 = pd.Series(closes).ewm(span=26, adjust=False).mean()
    macd = ema12 - ema26
    signal = pd.Series(macd).ewm(span=9, adjust=False).mean()
    return float(macd.iloc[-1])

def chan_score(df):
    """
    缠论结构打分（简化版）
    - 上升趋势: +2
    - 底分型: +1.5
    - 中枢上方: +2，强势离开: +1
    - 底背驰: +1.5
    - 放量配合: +0.5
    - 顶分型/顶背驰: 扣分
    """
    closes = df['close'].values.astype(float)
    highs = df['high'].values.astype(float)
    lows = df['low'].values.astype(float)
    volumes = df['volume'].values.astype(float)
    n = len(closes)
    score = 0.0
    details = {}

    # --- 趋势（20日斜率）---
    if n >= 20:
        slope = (closes[-1] / closes[-20] - 1) * 100
        if slope > 8:
            score += 2; trend = f'强势上升({slope:+.1f}%)'
        elif slope > 3:
            score += 1; trend = f'缓慢上升({slope:+.1f}%)'
        elif slope < -8:
            score -= 3; trend = f'下降({slope:+.1f}%)'
        elif slope < -3:
            score -= 1; trend = f'缓慢下降({slope:+.1f}%)'
        else:
            trend = f'横盘({slope:+.1f}%)'
        details['trend'] = trend

    # --- 中枢（最近20日）---
    look = min(20, n)
    zg = float(highs[-look:].max())
    zd = float(lows[-look:].min())
    z_width = zg - zd if zg > zd else 1
    current = float(closes[-1])

    if current > zg:
        score += 2
        position = '中枢上方强势'
        leave = (current - zg) / z_width
        if leave > 1.5: score += 1
    elif current < zd:
        score -= 2; position = '中枢下方弱势'
    else:
        position = '中枢内部震荡'

    details['zg'] = round(zg, 2)
    details['zd'] = round(zd, 2)
    details['position'] = position

    # --- 分型 ---
    if n >= 5:
        c = closes
        # 简化底分型判断
        is_bottom = (c[-3] < c[-2] and c[-3] < c[-4] and c[-2] > c[-1] and c[-2] > c[-5])
        is_top = (c[-3] > c[-2] and c[-3] > c[-4] and c[-2] < c[-1] and c[-2] < c[-5])
        if is_bottom:
            score += 1.5; details['fx'] = '底分型'
        elif is_top:
            score -= 1; details['fx'] = '顶分型'
        else:
            details['fx'] = '无分型'

    # --- MACD背驰 ---
    try:
        ema12_s = pd.Series(closes).ewm(span=12, adjust=False).mean()
        ema26_s = pd.Series(closes).ewm(span=26, adjust=False).mean()
        macd_s = ema12_s - ema26_s
        sig_s = pd.Series(macd_s).ewm(span=9, adjust=False).mean()
        hist_s = macd_s - sig_s
        recent_area = float(hist_s.iloc[-5:].sum())
        prev_area = float(hist_s.iloc[-10:-5].sum()) if len(hist_s) >= 10 else 0
        if recent_area > 0 and prev_area > 0 and recent_area > prev_area * 1.3:
            score += 1.5; details['bcie'] = '底背驰(+1.5)'
        elif recent_area < 0 and prev_area < 0 and abs(recent_area) > abs(prev_area) * 1.3:
            score -= 1.5; details['bcie'] = '顶背驰(-1.5)'
        else:
            details['bcie'] = '无背驰'
    except:
        details['bcie'] = 'N/A'

    # --- 成交量 ---
    if n >= 20:
        vol5 = float(volumes[-5:].mean())
        vol20 = float(volumes[-20:].mean())
        if vol5 > vol20 * 1.3 and score > 0:
            score += 0.5; details['vol'] = '放量配合'
        elif vol5 < vol20 * 0.7:
            details['vol'] = '缩量'
        else:
            details['vol'] = '量能正常'

    details['total_score'] = round(score, 1)
    return score, details

# ===== 2. 遍历筛选 =====
bs.login()
bs_session = bs
results = []
total = len(cy_codes)

for idx, code6 in enumerate(cy_codes):
    # 每200只重连一次，避免超时
    if idx > 0 and idx % 200 == 0:
        try: bs_session.logout()
        except: pass
        bs.login()
        bs_session = bs
    df = get_kline(code6, bs_session)
    if df is None:
        if (idx+1) % 200 == 0:
            print(f'  进度 {idx+1}/{total} (skip {total-idx-1} remaining)')
        continue

    closes = df['close'].values
    if len(closes) < 6:
        continue

    try:
        macd_val = calc_macd(closes)
        gain_5d = (closes[-1] / closes[-6] - 1) * 100
    except:
        continue

    if macd_val > 0 and gain_5d > 10:
        # 读名称
        concept_file = os.path.join(concept_dir, f'{code6}_concepts.json')
        name = code6
        try:
            with open(concept_file, 'r', encoding='utf-8') as f:
                cdata = json.load(f)
                name = cdata.get('stock_name') or cdata.get('name') or code6
        except:
            pass

        results.append({
            'code': code6,
            'name': name,
            'close': round(float(closes[-1]), 2),
            'macd': round(macd_val, 4),
            'gain_5d': round(gain_5d, 2),
            'df': df,
        })

    if (idx + 1) % 200 == 0:
        print(f'  进度 {idx+1}/{total}, 已筛出{len(results)}只符合条件的')

bs.logout()
print(f'\n符合 MACD>0 且5日涨幅>10%: {len(results)} 只')

# ===== 3. 缠论打分 =====
print('\n===== 缠论分析中 =====')
final_results = []
for i, r in enumerate(results):
    score, details = chan_score(r['df'])
    r['score'] = round(score, 1)
    r['details'] = details
    del r['df']
    final_results.append(r)
    if (i+1) % 20 == 0:
        print(f'  已分析 {i+1}/{len(results)}')

final_results.sort(key=lambda x: -x['score'])

print('\n===== 最终排名 =====')
for i, r in enumerate(final_results):
    d = r['details']
    print(f"{i+1:2d}. {r['code']} {r['name']:8s} 分数:{r['score']:+.1f} "
          f"| 5日涨幅:{r['gain_5d']:+.1f}% | {d.get('position','')} "
          f"| {d.get('trend','')} | {d.get('fx','')} | {d.get('bcie','')}")

# ===== 4. 写入 Excel Sheet6 =====
import openpyxl
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side

wb = load_workbook(r'D:\stock\tool\stock\data\板块轮动Top10_v4_含非Top3强势个股.xlsx')
ws6_name = 'MACD强势个股'
if ws6_name in wb.sheetnames:
    del wb[ws6_name]

ws6 = wb.create_sheet(ws6_name)

thin = Side(style='thin', color='CCCCCC')
border = Border(left=thin, right=thin, top=thin, bottom=thin)

headers = ['排名', '代码', '名称', '最新价', 'MACD', '5日涨幅%', '缠论分数',
           '位置', '趋势', '分型', '背驰', '量能', '备注']
col_widths = [6, 10, 14, 10, 10, 12, 10, 18, 22, 12, 14, 12, 34]

for i, (h, w) in enumerate(zip(headers, col_widths), 1):
    ws6.column_dimensions[get_column_letter(i)].width = w
    c = ws6.cell(row=1, column=i, value=h)
    c.font = Font(bold=True, color='FFFFFF', size=11)
    c.fill = PatternFill(start_color='2F5496', end_color='2F5496', fill_type='solid')
    c.alignment = Alignment(horizontal='center', vertical='center')
    c.border = border

ws6.freeze_panes = 'A2'

RANK_COLORS = ['C0392B', 'E74C3C', 'E67E22', 'F39C12', '27AE60',
               '2ECC71', '3498DB', '8E44AD', '16A085', '7F8C8D']

for row_i, r in enumerate(final_results, 2):
    d = r['details']
    rank = row_i - 2
    rank_color = RANK_COLORS[min(rank - 1, len(RANK_COLORS) - 1)]
    score_val = r['score']

    row_data = [
        rank,
        r['code'],
        r['name'],
        r['close'],
        r['macd'],
        r['gain_5d'],
        score_val,
        d.get('position', ''),
        d.get('trend', ''),
        d.get('fx', ''),
        d.get('bcie', ''),
        d.get('vol', ''),
        f"ZG:{d.get('zg','')} ZD:{d.get('zd','')}",
    ]
    for col_i, val in enumerate(row_data, 1):
        c = ws6.cell(row=row_i, column=col_i, value=val)
        c.border = border
        c.alignment = Alignment(horizontal='center', vertical='center')
        if col_i == 1:
            c.fill = PatternFill(start_color=rank_color, end_color=rank_color, fill_type='solid')
            c.font = Font(bold=True, color='FFFFFF')
        elif col_i == 7:
            if score_val >= 3:
                c.fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
                c.font = Font(bold=True, color='006100')
            elif score_val >= 1:
                c.fill = PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid')
                c.font = Font(bold=True, color='9C5700')
            elif score_val <= -1:
                c.fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
                c.font = Font(bold=True, color='9C0006')
        elif col_i == 6:
            if float(val) > 20: c.font = Font(bold=True, color='FF0000')
            elif float(val) > 15: c.font = Font(bold=True, color='FF4500')

wb.save(r'D:\stock\tool\stock\data\板块轮动Top10_v4_含非Top3强势个股.xlsx')
print(f'\nSheet6 已写入，共 {len(final_results)} 只')
