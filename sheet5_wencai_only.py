"""
Sheet5: 每日涨幅>6%但不在当日Top3板块内的个股
==============================================
方案：问财直接查每天涨幅>6%的股票（每天1次请求，10天=10次）
  - 每天问: "2026年8月7日涨幅超过6%的股票"
  - 返回结果就是所有>6%的个股
  - 排除当天Top3板块内的股票
  - 剩余的就是"独立强势个股"
"""
import sys
sys.path.insert(0, '.')
from stock_data_source import wencai
import json
from pathlib import Path
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import time

# ============ 1. 读取板块历史 ============
cache_file = Path('data/board_history_ths/history_20260728_20260820.json')
with open(cache_file, encoding='utf-8') as f:
    raw = json.load(f)

all_dates = set()
for sym, info in raw.items():
    for row in info.get('data', []):
        if row.get('d'):
            all_dates.add(row['d'])

dates_sorted = sorted(all_dates, reverse=True)
last_10 = dates_sorted[:10]

# 每天Top3板块
daily_top3 = {}
for date in last_10:
    day_results = []
    for sym, info in raw.items():
        name = info.get('name', sym)
        for row in info.get('data', []):
            if row.get('d') == date:
                pct = row.get('p', 0)
                day_results.append((pct, name))
                break
    day_results.sort(reverse=True)
    daily_top3[date] = {name for _, name in day_results[:3]}

all_boards = {}
for sym, info in raw.items():
    all_boards[info.get('name', sym)] = info.get('type', '')

def fmt_date(d):
    return f"{d[:4]}-{d[4:6]}-{d[6:]}"

# ============ 2. 每天查询涨幅>6%的个股 ============
print("查询每天涨幅>6%的个股...")
daily_other_stocks = {date: [] for date in last_10}

for date in last_10:
    ds = fmt_date(date)
    # 问财查询语句
    query = f"{ds}涨幅超过6%的股票"
    print(f"  {ds}: 查询中...", end='', flush=True)
    
    for attempt in range(5):
        try:
            result = wencai(query)
            df = result.get('datas')
            if df is not None and not df.empty:
                break
            time.sleep(2)
        except Exception as e:
            print(f"\n  错误: {e}")
            time.sleep(3)
    
    if df is None or df.empty:
        print(f" 无数据")
        continue
    
    # 获取该日期的板块涨跌幅数据（用于找个股所属板块）
    board_pct_map = {}
    for sym, info in raw.items():
        for row in info.get('data', []):
            if row.get('d') == date:
                board_pct_map[info.get('name', sym)] = row.get('p', 0)
                break
    
    count = 0
    for _, row in df.iterrows():
        try:
            pct = row.get('涨跌幅:前复权') or row.get('最新涨跌幅')
            if pct is None:
                continue
            pct = float(pct)
            if pct <= 6.0:
                continue
            code = str(row.get('股票代码', '')).strip()
            name = str(row.get('股票简称', '')).strip()
            if not code or not name:
                continue
            
            # 找这个股属于哪个板块（从问财返回中通常有板块字段）
            board = str(row.get('所属板块', '')).strip()
            if not board or board in ['-', '']:
                board = '其他'
            
            # 判断是否在Top3板块
            # 注意：这里用板块名称匹配，如果board名不完全匹配则需要额外处理
            in_top3 = False
            for top3_board in daily_top3.get(date, set()):
                if top3_board in board or board in top3_board:
                    in_top3 = True
                    break
            
            if not in_top3:
                board_pct = board_pct_map.get(board, 0)
                daily_other_stocks[date].append({
                    'date': date,
                    'stock_name': name,
                    'code': code,
                    'change': round(pct, 2),
                    'board': board,
                    'board_type': all_boards.get(board, '概念'),
                    'board_pct': board_pct,
                })
                count += 1
        except (ValueError, TypeError):
            continue
    
    print(f" {count} 只非Top3板块>6%")

# 排序
for date in last_10:
    daily_other_stocks[date].sort(key=lambda x: -x['change'])

# ============ 3. 写入Excel ============
print("\n写入Excel...")
out_path = 'data/板块轮动Top10_v4_含非Top3强势个股.xlsx'
wb = load_workbook('data/板块轮动Top10_v3_含每日Top3强势个股.xlsx')

if '非Top3板块强势个股' in wb.sheetnames:
    del wb['非Top3板块强势个股']

ws = wb.create_sheet('非Top3板块强势个股')

DAY_COLORS = [
    'D6E4F0', 'EBF5FB', 'D5F5E3', 'FEF9E7', 'FDEDEC',
    'F4ECF7', 'E8F8F5', 'FDF2E9', 'F0F3FF', 'F9EBEA',
]

thin = Side(style='thin', color='CCCCCC')
border = Border(left=thin, right=thin, top=thin, bottom=thin)

col_widths = [14, 12, 16, 18, 10, 12, 12]
col_headers = ['日期', '个股代码', '个股简称', '所属板块', '板块类型', '板块涨幅', '个股涨幅']
for i, (w, h) in enumerate(zip(col_widths, col_headers), 1):
    ws.column_dimensions[get_column_letter(i)].width = w
    c = ws.cell(row=1, column=i, value=h)
    c.font = Font(bold=True, color='FFFFFF', size=11)
    c.fill = PatternFill(start_color='2F5496', end_color='2F5496', fill_type='solid')
    c.alignment = Alignment(horizontal='center', vertical='center')
    c.border = border

ws.freeze_panes = 'A2'

row_num = 2
total_stocks = 0

for day_idx, date in enumerate(reversed(last_10)):
    day_color = DAY_COLORS[day_idx % len(DAY_COLORS)]
    items = daily_other_stocks[date]
    date_start_row = row_num

    if items:
        first = items[0]
        for col_i, val in enumerate([fmt_date(date), first['code'], first['stock_name'],
                                      first['board'], first['board_type'],
                                      f"{first['board_pct']:+.2f}%", f"{first['change']:+.2f}%"], 1):
            c = ws.cell(row=row_num, column=col_i, value=val)
            c.fill = PatternFill(start_color=day_color, end_color=day_color, fill_type='solid')
            c.border = border
            c.alignment = Alignment(horizontal='center', vertical='center')
            if col_i == 6:
                c.alignment = Alignment(horizontal='right', vertical='center')
                c.font = Font(bold=True, color='C00000')
            elif col_i == 7:
                c.alignment = Alignment(horizontal='right', vertical='center')
                c.font = Font(bold=True, color='C00000')
        ws.cell(row=row_num, column=1).font = Font(bold=True)
        row_num += 1

        for item in items[1:]:
            ws.cell(row=row_num, column=1, value='')
            for col_i, val in enumerate(['', item['code'], item['stock_name'], item['board'],
                                          item['board_type'],
                                          f"{item['board_pct']:+.2f}%", f"{item['change']:+.2f}%"], 1):
                c = ws.cell(row=row_num, column=col_i, value=val)
                c.fill = PatternFill(start_color=day_color, end_color=day_color, fill_type='solid')
                c.border = border
                c.alignment = Alignment(horizontal='center', vertical='center')
                if col_i == 6:
                    c.alignment = Alignment(horizontal='right', vertical='center')
                    c.font = Font(bold=True, color='C00000')
                elif col_i == 7:
                    c.alignment = Alignment(horizontal='right', vertical='center')
                    c.font = Font(bold=True, color='C00000')
            row_num += 1
    else:
        c_date = ws.cell(row=row_num, column=1, value=fmt_date(date))
        c_date.font = Font(bold=True)
        c_note = ws.cell(row=row_num, column=3, value='无>6%个股（不在Top3板块）')
        for col_i in range(1, 8):
            ws.cell(row=row_num, column=col_i).fill = PatternFill(start_color=day_color, end_color=day_color, fill_type='solid')
            ws.cell(row=row_num, column=col_i).border = border
        c_date.alignment = Alignment(horizontal='center', vertical='center')
        row_num += 1

    if date_start_row < row_num - 1:
        ws.merge_cells(start_row=date_start_row, start_column=1, end_row=row_num - 1, end_column=1)
        ws.cell(row=date_start_row, column=1).alignment = Alignment(horizontal='center', vertical='center')
        ws.cell(row=date_start_row, column=1).font = Font(bold=True)

    total_stocks += len(items)
    row_num += 1

wb.save(out_path)
print(f"\n已保存: {out_path}")
print(f"Sheet5 共 {total_stocks} 条")
