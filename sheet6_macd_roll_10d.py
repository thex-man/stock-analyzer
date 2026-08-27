# -*- coding: utf-8 -*-
"""
【已废弃】2026-08-27 移除

Sheet6/Sheet7 v2.0 及其衍生 sheets（信号消失追踪/概念聚合/上榜频次/回测/_chart_data）已被移除。
v4 Excel 现在只保留 8 个 sheet，不再调用本脚本。

如需重新启用，请参考 git 历史恢复。
"""
import sys
print('[DEPRECATED] sheet6_macd_roll_10d.py is deprecated as of 2026-08-27.', file=sys.stderr)
print('[DEPRECATED] MACD v2 衍生 sheets removed from v4 Excel.', file=sys.stderr)
print('[DEPRECATED] See git history if you need to restore this script.', file=sys.stderr)
sys.exit(0)

import os, sys, json, time, subprocess, requests
import pandas as pd
import numpy as np
import akshare as ak
from datetime import datetime

# ====== 配置 ======
CACHE_DIR = r'D:\stock\tool\stock\kline_cache'
EXCEL_FILE = r'D:\stock\tool\stock\data\板块轮动Top10_v4_含非Top3强势个股.xlsx'
START_DATE = '2026-08-01'
TODAY = '2026-08-24'

WC_DIR = r'C:\Users\s5631\AppData\Local\Programs\Python\Python313\Lib\site-packages\pywencai'


# ============ 0. 加载 Excel 工作簿（提前，供所有 sheet 使用） ============
import openpyxl
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.chart import LineChart, Reference

wb = load_workbook(EXCEL_FILE)


# ============ 1. 获取创业板股票名单 ============
print('===== Step1: 拉创业板股票名单（akshare）=====')
df_all = ak.stock_info_a_code_name()
cy_df = df_all[df_all['code'].str.startswith('30')].reset_index(drop=True)
stocks = [{'code': str(r['code']).strip(), 'name': str(r['name']).strip()}
          for _, r in cy_df.iterrows()]
print(f'创业板股票: {len(stocks)} 只')


# ============ 2. 拉 2026 年交易日历，过滤出截面日 ============
print(f'\n===== Step2: 拉交易日历（akshare）=====')
df_cal = ak.tool_trade_date_hist_sina()
df_cal['trade_date'] = pd.to_datetime(df_cal['trade_date']).dt.strftime('%Y-%m-%d')
trade_days = sorted([d for d in df_cal['trade_date']
                     if START_DATE <= d <= TODAY])
print(f'{START_DATE} ~ {TODAY} 共 {len(trade_days)} 个交易日:')
print('  ', trade_days)


# ============ 3. 读取所有 K 线缓存 ============
print(f'\n===== Step3: 读取 K 线缓存 =====')
all_klines = {}
miss_codes = []
for s in stocks:
    cache_file = os.path.join(CACHE_DIR, f'{s["code"]}_qfq.csv')
    if not os.path.exists(cache_file):
        miss_codes.append(s['code'])
        continue
    try:
        df = pd.read_csv(cache_file)
        df['date'] = pd.to_datetime(df['date']).dt.strftime('%Y-%m-%d')
        if not df.empty:
            all_klines[s['code']] = df
    except Exception as e:
        miss_codes.append(s['code'])
print(f'缓存命中: {len(all_klines)}, 缺失: {len(miss_codes)}')
if miss_codes:
    print(f'  缺失代码示例: {miss_codes[:5]}')


# ============ 4. 计算函数 ============
def calc_macd_at(df: pd.DataFrame, signal_date: str) -> float:
    df_until = df[df['date'] <= signal_date].reset_index(drop=True)
    closes = df_until['close'].values.astype(float)
    if len(closes) < 26:
        return None
    ema12 = pd.Series(closes).ewm(span=12, adjust=False).mean()
    ema26 = pd.Series(closes).ewm(span=26, adjust=False).mean()
    return float((ema12 - ema26).iloc[-1])


def calc_gain_nd_at(df: pd.DataFrame, signal_date: str, n: int) -> float:
    df_until = df[df['date'] <= signal_date].reset_index(drop=True)
    if len(df_until) < n + 1:
        return None
    close_now = float(df_until['close'].iloc[-1])
    close_n = float(df_until['close'].iloc[-(n + 1)])
    return (close_now / close_n - 1) * 100


def calc_gain_since(df: pd.DataFrame, signal_date: str, today: str = TODAY) -> float:
    df_until = df[df['date'] <= today].reset_index(drop=True)
    row_signal = df_until[df_until['date'] == signal_date]
    if row_signal.empty:
        return None
    close_signal = float(row_signal['close'].iloc[0])
    close_today = float(df_until['close'].iloc[-1])
    return (close_today / close_signal - 1) * 100


def chan_score(df: pd.DataFrame) -> tuple:
    """与原 sheet6 一致的缠论打分（用最新 K 线）"""
    closes = df['close'].values.astype(float)
    highs = df['high'].values.astype(float)
    lows = df['low'].values.astype(float)
    volumes = df['volume'].values.astype(float) if 'volume' in df.columns else np.zeros(len(closes))
    n = len(closes)
    score = 0.0
    details = {}

    if n >= 20:
        slope = (closes[-1] / closes[-20] - 1) * 100
        if slope > 8:
            score += 2; details['trend'] = f'强势上升({slope:+.1f}%)'
        elif slope > 3:
            score += 1; details['trend'] = f'缓慢上升({slope:+.1f}%)'
        elif slope < -8:
            score -= 3; details['trend'] = f'下降({slope:+.1f}%)'
        elif slope < -3:
            score -= 1; details['trend'] = f'缓慢下降({slope:+.1f}%)'
        else:
            details['trend'] = f'横盘({slope:+.1f}%)'

    look = min(20, n)
    zg = float(highs[-look:].max())
    zd = float(lows[-look:].min())
    z_width = zg - zd if zg > zd else 1
    current = float(closes[-1])
    if current > zg:
        score += 2; details['position'] = '中枢上方强势'
        if (current - zg) / z_width > 1.5: score += 1
    elif current < zd:
        score -= 2; details['position'] = '中枢下方弱势'
    else:
        details['position'] = '中枢内部震荡'
    details['zg'] = round(zg, 2); details['zd'] = round(zd, 2)

    if n >= 5:
        c = closes
        is_bottom = (c[-3] < c[-2] and c[-3] < c[-4] and c[-2] > c[-1] and c[-2] > c[-5])
        is_top = (c[-3] > c[-2] and c[-3] > c[-4] and c[-2] < c[-1] and c[-2] < c[-5])
        if is_bottom:
            score += 1.5; details['fx'] = '底分型'
        elif is_top:
            score -= 1; details['fx'] = '顶分型'
        else:
            details['fx'] = '无分型'

    try:
        ema12_s = pd.Series(closes).ewm(span=12, adjust=False).mean()
        ema26_s = pd.Series(closes).ewm(span=26, adjust=False).mean()
        macd_s = ema12_s - ema26_s
        sig_s = pd.Series(macd_s).ewm(span=9, adjust=False).mean()
        hist_s = macd_s - sig_s
        recent_area = float(hist_s.iloc[-5:].sum())
        prev_area = float(hist_s.iloc[-10:-5].sum()) if len(hist_s) >= 10 else 0
        if recent_area > 0 and prev_area > 0 and recent_area > prev_area * 1.3:
            score += 1.5; details['bcie'] = '底背驰(+1.5)'
        elif recent_area < 0 and prev_area < 0 and abs(recent_area) > abs(prev_area) * 1.3:
            score -= 1.5; details['bcie'] = '顶背驰(-1.5)'
        else:
            details['bcie'] = '无背驰'
    except:
        details['bcie'] = 'N/A'

    if n >= 20 and volumes.sum() > 0:
        vol5 = float(volumes[-5:].mean())
        vol20 = float(volumes[-20:].mean())
        if vol5 > vol20 * 1.3 and score > 0:
            score += 0.5; details['vol'] = '放量配合'
        elif vol5 < vol20 * 0.7:
            details['vol'] = '缩量'
        else:
            details['vol'] = '量能正常'

    details['total_score'] = round(score, 1)
    return score, details


# ============ 5. 每天遍历筛选（核心循环） ============
print(f'\n===== Step4: 滚动筛选 {len(trade_days)} 个交易日 =====')
sheet6_results = []   # MACD>0 + 5日涨幅>10%
sheet7_results = []   # MACD>0 + 10日涨幅>20%

t0 = time.time()
for day_idx, signal_date in enumerate(trade_days):
    day_s6 = []
    day_s7 = []
    for s in stocks:
        code6 = s['code']
        if code6 not in all_klines:
            continue
        df_k = all_klines[code6]
        try:
            macd_val = calc_macd_at(df_k, signal_date)
            gain_5d = calc_gain_nd_at(df_k, signal_date, 5)
            gain_10d = calc_gain_nd_at(df_k, signal_date, 10)
            if macd_val is None or gain_5d is None or gain_10d is None:
                continue

            # 取最新 K 线（用于缠论打分）
            df_latest = df_k[df_k['date'] <= TODAY].reset_index(drop=True)
            if len(df_latest) < 30:
                continue

            # Sheet6
            if macd_val > 0 and gain_5d > 10:
                score, details = chan_score(df_latest)
                gain_since = calc_gain_since(df_k, signal_date, TODAY)
                rec = {
                    'signal_date': signal_date,
                    'code': code6,
                    'name': s['name'],
                    'close': round(float(df_latest['close'].iloc[-1]), 2),
                    'macd': round(macd_val, 4),
                    'gain_nd': round(gain_5d, 2),
                    'gain_since': round(gain_since, 2) if gain_since is not None else None,
                    'score': round(score, 1),
                    'details': details,
                }
                day_s6.append(rec)
                sheet6_results.append(rec)

            # Sheet7
            if macd_val > 0 and gain_10d > 20:
                score, details = chan_score(df_latest)
                gain_since = calc_gain_since(df_k, signal_date, TODAY)
                rec = {
                    'signal_date': signal_date,
                    'code': code6,
                    'name': s['name'],
                    'close': round(float(df_latest['close'].iloc[-1]), 2),
                    'macd': round(macd_val, 4),
                    'gain_nd': round(gain_10d, 2),
                    'gain_since': round(gain_since, 2) if gain_since is not None else None,
                    'score': round(score, 1),
                    'details': details,
                }
                day_s7.append(rec)
                sheet7_results.append(rec)
        except Exception:
            pass

    print(f'  [{day_idx+1}/{len(trade_days)}] {signal_date}: '
          f'Sheet6 {len(day_s6)}只, Sheet7 {len(day_s7)}只, '
          f'累计 {time.time()-t0:.1f}s')

# Sheet6 按"日期降序 + 分数降序"排序
sheet6_results.sort(key=lambda x: (x['signal_date'], x['score']), reverse=True)
sheet7_results.sort(key=lambda x: (x['signal_date'], x['score']), reverse=True)

print(f'\n===== 汇总 =====')
print(f'Sheet6: {len(sheet6_results)} 行 ({len(trade_days)} 天 × 平均 {len(sheet6_results)/len(trade_days):.1f} 只/天)')
print(f'Sheet7: {len(sheet7_results)} 行 ({len(trade_days)} 天 × 平均 {len(sheet7_results)/len(trade_days):.1f} 只/天)')

# 每日计数
from collections import Counter
s6_by_day = Counter(r['signal_date'] for r in sheet6_results)
s7_by_day = Counter(r['signal_date'] for r in sheet7_results)
print('\n每日 Sheet6 计数:')
for d in sorted(s6_by_day.keys(), reverse=True):
    print(f'  {d}: {s6_by_day[d]} 只')
print('\n每日 Sheet7 计数:')
for d in sorted(s7_by_day.keys(), reverse=True):
    print(f'  {d}: {s7_by_day[d]} 只')


# ============ 7. D. 信号消失追踪 ============
def write_disappear_sheet(wb, sheet_name, results, all_klines, today=TODAY):
    """每只票的“掉出榜单”追踪（以今日为观测点）"""
    if sheet_name in wb.sheetnames:
        del wb[sheet_name]
    ws = wb.create_sheet(sheet_name)
    thin = Side(style='thin', color='CCCCCC')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    from collections import defaultdict
    by_code = defaultdict(list)
    for r in results:
        by_code[r['code']].append(r)

    rows = []
    for code6, recs in by_code.items():
        recs.sort(key=lambda x: x['signal_date'])
        days = sorted(set(r['signal_date'] for r in recs))
        if not days:
            continue
        last_day = days[-1]
        last_rec = [r for r in recs if r['signal_date'] == last_day][0]
        last_score = last_rec['score']
        # 掉出后 N 天最大涨幅 / 最大回撤（以 last_day 收盘价为基准，到 today）
        df_k = all_klines.get(code6)
        max_gain_since = None
        max_dd_since = None
        if df_k is not None:
            df_after = df_k[(df_k['date'] >= last_day) & (df_k['date'] <= today)].reset_index(drop=True)
            if len(df_after) >= 2:
                base_close = float(df_after['close'].iloc[0])
                max_high = float(df_after['high'].max())
                min_low = float(df_after['low'].min())
                last_close = float(df_after['close'].iloc[-1])
                max_gain_since = round((max_high / base_close - 1) * 100, 2)
                max_dd_since = round((min_low / base_close - 1) * 100, 2)
                gain_to_today = round((last_close / base_close - 1) * 100, 2)
            else:
                gain_to_today = None
        else:
            gain_to_today = None
        # 上榜连续性：连续天数 / 总间隔天数
        continuity = 0
        max_streak = 1
        streak = 1
        for i in range(1, len(days)):
            d_prev = datetime.strptime(days[i-1], '%Y-%m-%d')
            d_cur = datetime.strptime(days[i], '%Y-%m-%d')
            if (d_cur - d_prev).days <= 3:  # 考虑周末
                streak += 1
                max_streak = max(max_streak, streak)
            else:
                streak = 1
        rows.append({
            'code': code6,
            'name': recs[0]['name'],
            'cnt': len(days),
            'max_streak': max_streak,
            'last_day': last_day,
            'last_score': last_score,
            'max_gain': max_gain_since,
            'max_dd': max_dd_since,
            'gain_to_today': gain_to_today,
        })

    # 按掉出后表现排序：先按今天是否创新高排序
    rows.sort(key=lambda x: (
        -(x['max_gain'] or -999),
        -(x['gain_to_today'] or -999),
    ))

    headers = ['代码', '名称', '上榜天数', '最长连续',
               '末次上榜日', '末次上榜分数', '掉出后最大涨幅%', '掉出后最大回撤%', '末次上榜日→今日%']
    col_widths = [10, 14, 10, 12, 12, 14, 16, 16, 18]

    for i, (h, w) in enumerate(zip(headers, col_widths), 1):
        ws.column_dimensions[get_column_letter(i)].width = w
        c = ws.cell(row=1, column=i, value=h)
        c.font = Font(bold=True, color='FFFFFF', size=11)
        c.fill = PatternFill(start_color='E67E22', end_color='E67E22', fill_type='solid')
        c.alignment = Alignment(horizontal='center', vertical='center')
        c.border = border
    ws.freeze_panes = 'A2'
    ws.auto_filter.ref = ws.dimensions

    for row_i, r in enumerate(rows, 2):
        row_data = [
            r['code'], r['name'], r['cnt'], r['max_streak'],
            r['last_day'], r['last_score'],
            r['max_gain'] if r['max_gain'] is not None else '',
            r['max_dd'] if r['max_dd'] is not None else '',
            r['gain_to_today'] if r['gain_to_today'] is not None else '',
        ]
        for col_i, val in enumerate(row_data, 1):
            c = ws.cell(row=row_i, column=col_i, value=val)
            c.border = border
            c.alignment = Alignment(horizontal='center', vertical='center')
            if col_i == 7 and isinstance(val, (int, float)):
                if val > 10:
                    c.fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
                    c.font = Font(bold=True, color='006100')
                elif val > 0:
                    c.fill = PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid')
            elif col_i == 8 and isinstance(val, (int, float)) and val < -10:
                c.fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
                c.font = Font(bold=True, color='9C0006')

    return len(rows)


write_disappear_sheet(wb, 'MACD信号消失追踪_5日', sheet6_results, all_klines)
write_disappear_sheet(wb, 'MACD信号消失追踪_10日', sheet7_results, all_klines)


# ============ 8. E. 概念标签聚合 ============
CONCEPT_DIR = r'D:\stock\tool\stock\concept_data'


def write_concept_sheet(wb, sheet_name, results, gain_label):
    """按概念汇总上榜票"""
    if sheet_name in wb.sheetnames:
        del wb[sheet_name]
    ws = wb.create_sheet(sheet_name)
    thin = Side(style='thin', color='CCCCCC')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    from collections import defaultdict
    by_code = defaultdict(list)
    for r in results:
        by_code[r['code']].append(r)

    # code -> [concept_names]
    code_concepts = {}
    for code6 in by_code:
        cf = os.path.join(CONCEPT_DIR, f'{code6}_concepts.json')
        if not os.path.exists(cf):
            code_concepts[code6] = []
            continue
        try:
            with open(cf, 'r', encoding='utf-8') as f:
                cdata = json.load(f)
            code_concepts[code6] = [c['name'] for c in cdata.get('concepts', [])]
        except:
            code_concepts[code6] = []

    # concept -> [stocks_in_concept]
    concept_stocks = defaultdict(set)
    for code6, concepts in code_concepts.items():
        for c in concepts:
            concept_stocks[c].add(code6)

    # 只保留上榜过的票
    rows = []
    for concept, stock_set in concept_stocks.items():
        active_stocks = stock_set & set(by_code.keys())
        if not active_stocks:
            continue
        # 该概念下上榜过的总票数 / 总上榜次数 / 平均缠论分数
        cnt_stocks = len(active_stocks)
        cnt_total = sum(len(set(r['signal_date'] for r in by_code[c])) for c in active_stocks)
        all_scores = [r['score'] for c in active_stocks for r in by_code[c]]
        avg_score = sum(all_scores) / len(all_scores) if all_scores else 0
        all_gains = [r['gain_nd'] for c in active_stocks for r in by_code[c]]
        avg_gain = sum(all_gains) / len(all_gains) if all_gains else 0
        rows.append({
            'concept': concept,
            'cnt_stocks': cnt_stocks,
            'cnt_total': cnt_total,
            'avg_score': round(avg_score, 2),
            'avg_gain': round(avg_gain, 2),
        })

    # 按上榜总次数降序
    rows.sort(key=lambda x: -x['cnt_total'])

    headers = ['排名', '概念名称', f'上榜股票数', f'总上榜次数',
               f'平均{gain_label}涨幅%', '平均缠论分数']
    col_widths = [6, 24, 14, 14, 16, 14]

    for i, (h, w) in enumerate(zip(headers, col_widths), 1):
        ws.column_dimensions[get_column_letter(i)].width = w
        c = ws.cell(row=1, column=i, value=h)
        c.font = Font(bold=True, color='FFFFFF', size=11)
        c.fill = PatternFill(start_color='16A085', end_color='16A085', fill_type='solid')
        c.alignment = Alignment(horizontal='center', vertical='center')
        c.border = border
    ws.freeze_panes = 'A2'
    ws.auto_filter.ref = ws.dimensions

    RANK_COLORS = ['16A085', '1ABC9C', '27AE60', '2ECC71', '3498DB',
                   '2980B9', '8E44AD', '9B59B6', 'F39C12', 'E67E22']
    for row_i, r in enumerate(rows[:50], 2):  # 只输出 TOP 50 概念
        rank = row_i - 2
        rank_color = RANK_COLORS[min(rank, len(RANK_COLORS) - 1)]
        row_data = [
            rank + 1, r['concept'], r['cnt_stocks'], r['cnt_total'],
            r['avg_gain'], r['avg_score']
        ]
        for col_i, val in enumerate(row_data, 1):
            c = ws.cell(row=row_i, column=col_i, value=val)
            c.border = border
            c.alignment = Alignment(horizontal='center', vertical='center')
            if col_i == 1:
                c.fill = PatternFill(start_color=rank_color, end_color=rank_color, fill_type='solid')
                c.font = Font(bold=True, color='FFFFFF')

    return len(rows)


write_concept_sheet(wb, 'MACD概念聚合_5日', sheet6_results, '5日')
write_concept_sheet(wb, 'MACD概念聚合_10日', sheet7_results, '10日')


# ============ 9. F. 历史回测 ============
def write_backtest_sheet(wb, sheet_name, results, all_klines, today=TODAY):
    """简化回测：信号日次日开盘买入，持有 5 个交易日收盘卖出"""
    if sheet_name in wb.sheetnames:
        del wb[sheet_name]
    ws = wb.create_sheet(sheet_name)
    thin = Side(style='thin', color='CCCCCC')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    trade_days_full = sorted(set(r['signal_date'] for r in results))
    bt_rows = []
    for r in results:
        code6 = r['code']
        df_k = all_klines.get(code6)
        if df_k is None:
            continue
        # 找到信号日的下一天，作为买入日
        df_sorted = df_k.sort_values('date').reset_index(drop=True)
        idx_signal = df_sorted[df_sorted['date'] == r['signal_date']].index
        if len(idx_signal) == 0:
            continue
        idx_signal = idx_signal[0]
        if idx_signal + 6 >= len(df_sorted):
            continue  # 不够 5 个交易日
        buy_date = df_sorted.loc[idx_signal + 1, 'date']
        buy_open = float(df_sorted.loc[idx_signal + 1, 'open'])
        sell_date = df_sorted.loc[idx_signal + 5, 'date']
        sell_close = float(df_sorted.loc[idx_signal + 5, 'close'])
        if buy_open <= 0:
            continue
        ret = (sell_close / buy_open - 1) * 100
        # 期间最高 / 最低
        period_high = float(df_sorted.loc[idx_signal+1:idx_signal+5, 'high'].max())
        period_low = float(df_sorted.loc[idx_signal+1:idx_signal+5, 'low'].min())
        max_gain = (period_high / buy_open - 1) * 100
        max_dd = (period_low / buy_open - 1) * 100
        bt_rows.append({
            'signal_date': r['signal_date'],
            'code': code6,
            'name': r['name'],
            'buy_date': buy_date,
            'buy_open': round(buy_open, 2),
            'sell_date': sell_date,
            'sell_close': round(sell_close, 2),
            'ret': round(ret, 2),
            'max_gain': round(max_gain, 2),
            'max_dd': round(max_dd, 2),
        })

    # 写入详情
    headers = ['信号日', '代码', '名称', '买入日', '买入价', '卖出日', '卖出价',
               '收益%', '期间最高%', '期间最低%']
    col_widths = [12, 10, 14, 12, 10, 12, 10, 10, 12, 12]
    for i, (h, w) in enumerate(zip(headers, col_widths), 1):
        ws.column_dimensions[get_column_letter(i)].width = w
        c = ws.cell(row=1, column=i, value=h)
        c.font = Font(bold=True, color='FFFFFF', size=11)
        c.fill = PatternFill(start_color='8E44AD', end_color='8E44AD', fill_type='solid')
        c.alignment = Alignment(horizontal='center', vertical='center')
        c.border = border
    ws.freeze_panes = 'A2'
    ws.auto_filter.ref = ws.dimensions

    # 按信号日降序
    bt_rows.sort(key=lambda x: x['signal_date'], reverse=True)
    for row_i, r in enumerate(bt_rows, 2):
        row_data = [r['signal_date'], r['code'], r['name'],
                    r['buy_date'], r['buy_open'], r['sell_date'], r['sell_close'],
                    r['ret'], r['max_gain'], r['max_dd']]
        for col_i, val in enumerate(row_data, 1):
            c = ws.cell(row=row_i, column=col_i, value=val)
            c.border = border
            c.alignment = Alignment(horizontal='center', vertical='center')
            if col_i == 8:
                if val > 5:
                    c.fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
                    c.font = Font(bold=True, color='006100')
                elif val > 0:
                    c.fill = PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid')
                elif val < -5:
                    c.fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
                    c.font = Font(bold=True, color='9C0006')

    # 写入汇总（在下面 50 行后）
    if bt_rows:
        rets = [r['ret'] for r in bt_rows]
        wins = [r for r in rets if r > 0]
        big_wins = [r for r in rets if r > 5]
        losses = [r for r in rets if r < 0]
        avg_ret = sum(rets) / len(rets)
        win_rate = len(wins) / len(rets) * 100
        max_ret = max(rets)
        min_ret = min(rets)

        summary_row = len(bt_rows) + 4
        ws.cell(row=summary_row, column=1, value='汇总').font = Font(bold=True, size=12)
        summary_data = [
            ('总交易数', len(bt_rows)),
            ('平均收益%', round(avg_ret, 2)),
            ('胜率%', round(win_rate, 2)),
            ('最大收益%', round(max_ret, 2)),
            ('最大亏损%', round(min_ret, 2)),
            ('正收益次数', len(wins)),
            ('亏损次数', len(losses)),
            ('>5% 大胜次数', len(big_wins)),
        ]
        for i, (k, v) in enumerate(summary_data):
            c = ws.cell(row=summary_row + i + 1, column=1, value=k)
            c.font = Font(bold=True)
            c2 = ws.cell(row=summary_row + i + 1, column=2, value=v)
            c2.alignment = Alignment(horizontal='center')

    return len(bt_rows), avg_ret if bt_rows else 0


s6_bt_cnt, s6_avg = write_backtest_sheet(wb, 'MACD回测_5日', sheet6_results, all_klines)
s7_bt_cnt, s7_avg = write_backtest_sheet(wb, 'MACD回测_10日', sheet7_results, all_klines)


# ============ 10. G. 每日上榜数量折线图 ============
from collections import Counter
from openpyxl.chart import LineChart, Reference

# Sheet6 折线图
chart_s6 = LineChart()
chart_s6.title = 'Sheet6 每日上榜数量趋势 (5日>10%)'
chart_s6.style = 12
chart_s6.y_axis.title = '上榜数量'
chart_s6.x_axis.title = '日期'

# 数据写入临时位置（避免污染主 sheet）
s6_cnt_by_day = Counter(r['signal_date'] for r in sheet6_results)
s7_cnt_by_day = Counter(r['signal_date'] for r in sheet7_results)
days_sorted = sorted(set(list(s6_cnt_by_day.keys()) + list(s7_cnt_by_day.keys())))

# 写入临时表格区
if '_chart_data' in wb.sheetnames:
    del wb['_chart_data']
ws_tmp = wb.create_sheet('_chart_data')
ws_tmp.cell(row=1, column=1, value='日期')
ws_tmp.cell(row=1, column=2, value='Sheet6 5日')
ws_tmp.cell(row=1, column=3, value='Sheet7 10日')
for i, d in enumerate(days_sorted, 2):
    ws_tmp.cell(row=i, column=1, value=d)
    ws_tmp.cell(row=i, column=2, value=s6_cnt_by_day.get(d, 0))
    ws_tmp.cell(row=i, column=3, value=s7_cnt_by_day.get(d, 0))

data_s6 = Reference(ws_tmp, min_col=2, min_row=1, max_col=2, max_row=len(days_sorted)+1)
data_s7 = Reference(ws_tmp, min_col=3, min_row=1, max_col=3, max_row=len(days_sorted)+1)
cats = Reference(ws_tmp, min_col=1, min_row=2, max_row=len(days_sorted)+1)
chart_s6.add_data(data_s6, titles_from_data=True)
chart_s6.add_data(data_s7, titles_from_data=True)
chart_s6.set_categories(cats)
chart_s6.width = 20
chart_s6.height = 10
# 画在 Sheet6 v2 旁边
ws6_v2 = wb['MACD强势个股_v2']
ws6_v2.add_chart(chart_s6, 'P2')


# ============ 11. 写入 Excel（主 sheet 部分） ============
print(f'\n===== Step5: 写入 Excel =====')
import openpyxl
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side


def write_sheet(wb, sheet_name, results, gain_label):
    if sheet_name in wb.sheetnames:
        del wb[sheet_name]
    ws = wb.create_sheet(sheet_name)

    thin = Side(style='thin', color='CCCCCC')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    headers = ['信号日', '排名', '代码', '名称', '最新价', 'MACD',
               f'{gain_label}涨幅%', '信号日→今日%', '缠论分数',
               '位置', '趋势', '分型', '背驰', '量能', '备注']
    col_widths = [12, 6, 10, 14, 10, 10, 12, 14, 10, 18, 22, 12, 14, 12, 30]

    for i, (h, w) in enumerate(zip(headers, col_widths), 1):
        ws.column_dimensions[get_column_letter(i)].width = w
        c = ws.cell(row=1, column=i, value=h)
        c.font = Font(bold=True, color='FFFFFF', size=11)
        c.fill = PatternFill(start_color='2F5496', end_color='2F5496', fill_type='solid')
        c.alignment = Alignment(horizontal='center', vertical='center')
        c.border = border

    ws.auto_filter.ref = ws.dimensions
    ws.freeze_panes = 'A2'

    RANK_COLORS = ['C0392B', 'E74C3C', 'E67E22', 'F39C12', '27AE60',
                   '2ECC71', '3498DB', '8E44AD', '16A085', '7F8C8D']

    # 按日期分组，组内排名
    from collections import defaultdict
    by_day = defaultdict(list)
    for r in results:
        by_day[r['signal_date']].append(r)
    for d in by_day:
        by_day[d].sort(key=lambda x: -x['score'])

    flat = []
    for d in sorted(by_day.keys(), reverse=True):
        for r in by_day[d]:
            flat.append(r)

    for row_i, r in enumerate(flat, 2):
        d = r.get('details', {})
        score_val = r.get('score') or 0
        # 排名 = 组内排名
        rank_in_day = [x for x in by_day[r['signal_date']]].index(r) + 1
        rank_color = RANK_COLORS[min(rank_in_day - 1, len(RANK_COLORS) - 1)]
        row_data = [
            r['signal_date'],
            rank_in_day,
            r['code'],
            r['name'],
            r['close'],
            r['macd'],
            r['gain_nd'],
            r['gain_since'] if r['gain_since'] is not None else '',
            score_val,
            d.get('position', ''),
            d.get('trend', ''),
            d.get('fx', ''),
            d.get('bcie', ''),
            d.get('vol', ''),
            f"ZG:{d.get('zg','')} ZD:{d.get('zd','')}",
        ]
        for col_i, val in enumerate(row_data, 1):
            c = ws.cell(row=row_i, column=col_i, value=val)
            c.border = border
            c.alignment = Alignment(horizontal='center', vertical='center')
            if col_i == 2:
                c.fill = PatternFill(start_color=rank_color, end_color=rank_color, fill_type='solid')
                c.font = Font(bold=True, color='FFFFFF')
            elif col_i == 9 and score_val:
                if score_val >= 3:
                    c.fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
                    c.font = Font(bold=True, color='006100')
                elif score_val >= 1:
                    c.fill = PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid')
                    c.font = Font(bold=True, color='9C5700')
                elif score_val <= -1:
                    c.fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
                    c.font = Font(bold=True, color='9C0006')

write_sheet(wb, 'MACD强势个股_v2', sheet6_results, '5日')
write_sheet(wb, 'MACD强势个股_10日_v2', sheet7_results, '10日')


# ============ 跨 sheet 联动索引 ============
def build_code_index(wb, sheet_name):
    """扫描 v2 sheet，建立 code -> 第一个匹配行号的索引"""
    ws = wb[sheet_name]
    code_to_row = {}
    for row_i in range(2, ws.max_row + 1):
        code_val = ws.cell(row=row_i, column=3).value  # 代码列
        if code_val and code_val not in code_to_row:
            code_to_row[code_val] = row_i
    return code_to_row

s6_code_idx = build_code_index(wb, 'MACD强势个股_v2')
s7_code_idx = build_code_index(wb, 'MACD强势个股_10日_v2')
print(f'联动索引: Sheet6 {len(s6_code_idx)} 个代码, Sheet7 {len(s7_code_idx)} 个代码')


# ============ 7. 上榜频次 TOP N 汇总 ============
def write_freq_sheet(wb, sheet_name, results, gain_label, code_idx=None, target_sheet=None):
    """按股票代码汇总上榜频次。code_idx: {code6: row} 用于超链接跳转"""
    if sheet_name in wb.sheetnames:
        del wb[sheet_name]
    ws = wb.create_sheet(sheet_name)

    thin = Side(style='thin', color='CCCCCC')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    # 按 code 聚合
    from collections import defaultdict
    by_code = defaultdict(list)
    for r in results:
        by_code[r['code']].append(r)

    summary = []
    for code6, recs in by_code.items():
        recs.sort(key=lambda x: x['signal_date'])
        days = sorted(set(r['signal_date'] for r in recs))
        avg_score = sum(r['score'] for r in recs) / len(recs)
        # 末次上榜日→今日
        last_day = days[-1]
        last_rec = [r for r in recs if r['signal_date'] == last_day][0]
        gain_since_last = last_rec.get('gain_since')
        # 首日上榜日
        first_day = days[0]
        # 末次上榜日的平均涨幅
        avg_gain_nd = sum(r['gain_nd'] for r in recs) / len(recs)
        summary.append({
            'code': code6,
            'name': recs[0]['name'],
            'cnt': len(days),
            'first_day': first_day,
            'last_day': last_day,
            'avg_score': round(avg_score, 1),
            'avg_gain': round(avg_gain_nd, 2),
            'gain_since_last': gain_since_last,
        })

    # 按上榜天数降序
    summary.sort(key=lambda x: (-x['cnt'], -x['avg_score']))

    headers = ['排名', '代码', '名称', '上榜天数', '首次上榜日', '末次上榜日',
               f'平均{gain_label}涨幅%', '平均缠论分数', '末次上榜日→今日%']
    col_widths = [6, 10, 14, 10, 12, 12, 14, 12, 16]

    for i, (h, w) in enumerate(zip(headers, col_widths), 1):
        ws.column_dimensions[get_column_letter(i)].width = w
        c = ws.cell(row=1, column=i, value=h)
        c.font = Font(bold=True, color='FFFFFF', size=11)
        c.fill = PatternFill(start_color='C0392B', end_color='C0392B', fill_type='solid')
        c.alignment = Alignment(horizontal='center', vertical='center')
        c.border = border

    ws.freeze_panes = 'A2'
    ws.auto_filter.ref = ws.dimensions

    RANK_COLORS = ['C0392B', 'E74C3C', 'E67E22', 'F39C12', '27AE60',
                   '2ECC71', '3498DB', '8E44AD', '16A085', '7F8C8D']

    for row_i, s in enumerate(summary, 2):
        rank = row_i - 2
        rank_color = RANK_COLORS[min(rank, len(RANK_COLORS) - 1)]
        row_data = [
            rank + 1,
            s['code'],
            s['name'],
            s['cnt'],
            s['first_day'],
            s['last_day'],
            s['avg_gain'],
            s['avg_score'],
            s['gain_since_last'] if s['gain_since_last'] is not None else '',
        ]
        for col_i, val in enumerate(row_data, 1):
            c = ws.cell(row=row_i, column=col_i, value=val)
            c.border = border
            c.alignment = Alignment(horizontal='center', vertical='center')
            if col_i == 1:
                c.fill = PatternFill(start_color=rank_color, end_color=rank_color, fill_type='solid')
                c.font = Font(bold=True, color='FFFFFF')
            elif col_i == 2:
                # 代码列加超链接跳转到 v2 sheet
                if code_idx and val in code_idx and target_sheet:
                    target_row = code_idx[val]
                    c.hyperlink = f"#'{target_sheet}'!A{target_row}"
                    c.font = Font(color='0563C1', underline='single', bold=True)
                else:
                    c.font = Font(color='0563C1', underline='single')
            elif col_i == 4 and val >= 10:
                c.fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
                c.font = Font(bold=True, color='006100')
            elif col_i == 4 and val >= 5:
                c.fill = PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid')
                c.font = Font(bold=True, color='9C5700')

    return len(summary)


s6_freq = write_freq_sheet(wb, 'MACD上榜频次_5日', sheet6_results, '5日',
                            code_idx=s6_code_idx, target_sheet='MACD强势个股_v2')
s7_freq = write_freq_sheet(wb, 'MACD上榜频次_10日', sheet7_results, '10日',
                            code_idx=s7_code_idx, target_sheet='MACD强势个股_10日_v2')

wb.save(EXCEL_FILE)
print(f'\n写入完成: {EXCEL_FILE}')
print(f'  Sheet "MACD强势个股_v2": {len(sheet6_results)} 行')
print(f'  Sheet "MACD强势个股_10日_v2": {len(sheet7_results)} 行')
print(f'  Sheet "MACD上榜频次_5日": {s6_freq} 只去重后的股票')
print(f'  Sheet "MACD上榜频次_10日": {s7_freq} 只去重后的股票')

# 打印 TOP 10
print('\n===== 上榜频次 TOP 10 (Sheet6 5日) =====')
from collections import defaultdict
by_code = defaultdict(list)
for r in sheet6_results:
    by_code[r['code']].append(r)
top10 = sorted(by_code.items(), key=lambda x: -len(set(r['signal_date'] for r in x[1])))[:10]
for code6, recs in top10:
    days = sorted(set(r['signal_date'] for r in recs))
    avg_s = sum(r['score'] for r in recs) / len(recs)
    print(f'  {code6} {recs[0]["name"]:8s} {len(days)}天 '
          f'[{days[0]}~{days[-1]}] 均分{avg_s:+.1f}')

print('\n===== 上榜频次 TOP 10 (Sheet7 10日) =====')
by_code = defaultdict(list)
for r in sheet7_results:
    by_code[r['code']].append(r)
top10 = sorted(by_code.items(), key=lambda x: -len(set(r['signal_date'] for r in x[1])))[:10]
for code6, recs in top10:
    days = sorted(set(r['signal_date'] for r in recs))
    avg_s = sum(r['score'] for r in recs) / len(recs)
    print(f'  {code6} {recs[0]["name"]:8s} {len(days)}天 '
          f'[{days[0]}~{days[-1]}] 均分{avg_s:+.1f}')

print('\n===== 全 10 天版完成 =====')


# ============ 12. 生成 HTML 报告 ============
print(f'\n===== Step6: 生成 HTML 报告 =====')
import matplotlib
matplotlib.use('Agg')
import matplotlib.pyplot as plt
from io import BytesIO
import base64
from datetime import datetime as _dt

plt.rcParams['font.sans-serif'] = ['Microsoft YaHei', 'SimHei', 'DejaVu Sans']
plt.rcParams['axes.unicode_minus'] = False

REPORT_DIR = r'D:\stock\tool\stock\reports'
os.makedirs(REPORT_DIR, exist_ok=True)


def fig_to_base64(fig):
    """matplotlib 图转 base64"""
    buf = BytesIO()
    fig.savefig(buf, format='png', dpi=120, bbox_inches='tight')
    buf.seek(0)
    return base64.b64encode(buf.read()).decode()


# --- 图1: 每日上榜数量趋势 ---
fig1, ax1 = plt.subplots(figsize=(12, 5))
days_plot = sorted(set(list(s6_cnt_by_day.keys()) + list(s7_cnt_by_day.keys())))
s6_y = [s6_cnt_by_day.get(d, 0) for d in days_plot]
s7_y = [s7_cnt_by_day.get(d, 0) for d in days_plot]
ax1.plot(days_plot, s6_y, marker='o', linewidth=2, label='Sheet6 (5日>10%)', color='#3498DB')
ax1.plot(days_plot, s7_y, marker='s', linewidth=2, label='Sheet7 (10日>20%)', color='#E74C3C')
ax1.set_title('每日上榜数量趋势', fontsize=14, fontweight='bold')
ax1.set_xlabel('日期')
ax1.set_ylabel('上榜数量')
ax1.legend(loc='upper left')
ax1.grid(alpha=0.3)
plt.xticks(rotation=45)
img1 = fig_to_base64(fig1)
plt.close(fig1)

# --- 图2: TOP 10 概念（Sheet6 5日）---
# 从 Excel 读（刚写过的 sheet）
ws_concept = wb['MACD概念聚合_5日']
top_concepts = []
for row_i in range(2, 12):  # TOP 10
    name = ws_concept.cell(row=row_i, column=2).value
    cnt = ws_concept.cell(row=row_i, column=4).value
    if name and cnt:
        top_concepts.append((name, cnt))

if top_concepts:
    fig2, ax2 = plt.subplots(figsize=(10, 5))
    names = [c[0] for c in top_concepts]
    cnts = [c[1] for c in top_concepts]
    bars = ax2.barh(range(len(names)), cnts, color='#16A085')
    ax2.set_yticks(range(len(names)))
    ax2.set_yticklabels(names)
    ax2.invert_yaxis()
    ax2.set_title('TOP 10 概念 (Sheet6 5日 总上榜次数)', fontsize=14, fontweight='bold')
    ax2.set_xlabel('上榜次数')
    for bar, cnt in zip(bars, cnts):
        ax2.text(bar.get_width() + 5, bar.get_y() + bar.get_height()/2,
                 str(cnt), va='center', fontsize=10)
    img2 = fig_to_base64(fig2)
    plt.close(fig2)
else:
    img2 = ''

# --- 图3: Sheet6 上榜频次 TOP 10 ---
ws_freq = wb['MACD上榜频次_5日']
top_freq = []
for row_i in range(2, 12):
    name = ws_freq.cell(row=row_i, column=3).value
    cnt = ws_freq.cell(row=row_i, column=4).value
    if name and cnt:
        top_freq.append((name, cnt))

if top_freq:
    fig3, ax3 = plt.subplots(figsize=(10, 5))
    names = [f[0] for f in top_freq]
    cnts = [f[1] for f in top_freq]
    bars = ax3.barh(range(len(names)), cnts, color='#C0392B')
    ax3.set_yticks(range(len(names)))
    ax3.set_yticklabels(names)
    ax3.invert_yaxis()
    ax3.set_title('TOP 10 上榜频次 (Sheet6 5日)', fontsize=14, fontweight='bold')
    ax3.set_xlabel('上榜天数')
    for bar, cnt in zip(bars, cnts):
        ax3.text(bar.get_width() + 0.1, bar.get_y() + bar.get_height()/2,
                 str(cnt), va='center', fontsize=10)
    img3 = fig_to_base64(fig3)
    plt.close(fig3)
else:
    img3 = ''

# --- 汇总数字 ---
total_s6 = len(sheet6_results)
total_s7 = len(sheet7_results)
avg_s6_per_day = total_s6 / len(trade_days)
avg_s7_per_day = total_s7 / len(trade_days)
top_concept_name = top_concepts[0][0] if top_concepts else 'N/A'
top_freq_name = top_freq[0][0] if top_freq else 'N/A'
gen_time = _dt.now().strftime('%Y-%m-%d %H:%M:%S')

# --- 最近 10 个交易日数据 ---
recent_10_days = trade_days[-10:] if len(trade_days) >= 10 else trade_days
recent_10_set = set(recent_10_days)
s6_recent = [r for r in sheet6_results if r['signal_date'] in recent_10_set]
s7_recent = [r for r in sheet7_results if r['signal_date'] in recent_10_set]
print(f'最近 10 个交易日: {recent_10_days[0]} ~ {recent_10_days[-1]}')
print(f'  Sheet6 最近 10 天: {len(s6_recent)} 行')
print(f'  Sheet7 最近 10 天: {len(s7_recent)} 行')

# --- 回测汇总 ---
bt_summary_s6 = []
ws_bt6 = wb['MACD回测_5日']
for row in ws_bt6.iter_rows(values_only=True):
    if row[0] in ('总交易数', '平均收益%', '胜率%', '最大收益%', '最大亏损%'):
        bt_summary_s6.append((row[0], row[1]))


def _build_recent_10_table(prefix, rows, gain_label):
    """生成最近 10 天数据的 HTML 表格，含 data-date / data-prefix 供 JS 过滤"""
    if not rows:
        return f'<p style="color:#94a3b8;">无数据</p>'
    # 按日期+分数排
    rows_sorted = sorted(rows, key=lambda r: (r['signal_date'], -r['score']), reverse=True)
    out = [f'<h3 style="color:#60a5fa;margin-top:15px;">� {prefix.upper()} (MACD>0 + {gain_label}) · 共 {len(rows)} 行</h3>']
    out.append('<table>')
    out.append('<tr><th>信号日</th><th>代码</th><th>名称</th><th>最新价</th>'
               '<th>MACD</th><th>' + gain_label + '</th><th>信号日→今日%</th>'
               '<th>缠论分数</th><th>位置</th></tr>')
    for r in rows_sorted:
        d = r.get('details', {})
        pos = d.get('position', '')
        gain_since = r.get('gain_since')
        gain_since_str = f'{gain_since:+.1f}%' if gain_since is not None else ''
        out.append(
            f'<tr data-prefix="{prefix}" data-date="{r["signal_date"]}">'
            f'<td>{r["signal_date"]}</td>'
            f'<td>{r["code"]}</td>'
            f'<td>{r["name"]}</td>'
            f'<td>{r["close"]}</td>'
            f'<td>{r["macd"]:.3f}</td>'
            f'<td>{r["gain_nd"]:+.1f}%</td>'
            f'<td>{gain_since_str}</td>'
            f'<td>{r["score"]:+.1f}</td>'
            f'<td>{pos}</td>'
            f'</tr>'
        )
    out.append('</table>')
    return '\n'.join(out)


# --- HTML 模板 ---
html = f'''<!DOCTYPE html>
<html lang="zh-CN">
<head>
<meta charset="utf-8">
<title>MACD 强势个股报告 - {TODAY}</title>
<style>
body {{ font-family: "Microsoft YaHei", "微软雅黑", sans-serif; margin: 20px; background: #f5f7fa; color: #2c3e50; }}
h1 {{ color: #2c3e50; border-bottom: 3px solid #3498DB; padding-bottom: 10px; }}
h2 {{ color: #34495e; margin-top: 30px; padding: 8px 12px; background: #ecf0f1; border-left: 4px solid #3498DB; }}
.kpi-grid {{ display: grid; grid-template-columns: repeat(4, 1fr); gap: 15px; margin: 20px 0; }}
.kpi {{ background: white; padding: 15px; border-radius: 8px; box-shadow: 0 2px 4px rgba(0,0,0,0.1); text-align: center; }}
.kpi-label {{ font-size: 13px; color: #7f8c8d; margin-bottom: 5px; }}
.kpi-value {{ font-size: 24px; font-weight: bold; color: #2c3e50; }}
.kpi-sub {{ font-size: 11px; color: #95a5a6; margin-top: 5px; }}
.chart {{ background: white; padding: 15px; border-radius: 8px; box-shadow: 0 2px 4px rgba(0,0,0,0.1); margin: 15px 0; text-align: center; }}
.chart img {{ max-width: 100%; height: auto; }}
table {{ width: 100%; border-collapse: collapse; background: white; box-shadow: 0 2px 4px rgba(0,0,0,0.1); margin: 10px 0; }}
th {{ background: #2F5496; color: white; padding: 10px; text-align: center; font-size: 13px; }}
td {{ padding: 8px; text-align: center; border-bottom: 1px solid #ecf0f1; font-size: 13px; }}
tr:hover td {{ background: #f8f9fa; }}
.green {{ color: #27ae60; font-weight: bold; }}
.red {{ color: #c0392b; font-weight: bold; }}
.meta {{ color: #7f8c8d; font-size: 12px; text-align: right; margin-bottom: 20px; }}
.dark-section {{ background:#1a1a2e; color:#e2e8f0; padding:20px; border-radius:8px; margin:15px 0; box-shadow:0 2px 4px rgba(0,0,0,0.3); }}
.dark-section h2 {{ color:#60a5fa; margin-top:0; background:transparent; border-left:none; padding:0; }}
.dark-section table {{ background:#16213e; color:#e2e8f0; box-shadow:none; margin:10px 0; }}
.dark-section th {{ background:#2a3f6f; color:#fff; border-bottom:1px solid #2a2a4a; }}
.dark-section td {{ background:#16213e; color:#e2e8f0; border-bottom:1px solid #2a2a4a; }}
.dark-section tr:hover td {{ background:#1e2d4a; }}
.dark-section p {{ color:#94a3b8; }}
</style>
</head>
<body>
<h1>📈 MACD 强势个股报告</h1>
<div class="meta">生成时间: {gen_time} | 截面区间: {START_DATE} ~ {TODAY} ({len(trade_days)} 个交易日)</div>

<h2>📊 核心指标</h2>
<div class="kpi-grid">
  <div class="kpi">
    <div class="kpi-label">Sheet6 记录数</div>
    <div class="kpi-value">{total_s6}</div>
    <div class="kpi-sub">日均 {avg_s6_per_day:.1f} 只</div>
  </div>
  <div class="kpi">
    <div class="kpi-label">Sheet7 记录数</div>
    <div class="kpi-value">{total_s7}</div>
    <div class="kpi-sub">日均 {avg_s7_per_day:.1f} 只</div>
  </div>
  <div class="kpi">
    <div class="kpi-label">TOP 概念 (Sheet6)</div>
    <div class="kpi-value" style="font-size:18px;">{top_concept_name}</div>
    <div class="kpi-sub">{top_concepts[0][1] if top_concepts else 0} 次上榜</div>
  </div>
  <div class="kpi">
    <div class="kpi-label">最持续个股</div>
    <div class="kpi-value" style="font-size:18px;">{top_freq_name}</div>
    <div class="kpi-sub">{top_freq[0][1] if top_freq else 0} 天上榜</div>
  </div>
</div>

<h2>📈 每日上榜数量趋势</h2>
<div class="chart"><img src="data:image/png;base64,{img1}" alt="趋势"></div>

<h2>🔥 热门概念 TOP 10 (Sheet6 5日)</h2>
<div class="chart"><img src="data:image/png;base64,{img2}" alt="概念"></div>

<h2>🏆 上榜频次 TOP 10 (Sheet6 5日)</h2>
<div class="chart"><img src="data:image/png;base64,{img3}" alt="频次"></div>

<div class="dark-section">
<h2>📅 最近 10 个交易日 MACD 强势个股（可按日期搜索）</h2>
<p style="color:#94a3b8;margin-top:0;">两个表格分别对应 Excel 中的 <b>MACD强势个股_v2</b>（{total_s6} 行）和 <b>MACD强势个股_10日_v2</b>（{total_s7} 行）。在输入框中输入日期（如 <code>2026-08-13</code>）可快速查看某一天的股票。</p>
<div style="display:flex;gap:15px;flex-wrap:wrap;align-items:center;margin:10px 0;">
  <label style="color:#e2e8f0;">Sheet6 搜索日期：<input type="text" id="date_s6" placeholder="2026-08-13" oninput="filterByDate('s6')" style="padding:6px 10px;border-radius:4px;border:1px solid #2a3f6f;background:#16213e;color:#e2e8f0;"></label>
  <button onclick="document.getElementById('date_s6').value='';filterByDate('s6')" style="padding:6px 12px;background:#2a3f6f;color:#fff;border:none;border-radius:4px;cursor:pointer;">清除</button>
  <label style="color:#e2e8f0;">Sheet7 搜索日期：<input type="text" id="date_s7" placeholder="2026-08-13" oninput="filterByDate('s7')" style="padding:6px 10px;border-radius:4px;border:1px solid #2a3f6f;background:#16213e;color:#e2e8f0;"></label>
  <button onclick="document.getElementById('date_s7').value='';filterByDate('s7')" style="padding:6px 12px;background:#2a3f6f;color:#fff;border:none;border-radius:4px;cursor:pointer;">清除</button>
</div>
''' + _build_recent_10_table('s6', s6_recent, '5日涨幅%') + f'''
''' + _build_recent_10_table('s7', s7_recent, '10日涨幅%') + f'''
</div>

<div class="dark-section">
<h2>💰 Sheet6 回测结果 (次日开盘买，持有5日收盘卖)</h2>
<table>
  <tr><th>指标</th><th>值</th></tr>
''' + ''.join(f'<tr><td>{k}</td><td><b>{v}</b></td></tr>' for k, v in bt_summary_s6) + f'''
</table>
</div>

<div class="dark-section">
<h2>📋 Excel 文件 sheet 列表</h2>
<table>
  <tr><th>Sheet 名</th><th>说明</th></tr>
  <tr><td><b>MACD强势个股_v2</b></td><td>16 天滚动截面，{total_s6} 行，可按日期筛选</td></tr>
  <tr><td><b>MACD强势个股_10日_v2</b></td><td>16 天滚动截面，{total_s7} 行</td></tr>
  <tr><td><b>MACD上榜频次_5日 / _10日</b></td><td>每只票上榜天数排行（带超链接跳转）</td></tr>
  <tr><td><b>MACD信号消失追踪_5日 / _10日</b></td><td>每只票掉出榜单后的表现</td></tr>
  <tr><td><b>MACD概念聚合_5日 / _10日</b></td><td>按概念汇总上榜股票</td></tr>
  <tr><td><b>MACD回测_5日 / _10日</b></td><td>历史回测详情 + 汇总</td></tr>
</table>
</div>

<p style="color:#7f8c8d;font-size:12px;text-align:center;margin-top:30px;">
数据源: akshare 腾讯 K 线 + 本地计算 | 缓存: D:\\stock\\tool\\stock\\kline_cache\\</p>

<script>
function filterByDate(prefix) {{
  const input = document.getElementById('date_' + prefix);
  const val = input.value.trim();
  const rows = document.querySelectorAll('tr[data-prefix="' + prefix + '"]');
  let visibleCount = 0;
  rows.forEach(function(r) {{
    const match = !val || r.dataset.date === val || r.dataset.date.indexOf(val) >= 0;
    r.style.display = match ? '' : 'none';
    if (match) visibleCount++;
  }});
  const badge = document.getElementById('count_' + prefix);
  if (badge) badge.textContent = visibleCount;
}}
</script>

</body>
</html>
'''

report_file = os.path.join(REPORT_DIR, f'macd_{TODAY.replace("-", "")}.html')
with open(report_file, 'w', encoding='utf-8') as f:
    f.write(html)
print(f'HTML 报告: {report_file}')
print(f'  文件大小: {os.path.getsize(report_file) / 1024:.1f} KB')

# 同时生成一个固定别名，供 每日复盘看板 的 iframe 嵌入
latest_file = os.path.join(REPORT_DIR, 'macd_latest.html')
import shutil
shutil.copy2(report_file, latest_file)
print(f'最新别名: {latest_file}')

# 同时注入到“每日复盘看板”HTML
import subprocess as _sp
_inject = r'D:\stock\tool\stock\inject_macd_tab.py'
if os.path.exists(_inject):
    print('\n===== Step7: 注入到每日复盘看板 =====')
    rc = _sp.run(['python', _inject], capture_output=True, text=True)
    print(_inject, 'exit', _sp.returncode) if False else None  # noqa
    # 打印 inject 脚本输出（GBK 安全处理）
    try:
        out = rc.stdout.encode('utf-8', errors='replace').decode('utf-8')
    except Exception:
        out = rc.stdout
    for line in (out or '').splitlines()[-6:]:
        print('  ', line)
    if rc.returncode != 0:
        print('  stderr:', (rc.stderr or '')[:500])
else:
    print(f'未找到注入脚本: {_inject}')

print('\n===== 全部完成 =====')
