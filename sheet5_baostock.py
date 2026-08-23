"""
Sheet5: 每日涨幅>6%但不在当日Top3板块内的个股
==============================================
方案：baostock批量查K线 + 并发加速
  - 生成A股代码列表（约4000只：创业板300xxx + 科创板688xxx + 深市000xxx）
  - 20个worker并发查K线，每次查10天数据
  - 从pctChg字段筛选>6%的股票
  - 排除当天Top3板块内的股票
  - 每天只需1次请求（不限流）
"""
import baostock as bs
import json
import time
import sys
from pathlib import Path
from collections import defaultdict
from concurrent.futures import ThreadPoolExecutor, as_completed

# ============ 1. 读取板块历史 ============
cache_file = Path('data/board_history_ths/history_20260728_20260820.json')
with open(cache_file, encoding='utf-8') as f:
    raw = json.load(f)

all_dates = set()
for sym, info in raw.items():
    for row in info.get('data', []):
        if row.get('d'):
            all_dates.add(row['d'])

dates_sorted = sorted(all_dates, reverse=True)
last_10 = dates_sorted[:10]

# 每天Top3板块
daily_top3 = {}
for date in last_10:
    day_results = []
    for sym, info in raw.items():
        name = info.get('name', sym)
        for row in info.get('data', []):
            if row.get('d') == date:
                pct = row.get('p', 0)
                day_results.append((pct, name))
                break
    day_results.sort(reverse=True)
    daily_top3[date] = {name for _, name in day_results[:3]}

all_boards = {}
for sym, info in raw.items():
    all_boards[info.get('name', sym)] = info.get('type', '')

# 板块涨跌幅映射
board_pct_map = {}
for date in last_10:
    board_pct_map[date] = {}
    for sym, info in raw.items():
        for row in info.get('data', []):
            if row.get('d') == date:
                board_pct_map[date][info.get('name', sym)] = row.get('p', 0)
                break

# ============ 2. 生成代码列表 ============
def gen_codes():
    codes = []
    for i in range(688000, 689100):  # 科创板 ~1000只
        codes.append(f'sh.{i}')
    for i in range(300000, 304000):  # 创业板 ~4000只
        codes.append(f'sz.{i}')
    for i in range(1, 1000):  # 深市主板 ~1000只
        codes.append(f'sz.{i:06d}')
    return codes

# ============ 3. 单次查询函数（每次login+logout）============
def fetch_kline(code):
    """查单只股K线，返回 {date: pct}"""
    try:
        lg = bs.login()
        if lg.error_code != '0':
            return {}
        
        rs = bs.query_history_k_data_plus(
            code, 'date,close,pctChg',
            '2026-08-07', '2026-08-20',
            frequency='d', adjustflag='2')
        
        result = {}
        if rs.error_code == '0' and rs.data:
            for row in rs.data:
                try:
                    pct = float(row[2]) if row[2] else None
                    if pct is not None:
                        date_str = row[0].replace('-', '')
                        result[date_str] = round(pct, 2)
                except:
                    continue
        
        bs.logout()
        return result
    except Exception:
        try:
            bs.logout()
        except:
            pass
        return {}

# ============ 4. 并发查K线 ============
codes = gen_codes()
print(f"代码总数: {len(codes)}")

t0 = time.time()
stock_changes = {}  # {code: {date: pct}}

with ThreadPoolExecutor(max_workers=20) as pool:
    futures = {pool.submit(fetch_kline, code): code for code in codes}
    done = 0
    for future in as_completed(futures):
        code = futures[future]
        data = future.result()
        if data:
            stock_changes[code] = data
        done += 1
        if done % 200 == 0:
            elapsed = time.time() - t0
            print(f"  进度 {done}/{len(codes)}, 有效股票 {len(stock_changes)} 只, 耗时{elapsed:.0f}s")

elapsed = time.time() - t0
print(f"K线查询完成: {len(stock_changes)} 只有效股票, 耗时{elapsed:.0f}s ({elapsed/60:.1f}分钟)")

# ============ 5. 筛选>6%且不在Top3板块的个股 ============
# 由于没有细粒度概念板块成分股，用证监会行业来判断
# 先获取行业分类
print("获取行业分类...")
lg = bs.login()
rs = bs.query_stock_industry()
industry_data = {}
if rs.error_code == '0' and rs.data:
    for row in rs.data:
        code = row[1]   # 'sh.600000'
        industry = row[3]  # 'J66货币金融服务'
        name = row[2]
        if industry and industry.strip():
            industry_data[code] = {'name': name, 'industry': industry.strip()}
bs.logout()

# 行业 -> 同花顺板块名 映射（主要行业）
INDUSTRY_TO_BOARD = {
    'C27医药制造业': '医疗服务',
    'C39计算机、通信和其他电子设备制造业': '半导体',
    'C26化学原料和化学制品制造业': '化学制品',
    'C35专用设备制造业': '专用设备',
    'C36汽车制造业': '汽车整车',
    'C38电气机械和器材制造业': '电力设备',
    'K70房地产业': '房地产开发',
    'J66货币金融服务': '银行',
    'B06煤炭开采和洗选业': '煤炭开采',
    'A01农业': '种植业',
    'C13农副食品加工业': '食品加工',
    'C15酒、饮料和精制茶制造业': '饮料制造',
    'I64互联网和相关服务': '互联网服务',
    'I63电信、广播电视和卫星传输服务': '通信设备',
    'C34通用设备制造业': '通用设备',
    'C30非金属矿物制品业': '建筑材料',
    'C31黑色金属冶炼和压延加工业': '钢铁',
    'C32有色金属冶炼和压延加工业': '有色金属',
    'B07石油和天然气开采业': '油气开采',
    'C17纺织业': '纺织制造',
}

# 收集结果
daily_other_stocks = {date: [] for date in last_10}

for code, changes in stock_changes.items():
    name = industry_data.get(code, {}).get('name', code)
    industry = industry_data.get(code, {}).get('industry', '其他')
    industry_board = INDUSTRY_TO_BOARD.get(industry, industry)
    
    for date in last_10:
        pct = changes.get(date)
        if pct is None or pct <= 6.0:
            continue
        
        # 判断是否在当天Top3板块
        # 方式：看该股所属行业是否对应某个Top3板块
        in_top3 = False
        top3 = daily_top3.get(date, set())
        
        # 直接匹配行业板块名
        for tb in top3:
            if tb in industry_board or industry_board in tb:
                in_top3 = True
                break
            # 模糊匹配
            if tb in industry or industry in tb:
                in_top3 = True
                break
        
        # 也有可能在Top3的概念板块里（通过成分股数据判断）
        # 但这里没有细粒度概念数据，用行业近似
        if not in_top3:
            board_pct = board_pct_map.get(date, {}).get(industry_board, 0)
            daily_other_stocks[date].append({
                'date': date,
                'stock_name': name,
                'code': code,
                'change': pct,
                'board': industry_board,
                'board_type': '证监会行业',
                'board_pct': board_pct,
            })

for date in last_10:
    daily_other_stocks[date].sort(key=lambda x: -x['change'])

# ============ 6. 写入Excel ============
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

print("\n写入Excel...")
out_path = 'data/板块轮动Top10_v4_含非Top3强势个股.xlsx'
wb = load_workbook('data/板块轮动Top10_v3_含每日Top3强势个股.xlsx')

if '非Top3板块强势个股' in wb.sheetnames:
    del wb['非Top3板块强势个股']

ws = wb.create_sheet('非Top3板块强势个股')

DAY_COLORS = [
    'D6E4F0', 'EBF5FB', 'D5F5E3', 'FEF9E7', 'FDEDEC',
    'F4ECF7', 'E8F8F5', 'FDF2E9', 'F0F3FF', 'F9EBEA',
]

thin = Side(style='thin', color='CCCCCC')
border = Border(left=thin, right=thin, top=thin, bottom=thin)

def fmt_date(d):
    return f"{d[:4]}-{d[4:6]}-{d[6:]}"

col_widths = [14, 12, 16, 22, 12, 12, 12]
col_headers = ['日期', '个股代码', '个股简称', '所属证监会行业', '行业分类', '行业涨幅', '个股涨幅']
for i, (w, h) in enumerate(zip(col_widths, col_headers), 1):
    ws.column_dimensions[get_column_letter(i)].width = w
    c = ws.cell(row=1, column=i, value=h)
    c.font = Font(bold=True, color='FFFFFF', size=11)
    c.fill = PatternFill(start_color='2F5496', end_color='2F5496', fill_type='solid')
    c.alignment = Alignment(horizontal='center', vertical='center')
    c.border = border

ws.freeze_panes = 'A2'

row_num = 2
total_stocks = 0

for day_idx, date in enumerate(reversed(last_10)):
    day_color = DAY_COLORS[day_idx % len(DAY_COLORS)]
    items = daily_other_stocks[date]
    date_start_row = row_num

    if items:
        first = items[0]
        for col_i, val in enumerate([fmt_date(date), first['code'], first['stock_name'],
                                      first['board'], first['board_type'],
                                      f"{first['board_pct']:+.2f}%", f"{first['change']:+.2f}%"], 1):
            c = ws.cell(row=row_num, column=col_i, value=val)
            c.fill = PatternFill(start_color=day_color, end_color=day_color, fill_type='solid')
            c.border = border
            c.alignment = Alignment(horizontal='center', vertical='center')
            if col_i == 6:
                c.alignment = Alignment(horizontal='right', vertical='center')
                c.font = Font(bold=True, color='C00000')
            elif col_i == 7:
                c.alignment = Alignment(horizontal='right', vertical='center')
                c.font = Font(bold=True, color='C00000')
        ws.cell(row=row_num, column=1).font = Font(bold=True)
        row_num += 1

        for item in items[1:]:
            ws.cell(row=row_num, column=1, value='')
            for col_i, val in enumerate(['', item['code'], item['stock_name'], item['board'],
                                          item['board_type'],
                                          f"{item['board_pct']:+.2f}%", f"{item['change']:+.2f}%"], 1):
                c = ws.cell(row=row_num, column=col_i, value=val)
                c.fill = PatternFill(start_color=day_color, end_color=day_color, fill_type='solid')
                c.border = border
                c.alignment = Alignment(horizontal='center', vertical='center')
                if col_i == 6:
                    c.alignment = Alignment(horizontal='right', vertical='center')
                    c.font = Font(bold=True, color='C00000')
                elif col_i == 7:
                    c.alignment = Alignment(horizontal='right', vertical='center')
                    c.font = Font(bold=True, color='C00000')
            row_num += 1
    else:
        c_date = ws.cell(row=row_num, column=1, value=fmt_date(date))
        c_date.font = Font(bold=True)
        c_note = ws.cell(row=row_num, column=3, value='无>6%个股（不在Top3板块）')
        for col_i in range(1, 8):
            ws.cell(row=row_num, column=col_i).fill = PatternFill(start_color=day_color, end_color=day_color, fill_type='solid')
            ws.cell(row=row_num, column=col_i).border = border
        c_date.alignment = Alignment(horizontal='center', vertical='center')
        row_num += 1

    if date_start_row < row_num - 1:
        ws.merge_cells(start_row=date_start_row, start_column=1, end_row=row_num - 1, end_column=1)
        ws.cell(row=date_start_row, column=1).alignment = Alignment(horizontal='center', vertical='center')
        ws.cell(row=date_start_row, column=1).font = Font(bold=True)

    total_stocks += len(items)
    print(f"  {fmt_date(date)}: {len(items)} 只非Top3行业>6%个股")
    row_num += 1

wb.save(out_path)
print(f"\n已保存: {out_path}")
print(f"Sheet5 共 {total_stocks} 条")
