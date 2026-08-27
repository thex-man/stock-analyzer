# -*- coding: utf-8 -*-
"""
从08-25备份恢复MACD sheets到当前v4 Excel
当前v4已有Sheet1/2/3/4/5，缺MACD相关sheet
"""
import openpyxl
from openpyxl import load_workbook

bak_path = 'data/板块轮动Top10_v4_含非Top3强势个股.bak.20260825_194537.xlsx'
cur_path = 'data/板块轮动Top10_v4_含非Top3强势个股.xlsx'

wb_bak = load_workbook(bak_path, data_only=True)
wb_cur = load_workbook(cur_path)

# MACD sheets to restore (skip today's - use backup data for these)
macd_sheets = [
    'MACD强势个股',
    'MACD5日_最近10日',
    'MACD强势个股_10日',
    'MACD信号消失追踪_5日',
    'MACD信号消失追踪_10日',
    'MACD概念聚合_5日',
    'MACD概念聚合_10日',
    'MACD回测_5日',
    'MACD回测_10日',
    'MACD强势个股_v2',
    'MACD强势个股_10日_v2',
    'MACD上榜频次_5日',
    'MACD上榜频次_10日',
]

print('从备份恢复 MACD sheets:')
restored = []
for sname in macd_sheets:
    if sname not in wb_bak.sheetnames:
        print(f'  [跳过] {sname} 不在备份中')
        continue
    if sname in wb_cur.sheetnames:
        print(f'  [跳过] {sname} 已在当前v4中')
        continue

    ws_bak = wb_bak[sname]
    ws_new = wb_cur.create_sheet(sname)

    # 复制表头（第1行）
    for col in range(1, ws_bak.max_column + 1):
        v = ws_bak.cell(row=1, column=col).value
        ws_new.cell(row=1, column=col, value=v)

    # 复制数据行
    for row in range(2, ws_bak.max_row + 1):
        for col in range(1, ws_bak.max_column + 1):
            v = ws_bak.cell(row=row, column=col).value
            ws_new.cell(row=row, column=col, value=v)

    print(f'  [OK] {sname}: {ws_bak.max_row} 行')
    restored.append(sname)

wb_bak.close()
wb_cur.save(cur_path)

wb_v = load_workbook(cur_path, data_only=True)
print(f'\n当前v4全部sheets ({len(wb_v.sheetnames)}个):')
for s in wb_v.sheetnames:
    ws = wb_v[s]
    print(f'  {s}: {ws.max_row}行')
wb_v.close()

print(f'\n已保存: {cur_path}')
print(f'恢复 {len(restored)} 个 MACD sheets')
