# -*- coding: utf-8 -*-
"""
从备份恢复 Sheet4/5 历史数据，保留当前v4的Sheet1/2/3
逻辑：当前v4 = 今日跑的（Sheet1/2/3新数据 + Sheet4/5只有今日）
目标：v4 = Sheet1/2/3（今日） + Sheet4/5（今日+历史）
"""
import openpyxl
from openpyxl import load_workbook
from copy import copy

bak_path = 'data/板块轮动Top10_v4_含非Top3强势个股.bak.20260825_194537.xlsx'
cur_path = 'data/板块轮动Top10_v4_含非Top3强势个股.xlsx'

wb_bak = load_workbook(bak_path, data_only=True)
wb_cur = load_workbook(cur_path)

# 备份中的历史日期（08-11~08-24，共10天）
# 今日(08-25)的数据用当前v4的，备份里的今日数据直接丢弃

def copy_sheet_keep_current_header(wb_src, wb_dst, src_name, dst_name, today='2026-08-25'):
    """从备份复制sheet，但跳过今日（保留目标文件的今日数据）"""
    ws_src = wb_src[src_name]
    ws_dst = wb_dst[dst_name]

    # 清理目标sheet的历史行（保留表头第1行 + 今日数据）
    # 找到今日之后的所有行并删除
    max_row = ws_dst.max_row
    rows_to_delete = []
    for r in range(2, max_row + 1):
        v = ws_dst.cell(row=r, column=1).value
        if v and str(v)[:10] != today:
            rows_to_delete.append(r)

    for r in reversed(rows_to_delete):
        ws_dst.delete_rows(r)
    print(f'  [{dst_name}] 清理旧历史: 删除 {len(rows_to_delete)} 行，保留今日数据')

    # 从备份追加历史数据（跳过今日）
    src_row_count = 0
    skip_count = 0
    for row in ws_src.iter_rows(min_row=2, values_only=True):
        date_val = str(row[0])[:10] if row[0] else ''
        if date_val == today:
            skip_count += 1
            continue
        src_row_count += 1
        new_row = ws_dst.max_row + 1
        for col_i, val in enumerate(row, 1):
            ws_dst.cell(row=new_row, column=col_i, value=val)

    print(f'  [{dst_name}] 从备份追加: {src_row_count} 行历史（跳过今日 {skip_count} 行）')
    return src_row_count

# 处理 Sheet4
print('=== Sheet4 每日Top3强势个股 ===')
s4_count = copy_sheet_keep_current_header(wb_bak, wb_cur, '每日Top3强势个股', '每日Top3强势个股')

# 处理 Sheet5
print('\n=== Sheet5 非Top3板块强势个股 ===')
s5_count = copy_sheet_keep_current_header(wb_bak, wb_cur, '非Top3板块强势个股', '非Top3板块强势个股')

wb_bak.close()
wb_cur.save(cur_path)

# 验证
wb_verify = load_workbook(cur_path, data_only=True)
ws4_v = wb_verify['每日Top3强势个股']
ws5_v = wb_verify['非Top3板块强势个股']
dates4 = sorted(set(str(ws4_v.cell(r,1).value)[:10] for r in range(2, ws4_v.max_row+1) if ws4_v.cell(r,1).value and '2026' in str(ws4_v.cell(r,1).value)))
dates5 = sorted(set(str(ws5_v.cell(r,1).value)[:10] for r in range(2, ws5_v.max_row+1) if ws5_v.cell(r,1).value and '2026' in str(ws5_v.cell(r,1).value)))
print(f'\n验证 Sheet4 日期: {dates4}')
print(f'验证 Sheet5 日期: {dates5}')
wb_verify.close()

print(f'\n已保存: {cur_path}')
