# coding: utf-8
import datetime
import tkinter as tk
import pandas.core.frame
import pandas as pd
import pywencai
from tkinter import ttk, scrolledtext
from datetime import datetime, timedelta, time
import re
from datetime import date
from ttkbootstrap import Style
import openpyxl
import os
import warnings
warnings.filterwarnings('ignore')

# 尝试导入akshare和matplotlib，如果失败则提供替代方案
try:
    import akshare as ak
    import matplotlib.pyplot as plt
    import matplotlib.dates as mdates
    from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg
    from matplotlib.figure import Figure
    
    # 设置中文字体
    plt.rcParams['font.sans-serif'] = ['SimHei', 'Microsoft YaHei']
    plt.rcParams['axes.unicode_minus'] = False
    
    AKSHARE_AVAILABLE = True
except ImportError as e:
    print(f"警告: 无法导入akshare或matplotlib: {e}")
    print("K线图功能将不可用，请安装: pip install akshare matplotlib")
    AKSHARE_AVAILABLE = False

class KLineGUI:
    """K线图GUI界面，点击股票名自动加载K线图"""
    
    def __init__(self, stock_data):
        self.stock_data = stock_data
        self.current_stock = None
        self.current_kline_data = None
        
        # 创建主窗口
        self.root = tk.Tk()
        self.root.title("股票信息查看器 - 点击股票自动显示K线图")
        self.root.geometry("1400x900")
        
        # 创建样式
        self.style = Style(theme='darkly')
        
        # 创建界面
        self.create_widgets()
        
    def create_widgets(self):
        """创建界面组件，移除加载按钮，优化交互"""
        # 主框架
        main_frame = ttk.Frame(self.root, padding="10")
        main_frame.grid(row=0, column=0, sticky=(tk.W, tk.E, tk.N, tk.S))
        
        # 配置网格权重
        self.root.columnconfigure(0, weight=1)
        self.root.rowconfigure(0, weight=1)
        main_frame.columnconfigure(1, weight=3)
        main_frame.columnconfigure(0, weight=1)
        main_frame.rowconfigure(1, weight=1)
        
        # ---------------------- 左侧：股票列表区域 ----------------------
        stock_list_frame = ttk.LabelFrame(main_frame, text="股票列表（点击股票名加载K线图）", padding="10")
        stock_list_frame.grid(row=0, column=0, rowspan=2, sticky=(tk.W, tk.E, tk.N, tk.S), padx=(0, 10), pady=(0, 10))
        stock_list_frame.columnconfigure(0, weight=1)
        stock_list_frame.rowconfigure(0, weight=1)
        
        # 创建股票列表（股票名称列可点击）
        columns = ("序号", "股票简称", "股票代码", "最新价", "经营范围(%)")
        self.stock_tree = ttk.Treeview(stock_list_frame, columns=columns, show="headings", height=20)
        
        # 设置列标题和宽度
        self.stock_tree.heading("序号", text="序号")
        self.stock_tree.heading("股票简称", text="股票简称")  # 点击此列会触发加载
        self.stock_tree.heading("股票代码", text="股票代码")
        self.stock_tree.heading("最新价", text="最新价")
        self.stock_tree.heading("经营范围(%)", text="经营范围(%)")
        
        self.stock_tree.column("序号", width=50, anchor="center")
        self.stock_tree.column("股票简称", width=100, anchor="center")  # 股票名称列
        self.stock_tree.column("股票代码", width=100, anchor="center")
        self.stock_tree.column("最新价", width=80, anchor="center")
        self.stock_tree.column("经营范围(%)", width=100, anchor="center")
        
        # 添加滚动条
        tree_scroll = ttk.Scrollbar(stock_list_frame, orient="vertical", command=self.stock_tree.yview)
        self.stock_tree.configure(yscrollcommand=tree_scroll.set)
        
        # 布局列表和滚动条
        self.stock_tree.grid(row=0, column=0, sticky=(tk.W, tk.E, tk.N, tk.S))
        tree_scroll.grid(row=0, column=1, sticky=(tk.N, tk.S))
        
        # 填充股票数据
        self.populate_stock_list()
        
        # 绑定列表点击事件（点击任意行自动加载K线图）
        self.stock_tree.bind("<<TreeviewSelect>>", self.on_stock_row_selected)
        
        # ---------------------- 右侧：K线图和详情区域 ----------------------
        # 控制面板（移除加载按钮，保留状态提示）
        control_frame = ttk.LabelFrame(main_frame, text="状态", padding="10")
        control_frame.grid(row=0, column=1, sticky=(tk.W, tk.E), pady=(0, 10))
        
        # 状态标签（显示当前操作状态）
        self.status_label = ttk.Label(control_frame, text="请从左侧选择股票")
        self.status_label.grid(row=0, column=0, padx=(10, 0))
        
        # 图表显示区域
        chart_frame = ttk.LabelFrame(main_frame, text="K线图", padding="10")
        chart_frame.grid(row=1, column=1, sticky=(tk.W, tk.E, tk.N, tk.S), pady=(0, 10))
        chart_frame.columnconfigure(0, weight=1)
        chart_frame.rowconfigure(0, weight=1)
        
        # 创建matplotlib图形
        self.fig = Figure(figsize=(12, 8), dpi=100)
        self.canvas = FigureCanvasTkAgg(self.fig, chart_frame)
        self.canvas.get_tk_widget().grid(row=0, column=0, sticky=(tk.W, tk.E, tk.N, tk.S))
        
        # 信息显示区域
        info_frame = ttk.LabelFrame(main_frame, text="股票详情", padding="10")
        info_frame.grid(row=2, column=1, sticky=(tk.W, tk.E), pady=(10, 0))
        
        self.info_text = scrolledtext.ScrolledText(info_frame, height=6, width=80)
        self.info_text.grid(row=0, column=0, sticky=(tk.W, tk.E))
        
    def populate_stock_list(self):
        """填充股票数据到列表"""
        for idx, stock in enumerate(self.stock_data, 1):
            name = stock.get('股票简称', '未知')
            code = stock.get('股票代码', '未知')
            price = stock.get('最新价', '--')
            change = stock.get('经营范围', '--')
            
            self.stock_tree.insert("", tk.END, values=(idx, name, code, price, change))
    
    def on_stock_row_selected(self, event):
        """点击股票行时自动加载K线图"""
        selected_items = self.stock_tree.selection()
        if not selected_items:
            return
        
        # 获取选中行数据
        selected_item = selected_items[0]
        row_data = self.stock_tree.item(selected_item, "values")
        if not row_data:
            return
        
        # 解析股票信息
        name = row_data[1]  # 股票简称
        code = row_data[2]  # 股票代码
        
        # 匹配股票数据
        for stock in self.stock_data:
            if stock['股票简称'] == name and stock['股票代码'] == code:
                self.current_stock = stock
                self.status_label.config(text=f"正在加载 {name} 的K线图...")
                self.root.update()  # 实时更新状态文字
                
                # 自动加载K线图
                self.load_kline_chart()
                # 更新详情
                self.update_info_display()
                break
    
    def update_info_display(self):
        """更新股票详情"""
        if self.current_stock:
            info = f"股票简称: {self.current_stock['股票简称']}\n"
            info += f"股票代码: {self.current_stock['股票代码']}\n"
            info += f"最新价格: {self.current_stock['最新价']}\n"
            info += f"经营范围: {self.current_stock['经营范围']}%\n"
            
            self.info_text.delete(1.0, tk.END)
            self.info_text.insert(1.0, info)
    
    def load_kline_chart(self):
        """自动加载K线图（原按钮触发的逻辑迁移到这里）"""
        if not self.current_stock:
            self.status_label.config(text="未选择股票")
            return
        
        if not AKSHARE_AVAILABLE:
            self.status_label.config(text="akshare未安装，无法加载K线图")
            return
        
        try:
            stock_code = self.current_stock['股票代码']
            kline_data = get_stock_kline_data(stock_code)
            
            if kline_data is not None:
                self.current_kline_data = kline_data
                self.plot_kline_chart_gui()
                self.status_label.config(text=f"已加载 {self.current_stock['股票简称']} 的K线图")
            else:
                self.status_label.config(text=f"无法获取 {self.current_stock['股票简称']} 的K线数据")
                
        except Exception as e:
            self.status_label.config(text=f"加载失败: {str(e)}")
    
    def plot_kline_chart_gui(self):
        """绘制K线图（保持原有逻辑）"""
        if self.current_kline_data is None or self.current_kline_data.empty:
            return
        
        self.fig.clear()
        ax1 = self.fig.add_subplot(211)
        ax2 = self.fig.add_subplot(212)
        
        dates = pd.to_datetime(self.current_kline_data['日期'])
        opens = self.current_kline_data['开盘']
        highs = self.current_kline_data['最高']
        lows = self.current_kline_data['最低']
        closes = self.current_kline_data['收盘']
        volumes = self.current_kline_data['成交量']
        
        ma3 = closes.rolling(window=3).mean()
        ma5 = closes.rolling(window=5).mean()
        ma13 = closes.rolling(window=13).mean()
        
        price_changes = ((closes - opens) / opens * 100).round(2)
        
        for i in range(len(dates)):
            if closes[i] >= opens[i]:
                color = 'red'
                alpha = 0.8
            else:
                color = 'green'
                alpha = 0.8
            
            ax1.plot([dates[i], dates[i]], [lows[i], highs[i]], color='black', linewidth=1)
            rect_height = abs(closes[i] - opens[i])
            rect_bottom = min(opens[i], closes[i])
            rect = plt.Rectangle((dates[i] - timedelta(hours=6), rect_bottom), 
                               timedelta(hours=12), rect_height, 
                               facecolor=color, alpha=alpha, edgecolor='black')
            ax1.add_patch(rect)
        
        ax1.plot(dates, ma3, color='orange', linewidth=2, label='MA3', alpha=0.8)
        ax1.plot(dates, ma5, color='blue', linewidth=2, label='MA5', alpha=0.8)
        ax1.plot(dates, ma13, color='purple', linewidth=2, label='MA13', alpha=0.8)
        
        max_change_idx = price_changes.idxmax()
        min_change_idx = price_changes.idxmin()
        
        if not pd.isna(max_change_idx):
            max_date = dates[max_change_idx]
            max_price = highs[max_change_idx]
            max_change = price_changes[max_change_idx]
            ax1.annotate(f'最高: {max_change}%', 
                        xy=(max_date, max_price), 
                        xytext=(10, 10), textcoords='offset points',
                        bbox=dict(boxstyle='round,pad=0.3', facecolor='yellow', alpha=0.7),
                        arrowprops=dict(arrowstyle='->', connectionstyle='arc3,rad=0'),
                        fontsize=10, fontweight='bold')
        
        if not pd.isna(min_change_idx):
            min_date = dates[min_change_idx]
            min_price = lows[min_change_idx]
            min_change = price_changes[min_change_idx]
            ax1.annotate(f'最低: {min_change}%', 
                        xy=(min_date, min_price), 
                        xytext=(10, -10), textcoords='offset points',
                        bbox=dict(boxstyle='round,pad=0.3', facecolor='lightblue', alpha=0.7),
                        arrowprops=dict(arrowstyle='->', connectionstyle='arc3,rad=0'),
                        fontsize=10, fontweight='bold')
        
        ax1.set_ylabel('价格 (元)', fontsize=12)
        ax1.grid(True, alpha=0.3)
        ax1.set_title(f'{self.current_stock["股票简称"]}({self.current_stock["股票代码"]}) K线图 + 移动平均线', 
                     fontsize=14, fontweight='bold')
        ax1.legend(loc='upper left', fontsize=10)
        
        colors = ['red' if close >= open else 'green' for close, open in zip(closes, opens)]
        ax2.bar(dates, volumes, color=colors, alpha=0.7)
        ax2.set_ylabel('成交量', fontsize=12)
        ax2.set_xlabel('日期', fontsize=12)
        ax2.grid(True, alpha=0.3)
        
        ax1.xaxis.set_major_formatter(mdates.DateFormatter('%m-%d'))
        ax2.xaxis.set_major_formatter(mdates.DateFormatter('%m-%d'))
        
        self.fig.tight_layout()
        self.canvas.draw()
    
    def run(self):
        """运行GUI"""
        self.root.mainloop()

# 以下函数保持不变（与之前代码一致）
def clean_stock_code(stock_code):
    if stock_code:
        clean_code = stock_code.split('.')[0]
        return clean_code
    return stock_code

def get_stock_kline_data(stock_code, period='daily', start_date=None, end_date=None):
    if not AKSHARE_AVAILABLE:
        print("akshare不可用，无法获取K线数据")
        return None
    
    try:
        clean_code = clean_stock_code(stock_code)
        if not start_date:
            start_date = (datetime.now() - timedelta(days=30)).strftime('%Y%m%d')
        if not end_date:
            end_date = datetime.now().strftime('%Y%m%d')
        
        print(f"正在获取股票 {clean_code} 的K线数据...")
        stock_data = ak.stock_zh_a_hist(symbol=clean_code, period=period, 
                                       start_date=start_date, end_date=end_date, 
                                       adjust="qfq")
        
        if stock_data.empty:
            print(f"无法获取股票 {clean_code} 的K线数据")
            return None
        
        print(f"成功获取股票 {clean_code} 的K线数据，共 {len(stock_data)} 条记录")
        return stock_data
        
    except Exception as e:
        print(f"获取股票 {stock_code} K线数据时出错: {e}")
        return None

def plot_kline_chart(stock_data, stock_name, stock_code, save_path=None):
    if not AKSHARE_AVAILABLE:
        print("matplotlib不可用，无法绘制K线图")
        return False
    
    try:
        if stock_data is None or stock_data.empty:
            print(f"股票 {stock_name} 没有数据可绘制")
            return False
        
        fig, (ax1, ax2) = plt.subplots(2, 1, figsize=(12, 8), height_ratios=[3, 1])
        fig.suptitle(f'{stock_name}({stock_code}) K线图', fontsize=16, fontweight='bold')
        
        dates = pd.to_datetime(stock_data['日期'])
        opens = stock_data['开盘']
        highs = stock_data['最高']
        lows = stock_data['最低']
        closes = stock_data['收盘']
        volumes = stock_data['成交量']
        
        ma3 = closes.rolling(window=3).mean()
        ma5 = closes.rolling(window=5).mean()
        ma13 = closes.rolling(window=13).mean()
        price_changes = ((closes - opens) / opens * 100).round(2)
        
        for i in range(len(dates)):
            if closes[i] >= opens[i]:
                color = 'red'
                alpha = 0.8
            else:
                color = 'green'
                alpha = 0.8
            
            ax1.plot([dates[i], dates[i]], [lows[i], highs[i]], color='black', linewidth=1)
            rect_height = abs(closes[i] - opens[i])
            rect_bottom = min(opens[i], closes[i])
            rect = plt.Rectangle((dates[i] - timedelta(hours=6), rect_bottom), 
                               timedelta(hours=12), rect_height, 
                               facecolor=color, alpha=alpha, edgecolor='black')
            ax1.add_patch(rect)
        
        ax1.plot(dates, ma3, color='orange', linewidth=2, label='MA3', alpha=0.8)
        ax1.plot(dates, ma5, color='blue', linewidth=2, label='MA5', alpha=0.8)
        ax1.plot(dates, ma13, color='purple', linewidth=2, label='MA13', alpha=0.8)
        
        max_change_idx = price_changes.idxmax()
        min_change_idx = price_changes.idxmin()
        
        if not pd.isna(max_change_idx):
            max_date = dates[max_change_idx]
            max_price = highs[max_change_idx]
            max_change = price_changes[max_change_idx]
            ax1.annotate(f'最高: {max_change}%', 
                        xy=(max_date, max_price), 
                        xytext=(10, 10), textcoords='offset points',
                        bbox=dict(boxstyle='round,pad=0.3', facecolor='yellow', alpha=0.7),
                        arrowprops=dict(arrowstyle='->', connectionstyle='arc3,rad=0'),
                        fontsize=10, fontweight='bold')
        
        if not pd.isna(min_change_idx):
            min_date = dates[min_change_idx]
            min_price = lows[min_change_idx]
            min_change = price_changes[min_change_idx]
            ax1.annotate(f'最低: {min_change}%', 
                        xy=(min_date, min_price), 
                        xytext=(10, -10), textcoords='offset points',
                        bbox=dict(boxstyle='round,pad=0.3', facecolor='lightblue', alpha=0.7),
                        arrowprops=dict(arrowstyle='->', connectionstyle='arc3,rad=0'),
                        fontsize=10, fontweight='bold')
        
        ax1.set_ylabel('价格 (元)', fontsize=12)
        ax1.grid(True, alpha=0.3)
        ax1.set_title('K线图 + 移动平均线', fontsize=14)
        ax1.legend(loc='upper left', fontsize=10)
        
        colors = ['red' if close >= open else 'green' for close, open in zip(closes, opens)]
        ax2.bar(dates, volumes, color=colors, alpha=0.7)
        ax2.set_ylabel('成交量', fontsize=12)
        ax2.set_xlabel('日期', fontsize=12)
        ax2.grid(True, alpha=0.3)
        
        ax1.xaxis.set_major_formatter(mdates.DateFormatter('%m-%d'))
        ax2.xaxis.set_major_formatter(mdates.DateFormatter('%m-%d'))
        
        plt.tight_layout()
        
        if save_path:
            if not os.path.exists(save_path):
                os.makedirs(save_path)
            filename = f"{stock_name}_{stock_code}_kline.png"
            filepath = os.path.join(save_path, filename)
            plt.savefig(filepath, dpi=300, bbox_inches='tight')
            print(f"K线图已保存: {filepath}")
        
        plt.show()
        return True
        
    except Exception as e:
        print(f"绘制股票 {stock_name} K线图时出错: {e}")
        return False

def get_stock_data():
    now = datetime.now()
    today = date.today()
    
    try:
        if now.time() < time(9, 25):
            stock_query_result = pywencai.get(
                query='创业板，前七个交易日的经营范围大于29%，非新股,非ST股', 
                sort_key='', 
                sort_order='asc'
            )
            print("使用昨天的日期:")
        else:
            stock_query_result = pywencai.get(
                query='创业板，前七个交易日的经营范围大于29%，非新股,非ST股', 
                sort_key='', 
                sort_order='asc',
                query_type='stock'
            )
            print("使用今天的日期:")
        
        print(stock_query_result)
        
        if isinstance(stock_query_result, pandas.core.frame.DataFrame):
            print("返回的是DataFrame格式")
            print(f"数据形状: {stock_query_result.shape}")
            print("列名:", stock_query_result.columns.tolist())
            
            stock_data = []
            if '股票简称' in stock_query_result.columns:
                for index, row in stock_query_result.iterrows():
                    stock_info = {
                        '股票简称': row.get('股票简称', ''),
                        '股票代码': row.get('股票代码', ''),
                        '最新价': row.get('最新价', ''),
                        '经营范围': row.get('经营范围', '') if '经营范围' in stock_query_result.columns else ''
                    }
                    stock_data.append(stock_info)
                
                stock_names = [item['股票简称'] for item in stock_data if item['股票简称']]
                print(f"股票简称列表: {stock_names}")
                print(f"共获取到 {len(stock_names)} 只股票")
                return stock_data
            else:
                print("DataFrame中没有找到'股票简称'列")
                print("可用的列名:", stock_query_result.columns.tolist())
                return []
        else:
            print(f"返回的数据类型是: {type(stock_query_result)}")
            result_str = str(stock_query_result)
            stock_codes = re.findall(r'30\d{4}', result_str)
            print("提取的股票代码:", stock_codes)
            
            stock_data = []
            for code in stock_codes:
                try:
                    stock_info = pywencai.get(
                        query=code, 
                        sort_key='', 
                        sort_order='asc',
                        query_type='stock'
                    )
                    
                    if isinstance(stock_info, pandas.core.frame.DataFrame):
                        if "股票简称" in stock_info.columns:
                            stock_data.append({
                                '股票简称': stock_info["股票简称"].iloc[0],
                                '股票代码': code,
                                '最新价': '',
                                '经营范围': ''
                            })
                        break

                except Exception as e:
                    print(f"获取股票 {code} 信息时出错: {e}")
                    continue
            
            stock_names = [item['股票简称'] for item in stock_data if item['股票简称']]
            print("股票简称列表:", stock_names)
            return stock_data
        
    except Exception as e:
        print(f"获取股票数据时出错: {e}")
        import traceback
        traceback.print_exc()
        return []

def generate_kline_charts(stock_data, save_path='kline_charts'):
    if not stock_data:
        print("没有股票数据可生成K线图")
        return
    
    if not AKSHARE_AVAILABLE:
        print("akshare或matplotlib不可用，无法生成K线图")
        print("请安装: pip install akshare matplotlib")
        return
    
    print(f"\n开始为 {len(stock_data)} 只股票生成K线图...")
    
    success_count = 0
    for i, stock_info in enumerate(stock_data, 1):
        stock_name = stock_info.get('股票简称', '')
        stock_code = stock_info.get('股票代码', '')
        
        if not stock_code:
            print(f"{i}. 跳过 {stock_name} - 无股票代码")
            continue
        
        print(f"{i}. 正在生成 {stock_name}({stock_code}) 的K线图...")
        kline_data = get_stock_kline_data(stock_code)
        
        if kline_data is not None:
            if plot_kline_chart(kline_data, stock_name, stock_code, save_path):
                success_count += 1
                print(f"    ✓ {stock_name} K线图生成成功")
            else:
                print(f"    ✗ {stock_name} K线图生成失败")
        else:
            print(f"    ✗ {stock_name} 无法获取K线数据")
    
    print(f"\nK线图生成完成！成功生成 {success_count}/{len(stock_data)} 只股票的K线图")

def save_to_excel(stock_data):
    try:
        file_path = 'D:\\stock\\tool\\data\\everyday_data.xlsx'
        
        if not os.path.exists(file_path):
            print(f"Excel文件不存在: {file_path}")
            return False
        
        workbook = openpyxl.load_workbook(file_path)
        sheet = workbook['Sheet1']
        today = date.today().strftime("%Y/%m/%d")
        print("今天的日期为:", today)
        
        next_row = sheet.max_row + 1
        sheet.cell(row=next_row, column=1, value=today)
        
        for i, stock_info in enumerate(stock_data, start=2):
            sheet.cell(row=next_row, column=i, value=stock_info.get('股票简称', ''))
            sheet.cell(row=next_row, column=i+len(stock_data), value=stock_info.get('股票代码', ''))
            sheet.cell(row=next_row, column=i+len(stock_data)*2, value=stock_info.get('最新价', ''))
            sheet.cell(row=next_row, column=i+len(stock_data)*3, value=stock_info.get('经营范围', ''))
        
        workbook.save(file_path)
        print(f"数据已保存到Excel文件，共保存了 {len(stock_data)} 只股票")
        return True
        
    except Exception as e:
        print(f"保存Excel文件时出错: {e}")
        return False

def print_stock_summary(stock_data):
    if not stock_data:
        print("没有股票数据可显示")
        return
    
    print("\n=== 股票数据汇总 ===")
    print(f"日期: {date.today().strftime('%Y/%m/%d')}")
    print(f"股票数量: {len(stock_data)}")
    print("\n股票列表:")
    for i, stock in enumerate(stock_data, 1):
        print(f"{i:2d}. {stock.get('股票简称', 'N/A')} ({stock.get('股票代码', 'N/A')}) - 价格: {stock.get('最新价', 'N/A')} - 经营范围: {stock.get('经营范围', 'N/A')}")
    print("=" * 30)

def main():
    print("开始获取股票数据...")
    
    stock_data = get_stock_data()
    
    if stock_data:
        print(f"成功获取到 {len(stock_data)} 只股票")
        print_stock_summary(stock_data)
        save_to_excel(stock_data)
        
        if AKSHARE_AVAILABLE:
            user_input = input("\n是否要启动股票查看器界面？(y/n): ").lower().strip()
            if user_input in ['y', 'yes', '是']:
                print("启动股票查看器界面...")
                gui = KLineGUI(stock_data)
                gui.run()
            else:
                user_input2 = input("是否要为这些股票生成K线图文件？(y/n): ").lower().strip()
                if user_input2 in ['y', 'yes', '是']:
                    generate_kline_charts(stock_data)
                else:
                    print("跳过K线图生成")
        else:
            print("\n注意: akshare或matplotlib未安装，无法使用界面功能")
            print("如需界面功能，请安装: pip install akshare matplotlib ttkbootstrap")
    else:
        print("没有获取到股票数据")

if __name__ == "__main__":
    main()