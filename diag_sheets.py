# -*- coding: utf-8 -*-
"""诊断 Sheet4 和 Sheet5 的数据问题"""
import json
from pathlib import Path
from openpyxl import load_workbook

# 读取缓存看今日行业情况
cache_files = sorted(Path('data/board_history_ths').glob('history_*.json'))
with open(cache_files[-1], encoding='utf-8') as f:
    raw = json.load(f)

all_dates = set()
for sym, info in raw.items():
    for row in info.get('data', []):
        if row.get('d'):
            all_dates.add(row['d'])
dates_sorted = sorted(all_dates, reverse=True)
last_10 = dates_sorted[:10]
print("缓存中的日期:", last_10)

# 今日行业 Top3
target = '20260826'
today_data = []
for sym, info in raw.items():
    if info.get('type') == '行业':
        for row in info.get('data', []):
            if row.get('d') == target:
                today_data.append((row.get('p', 0), info.get('name', sym), sym))
                break
today_data.sort(reverse=True)
print(f"\n今日({target})行业 Top10:")
for pct, name, sym in today_data[:10]:
    print(f"  {name:20s} {pct:+.2f}%")

# 读取 v4 Excel 看 Sheet4 内容
print("\n--- Sheet4 每日Top3强势个股 ---")
wb = load_workbook('data/板块轮动Top10_v4_含非Top3强势个股.xlsx', data_only=True)
ws4 = wb['每日Top3强势个股']
print(f"总行数: {ws4.max_row}")
# 前20行
for row in ws4.iter_rows(min_row=1, max_row=20, values_only=True):
    print(f"  {row}")

print("\n--- Sheet5 非Top3强势个股 ---")
if '非Top3板块强势个股' in wb.sheetnames:
    ws5 = wb['非Top3板块强势个股']
    print(f"总行数: {ws5.max_row}")
    for row in ws5.iter_rows(min_row=1, max_row=20, values_only=True):
        print(f"  {row}")
else:
    print("Sheet5 不存在")
