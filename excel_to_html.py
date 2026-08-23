"""
将 Excel 转换为可交互的 HTML 可视化页面
"""
import openpyxl
from openpyxl.styles import PatternFill

wb = openpyxl.load_workbook('data/板块轮动Top10_v4_含非Top3强势个股.xlsx')

def get_cell_color(cell):
    fill = cell.fill
    if fill and fill.fill_type == 'solid':
        fg = fill.fgColor
        if fg.type == 'rgb':
            s = str(fg.rgb).strip().upper().lstrip('#')
            return s[2:] if len(s) == 8 else s  # 去掉AA前缀
    return None

# 收集各 Sheet 数据（保持独立引用）
s1_rows = []  # 行业板块
for row in wb['行业板块'].iter_rows(values_only=False):
    s1_rows.append([{'value': c.value, 'color': get_cell_color(c)} for c in row])

s2_rows = []  # 概念板块
for row in wb['概念板块'].iter_rows(values_only=False):
    s2_rows.append([{'value': c.value, 'color': get_cell_color(c)} for c in row])

s3_rows = []  # 色卡图例
for row in wb['色卡图例'].iter_rows(values_only=False):
    s3_rows.append([{'value': c.value, 'color': get_cell_color(c)} for c in row])

s4_rows = []  # 每日Top3强势个股
for row in wb['每日Top3强势个股'].iter_rows(values_only=False):
    s4_rows.append([{'value': c.value, 'color': get_cell_color(c)} for c in row])

s5_rows = []  # 非Top3板块强势个股
for row in wb['非Top3板块强势个股'].iter_rows(values_only=False):
    s5_rows.append([{'value': c.value, 'color': get_cell_color(c)} for c in row])

s6_rows = []  # MACD强势个股（5日>10%）
if 'MACD强势个股' in wb.sheetnames:
    for row in wb['MACD强势个股'].iter_rows(values_only=False):
        s6_rows.append([{'value': c.value, 'color': get_cell_color(c)} for c in row])

s7_rows = []  # MACD强势个股（10日>20%）
if 'MACD强势个股_10日' in wb.sheetnames:
    for row in wb['MACD强势个股_10日'].iter_rows(values_only=False):
        s7_rows.append([{'value': c.value, 'color': get_cell_color(c)} for c in row])

def build_sheet1_html(label, rows):
    # ---- 计算出现次数>2且每次涨幅>1%的板块并分配颜色 ----
    board_info = {}  # name -> {'count': N, 'gains': [g1, g2...]}
    for r_idx, row in enumerate(rows):
        if r_idx == 0:
            continue
        for c_idx, cell in enumerate(row):
            if c_idx == 0:
                continue  # 跳过日期列
            val = cell['value'] or ''
            name = val.split(' ')[0].strip()
            if not name:
                continue
            # 提取涨幅数字
            import re
            m = re.search(r'([+-]?\d+\.?\d*)%', val)
            gain = float(m.group(1)) if m else 0.0
            if name not in board_info:
                board_info[name] = {'count': 0, 'gains': []}
            board_info[name]['count'] += 1
            board_info[name]['gains'].append(gain)

    PALETTE = [
        'FF6B6B', 'FF9F43', 'FECA57', '48DBFB', '1DD1A1',
        'A55EEA', 'FD79A8', '00CEC9', '6C5CE7', 'FDCB6E',
        '74B9FF', '55EFC4', 'FF7675', 'D63031', 'E17055',
        '009432', '0652DD', 'FFC312', '1289A7', '12CBC4',
        'EE5A24', 'B53405', '686DE0', '22A6B3', 'F19066',
        'FDA7DF', 'D980FA', 'FAB1A0', '7ED6DF', 'C44569',
        '778BEB', '70A1FF', '7BED9F', 'ECCC68', 'FF6348',
        'A3CB38', '55A3FF', 'FED330', '26DE81', '2BCBBA',
        'EB5D68', '4B7BEC', '45AAF2', 'FCAB10', 'F8B739',
    ]
    # 出现次数>2 且超过2次涨幅>1%
    repeat_boards = {
        b: info['count']
        for b, info in board_info.items()
        if info['count'] > 2 and sum(1 for g in info['gains'] if g > 1.0) > 2
    }
    board_color = {}
    for i, board in enumerate(sorted(repeat_boards)):
        board_color[board] = PALETTE[i % len(PALETTE)]

    # 生成 CSS
    css = ''
    for board, color in board_color.items():
        safe = board.replace('(', '').replace(')', '').replace('.', '').replace('/', '')
        css += f'.board-{safe} {{ background:#{color} !important; color:#1a1a2e !important; font-weight:600; }}\n'

    # ---- 涨幅合计汇总表（合计>10%按降序） ----
    gain_boards = {
        b: sum(info['gains'])
        for b, info in board_info.items()
        if sum(info['gains']) > 10
    }
    gain_sorted = sorted(gain_boards.items(), key=lambda x: -x[1])

    gain_css = ''
    for board, total in gain_sorted:
        safe = board.replace('(', '').replace(')', '').replace('.', '').replace('/', '')
        gain_css += f'.gs-{safe} {{ background:rgba(255,200,80,0.15) !important; }}\n'

    html = f'<style>{css}{gain_css}</style>'
    html += f'<div id="tab-{label}" class="panel">'
    if gain_sorted:
        html += '<div class="repeat-summary">'
        html += f'<h4>&#x1F4C8; 涨幅合计 Top（合计>10%，共{len(gain_sorted)}个板块）</h4>'
        html += '<table class="repeat-table"><thead><tr><th>板块</th><th>合计涨幅</th><th>上榜次数</th><th>平均涨幅</th></tr></thead><tbody>'
        for board, total in gain_sorted:
            safe = board.replace('(', '').replace(')', '').replace('.', '').replace('/', '')
            info = board_info[board]
            avg = total / info['count'] if info['count'] else 0
            html += (f'<tr class="gs-{safe}">'
                     f'<td style="text-align:left;font-weight:600">{board}</td>'
                     f'<td style="text-align:center;font-weight:700;color:#f1c40f">{total:+.2f}%</td>'
                     f'<td style="text-align:center">{info["count"]}</td>'
                     f'<td style="text-align:center;color:#94a3b8">{avg:+.2f}%</td>'
                     f'</tr>')
        html += '</tbody></table></div>'
    html += '<div class="table-wrap"><table>'
    for r_idx, row in enumerate(rows):
        html += '<tr>'
        for c_idx, cell in enumerate(row):
            val = cell['value'] or ''
            cls = 'date-cell' if c_idx == 0 and r_idx > 0 else ''
            if c_idx > 0 and val:
                board_name = val.split(' ')[0].strip()
                if board_name in board_color:
                    safe = board_name.replace('(', '').replace(')', '').replace('.', '').replace('/', '')
                    cls = f'board-{safe}'
            html += f'<td class="{cls}">{val}</td>'
        html += '</tr>'
    html += '</table></div></div>'
    return html

def build_legend_html(rows):
    hy, gn = [], []
    for r_idx, row in enumerate(rows):
        if r_idx == 0: continue
        hy_val = row[0]['value'] if len(row) > 0 else ''
        hy_cnt = row[1]['value'] if len(row) > 1 else ''
        gn_val = row[2]['value'] if len(row) > 2 else ''
        gn_cnt = row[3]['value'] if len(row) > 3 else ''
        if hy_val:
            hy.append(f'<div class="legend-item"><span class="legend-dot" style="background:#F5B942"></span>{hy_val} <span style="color:#94a3b8">{hy_cnt}</span></div>')
        if gn_val:
            gn.append(f'<div class="legend-item"><span class="legend-dot" style="background:#F5B942"></span>{gn_val} <span style="color:#94a3b8">{gn_cnt}</span></div>')
    return f'''<div id="tab-色卡图例" class="panel">
<div class="legend-grid">
  <div class="legend-col"><h3>行业板块（上榜次数>2）</h3>{"".join(hy)}</div>
  <div class="legend-col"><h3>概念板块（上榜次数>2）</h3>{"".join(gn)}</div>
</div>
<p class="legend-note">★ 灰色=出现1-2次，不涂色 &nbsp;|&nbsp; 浅黄=3次 &nbsp;|&nbsp; 橙色=4次 &nbsp;|&nbsp; 浅红=5次 &nbsp;|&nbsp; 深红=6次+</p>
</div>'''

def build_sheet6_html(rows, gain_label='5日涨幅%', tab_id='MACD强势个股', tab_title='MACD强势个股', table_id='table6', filter_id='filter6', stats_id='stats6'):
    """MACD强势个股 — 按缠论分数排名（参数化：支持 5日/10日 涨幅两个变体）"""
    cols = ['排名', '代码', '名称', '最新价', 'MACD', gain_label, '缠论分数',
            '位置', '趋势', '分型', '背驰', '量能', '备注']

    # 按缠论分数分档着色
    score_colors = {
        'high':   'C6EFCE',  # 绿色 分数>=3
        'mid':    'FFEB9C',  # 黄色 分数1~2
        'low':    'FFC7CE',  # 红色 分数<0
    }
    score_text_colors = {
        'high':   '006100',
        'mid':    '9C5700',
        'low':    '9C0006',
    }

    # 排名颜色
    RANK_COLORS = [
        'C0392B','E74C3C','E67E22','F39C12','27AE60',
        '2ECC71','3498DB','8E44AD','16A085','7F8C8D',
    ]

    html = '''<style>
      .s6-table { width:100%; border-collapse:collapse; font-size:0.85rem; }
      .s6-table th { background:#1e2d4a; color:#93c5fd; padding:10px 12px; text-align:center; font-weight:600; border:1px solid #2a3f6f; position:sticky; top:0; }
      .s6-table td { padding:8px 12px; text-align:center; border:1px solid #2a2a4a; }
      .s6-table tr:hover td { background:#2a2f4a; }
      .s6-rank { font-weight:700; color:#fff; }
      .s6-score-high { background:#C6EFCE !important; color:#006100 !important; font-weight:700; }
      .s6-score-mid  { background:#FFEB9C !important; color:#9C5700 !important; font-weight:700; }
      .s6-score-low  { background:#FFC7CE !important; color:#9C0006 !important; }
      .s6-pct-hot { color:#ff4444 !important; font-weight:700; }
    </style>
    <div id="tab-{tab_id}" class="panel">
    '''.replace('{tab_id}', tab_id)
    html += '<div class="filter-bar">'
    html += f'<label>🔍 <input type="text" id="{filter_id}" oninput="filterTable(\'{filter_id}\')" placeholder="搜索代码/名称..."></label>'
    html += '</div>'
    html += f'<div class="stats" id="{stats_id}"></div>'
    html += f'<div class="table-wrap"><table class="s6-table" id="{table_id}">'
    html += '<thead><tr>' + ''.join(f'<th>{c}</th>' for c in cols) + '</tr></thead>'
    html += '<tbody>'

    for r_idx, row in enumerate(rows):
        if r_idx == 0:
            continue
        cells = [c['value'] for c in row]
        # cells: [rank, code, name, close, macd, gain_5d, score, position, trend, fx, bcie, vol, note]
        rank = cells[0] if len(cells) > 0 else ''
        score_val = float(cells[6]) if (len(cells) > 6 and cells[6] is not None) else 0
        gain_val = float(cells[5]) if (len(cells) > 5 and cells[5] is not None) else 0
        rank_color = RANK_COLORS[min(int(rank)-1, len(RANK_COLORS)-1)] if str(rank).isdigit() else '7F8C8D'

        if score_val >= 3:
            score_cls = 's6-score-high'
        elif score_val >= 1:
            score_cls = 's6-score-mid'
        elif score_val <= -1:
            score_cls = 's6-score-low'
        else:
            score_cls = ''

        gain_cls = 's6-pct-hot' if gain_val > 20 else ('' if gain_val > 15 else '')

        html += '<tr>'
        for i, v in enumerate(cells):
            if i == 0:  # 排名
                cls = f's6-rank' + f' background:#{rank_color} !important; color:#fff'
                # inline style for rank cell
                html += f"<td style='background:#{rank_color};color:#fff;font-weight:700'>{v}</td>"
            elif i == 6:  # 缠论分数
                html += f"<td class='{score_cls}'>{v}</td>"
            elif i == 5 and gain_cls:  # 5日涨幅
                html += f"<td style='color:#ff4444;font-weight:700'>{v}</td>"
            else:
                html += f"<td>{v or ''}</td>"
        html += '</tr>'
    html += '</tbody></table></div></div>'
    return html


def build_sheet45_html(rows, sheet_idx, sheet_label, cols):
    from collections import Counter

    # ---- 统计个股出现次数及日期 ----
    stock_counter = Counter()
    stock_dates = {}  # name -> sorted set of dates
    current_date = None
    for r_idx, row in enumerate(rows):
        if r_idx == 0:
            continue
        if row[0]['value']:
            current_date = row[0]['value']
        if sheet_idx == 4:
            v = row[4]['value'] if len(row) > 4 else None
        else:
            v = row[2]['value'] if len(row) > 2 else None
        if v:
            name = str(v).split(' ')[0].strip()
            if name:
                stock_counter[name] += 1
                date_str = str(current_date)[:10] if current_date else ''
                if date_str:
                    stock_dates.setdefault(name, set()).add(date_str)

    PALETTE = [
        'C0392B', 'D35400', 'E67E22', 'F39C12', '27AE60',
        '16A085', '2980B9', '8E44AD', '2C3E50', '7F8C8D',
        'E74C3C', 'E67E22', 'F1C40F', '2ECC71', '1ABC9C',
        '3498DB', '9B59B6', '34495E', 'E91E63', '00BCD4',
        'FF5722', '795548', '607D8B', 'FF9800', 'FFEB3B',
        'CDDC39', '8BC34A', '00E676', '1DE9B6', '00B0FF',
        '651FFF', 'E040FB', 'FF4081', 'FF6E40', 'FFD740',
    ]
    repeat_stocks = {b: c for b, c in stock_counter.items() if c > 1}
    stock_color = {}
    for i, stock in enumerate(sorted(repeat_stocks, key=lambda x: -repeat_stocks[x])):
        stock_color[stock] = PALETTE[i % len(PALETTE)]

    # 生成 CSS
    css = ''
    for stock, color in stock_color.items():
        safe = stock.replace('(', '').replace(')', '').replace('.', '').replace('/', '').replace(' ', '_')
        css += f'.s{sheet_idx}-stock-{safe} {{ background:#{color} !important; color:#000 !important; font-weight:700; }}\n'
        css += f'.s{sheet_idx}-stock-row-{safe} td {{ background:rgba({int(color[0:2],16)},{int(color[2:4],16)},{int(color[4:6],16)},0.2) !important; }}\n'

    html = f'<style>{css}</style>'
    html += f'<div id="tab-{sheet_label}" class="panel">'
    html += '<div class="filter-bar">'
    html += f'<label>🔍 <input type="text" id="filter{sheet_idx}" oninput="filterTable({sheet_idx})" placeholder="搜索个股/板块..."></label>'
    html += f'<select id="dateFilter{sheet_idx}" onchange="filterTable({sheet_idx})"><option value="">全部日期</option>'
    dates = []
    for r_idx, row in enumerate(rows):
        if r_idx == 0:
            continue
        if row[0]['value']:
            dates.append(str(row[0]['value'])[:10])
    for d in sorted(set(dates), reverse=True):
        html += f'<option value="{d}">{d}</option>'
    html += '</select></div>'
    html += f'<div class="stats" id="stats{sheet_idx}"></div>'

    # ---- 重复出现个股汇总表 ----
    if repeat_stocks:
        html += '<div class="repeat-summary">'
        html += f'<h4>🔁 重复出现个股（出现>{1}次，共{len(repeat_stocks)}只）</h4>'
        html += '<table class="repeat-table"><thead><tr><th>个股</th><th>次数</th><th>出现日期</th></tr></thead><tbody>'
        for stock in sorted(repeat_stocks, key=lambda x: -repeat_stocks[x]):
            color = stock_color[stock]
            safe = stock.replace('(', '').replace(')', '').replace('.', '').replace('/', '').replace(' ', '_')
            cnt = stock_counter[stock]
            date_list = sorted(stock_dates.get(stock, []), reverse=True)
            dates_str = ' / '.join(date_list)
            html += (f'<tr class="s{sheet_idx}-stock-row-{safe}">'
                     f'<td class="s{sheet_idx}-stock-{safe}">{stock}</td>'
                     f'<td style="text-align:center;font-weight:700">{cnt}</td>'
                     f'<td style="text-align:left;color:#94a3b8;font-size:0.8rem">{dates_str}</td>'
                     f'</tr>')
        html += '</tbody></table></div>'

    html += f'<div class="table-wrap"><table id="table{sheet_idx}">'
    html += f'<thead><tr>{"".join(f"<th>{c}</th>" for c in cols)}</tr></thead><tbody>'

    current_date = None
    for r_idx, row in enumerate(rows):
        if r_idx == 0:
            continue
        if row[0]['value']:
            current_date = row[0]['value']
        cells = [c['value'] for c in row]
        # Sheet4 日期列为合并单元格，后续行日期为空，用 current_date 补全
        if sheet_idx == 4 and cells[0] is None and current_date:
            cells[0] = current_date
        pct_idx = 4 if sheet_idx == 4 else 5
        pct_val = cells[pct_idx] if len(cells) > pct_idx else ''
        pct_cls = 'pct-positive' if pct_val and '+' in str(pct_val) else ''

        # ---- 个股涂色 ----
        cell_classes = ['', '', '', '', '', '']
        if sheet_idx == 4 and len(cells) > 4:
            stock_name = str(cells[4]).split(' ')[0].strip() if cells[4] else ''
            if stock_name in stock_color:
                safe = stock_name.replace('(', '').replace(')', '').replace('.', '').replace('/', '').replace(' ', '_')
                cell_classes[4] = f's{sheet_idx}-stock-{safe}'
        elif sheet_idx == 5 and len(cells) > 2:
            stock_name = str(cells[2]).strip() if cells[2] else ''
            if stock_name in stock_color:
                safe = stock_name.replace('(', '').replace(')', '').replace('.', '').replace('/', '').replace(' ', '_')
                cell_classes[2] = f's{sheet_idx}-stock-{safe}'

        html += f"<tr>"
        for i, v in enumerate(cells):
            cls = 'date-cell' if i == 0 else (cell_classes[i] if cell_classes[i] else pct_cls)
            html += f"<td class='{cls}'>{v or ''}</td>"
        html += '</tr>'
    html += '</tbody></table></div></div>'
    return html

# ===== Build all panels =====
panels_html = (
    build_sheet1_html('行业板块', s1_rows) +
    build_sheet1_html('概念板块', s2_rows) +
    build_legend_html(s3_rows) +
    build_sheet45_html(s4_rows, 4, '每日Top3强势个股',
                       ['日期', '板块', '板块涨幅', '类型', '强势个股', '个股涨幅']) +
    build_sheet45_html(s5_rows, 5, '非Top3板块强势个股',
                       ['日期', '个股代码', '个股简称', '所属概念', '板块涨幅', '个股涨幅']) +
    (build_sheet6_html(s6_rows, gain_label='5日涨幅%', tab_id='MACD强势个股', table_id='table6', filter_id='filter6', stats_id='stats6') if s6_rows else '') +
    (build_sheet6_html(s7_rows, gain_label='10日涨幅%', tab_id='MACD强势个股_10日', table_id='table7', filter_id='filter7', stats_id='stats7') if s7_rows else '')
)

# ===== Full HTML =====
html = f'''<!DOCTYPE html>
<html lang="zh">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1.0">
<title>每日复盘看板</title>
<style>
* {{ box-sizing: border-box; margin: 0; padding: 0; }}
body {{ font-family: -apple-system, BlinkMacSystemFont, 'Segoe UI', Roboto, sans-serif; background: #1a1a2e; color: #eee; min-height: 100vh; }}
.header {{ background: linear-gradient(135deg, #16213e, #0f3460); padding: 20px 32px; border-bottom: 1px solid #2a5298; }}
.header h1 {{ font-size: 1.5rem; color: #e2e8f0; }}
.header p {{ font-size: 0.85rem; color: #94a3b8; margin-top: 4px; }}
.tabs {{ display: flex; gap: 4px; padding: 16px 32px 0; background: #16213e; flex-wrap: wrap; }}
.tab {{ padding: 10px 20px; background: #1e2d4a; border: none; border-radius: 8px 8px 0 0; color: #94a3b8; cursor: pointer; font-size: 0.9rem; transition: all 0.2s; }}
.tab:hover {{ background: #2a3f6f; color: #e2e8f0; }}
.tab.active {{ background: #1a1a2e; color: #60a5fa; font-weight: 600; }}
.panel {{ display: none; padding: 24px 32px; }}
.panel.active {{ display: block; }}
table {{ width: 100%; border-collapse: collapse; font-size: 0.85rem; }}
th {{ background: #1e2d4a; color: #93c5fd; padding: 10px 12px; text-align: center; font-weight: 600; border: 1px solid #2a3f6f; position: sticky; top: 0; }}
td {{ padding: 8px 12px; text-align: center; border: 1px solid #2a2a4a; }}
tr:hover td {{ background: #2a2f4a; }}
.rank-cell {{ text-align: left; padding-left: 16px; }}
.color-f5b942 {{ background: #F5B942 !important; color: #1a1a2e !important; font-weight: 600; }}
.color-fff9c4 {{ background: #FFF9C4 !important; color: #1a1a2e !important; }}
.color-c00000 {{ background: #C00000 !important; color: #fff !important; font-weight: 600; }}
.color-f4a0a0 {{ background: #F4A0A0 !important; color: #1a1a2e !important; }}
.date-cell {{ font-weight: 700; background: #1e2d4a; color: #93c5fd; }}
.stock-name {{ text-align: left; padding-left: 12px; }}
.pct-positive {{ color: #f87171; font-weight: 700; }}
.filter-bar {{ margin-bottom: 16px; display: flex; gap: 12px; align-items: center; flex-wrap: wrap; }}
.filter-bar input, .filter-bar select {{ padding: 8px 12px; background: #1e2d4a; border: 1px solid #2a3f6f; color: #e2e8f0; border-radius: 6px; font-size: 0.85rem; }}
.filter-bar label {{ color: #94a3b8; font-size: 0.85rem; }}
.legend-grid {{ display: grid; grid-template-columns: 1fr 1fr; gap: 16px; }}
.legend-col h3 {{ color: #93c5fd; margin-bottom: 8px; font-size: 0.9rem; }}
.legend-item {{ display: flex; align-items: center; gap: 8px; padding: 4px 0; font-size: 0.85rem; }}
.legend-dot {{ width: 12px; height: 12px; border-radius: 50%; display: inline-block; }}
.legend-note {{ margin-top: 12px; font-size: 0.8rem; color: #94a3b8; }}
.table-wrap {{ overflow-x: auto; border-radius: 8px; }}
.table-wrap table {{ min-width: 800px; }}
.repeat-summary {{ margin-bottom: 20px; }}
.repeat-summary h4 {{ color: #93c5fd; margin-bottom: 10px; font-size: 0.9rem; }}
.repeat-table {{ width: 100%; border-collapse: collapse; font-size: 0.82rem; margin-bottom: 16px; }}
.repeat-table th {{ background: #1e2d4a; color: #93c5fd; padding: 8px 12px; text-align: center; border: 1px solid #2a3f6f; }}
.repeat-table td {{ padding: 6px 12px; border: 1px solid #2a2a4a; text-align: center; }}
.repeat-table tr:hover td {{ background: #2a2f4a !important; }}
.stats {{ display: flex; gap: 16px; margin-bottom: 20px; flex-wrap: wrap; }}
.stat-card {{ background: #1e2d4a; border-radius: 8px; padding: 16px 24px; flex: 1; min-width: 120px; }}
.stat-card .num {{ font-size: 1.8rem; font-weight: 700; color: #60a5fa; }}
.stat-card .label {{ font-size: 0.8rem; color: #94a3b8; margin-top: 4px; }}
</style>
</head>
<body>
<div class="header">
  <h1>📊 每日复盘看板</h1>
  <p id="data-period"></p>
</div>
<div class="tabs">
  <button class="tab active" onclick="showTab('行业板块', this)">行业板块 Top10</button>
  <button class="tab" onclick="showTab('概念板块', this)">概念板块 Top10</button>
  <button class="tab" onclick="showTab('色卡图例', this)">色卡图例</button>
  <button class="tab" onclick="showTab('每日Top3强势个股', this)">每日Top3强势个股</button>
  <button class="tab" onclick="showTab('非Top3板块强势个股', this)">非Top3板块强势个股</button>
  <button class="tab" onclick="showTab('MACD强势个股', this)">MACD强势个股（5日>10%）</button>
  <button class="tab" onclick="showTab('MACD强势个股_10日', this)">MACD强势个股（10日>20%）</button>
</div>
{panels_html}
<script>
function showTab(name, btn) {{
  document.querySelectorAll('.panel').forEach(p => p.classList.remove('active'));
  document.querySelectorAll('.tab').forEach(t => t.classList.remove('active'));
  document.getElementById('tab-' + name).classList.add('active');
  if (btn) btn.classList.add('active');
  const map = {{'每日Top3强势个股':4,'非Top3板块强势个股':5,'MACD强势个股':6}};
  if (map[name]) updateStats(map[name]);
}}

function filterTable(arg) {{
  // 兼容传入: 'filter6'/'filter7'/'table6'/'table7'/'dateFilter6'/数字 4/5/6/7
  let n;
  if (typeof arg === 'string') {{
    const m = arg.match(/(?:filter|table|stats|dateFilter)(\d+)/);
    n = m ? m[1] : (arg.replace(/[^0-9]/g,'') || '6');
  }} else {{ n = String(arg); }}
  const filterEl = document.getElementById('filter'+n);
  const tableId  = 'table'+n;
  const statsId  = 'stats'+n;
  const text = filterEl ? filterEl.value.toLowerCase() : '';
  const dateVal = document.getElementById('dateFilter'+n)?.value || '';
  document.querySelectorAll('#'+tableId+' tbody tr').forEach(row => {{
    const cells = Array.from(row.cells);
    const matchText = cells.some(c => c.textContent.toLowerCase().includes(text));
    const matchDate = !dateVal || cells[0].textContent.includes(dateVal);
    row.style.display = (matchText && matchDate) ? '' : 'none';
  }});
  const el = document.getElementById(statsId);
  if (!el) return;
  const visible = Array.from(document.querySelectorAll('#'+tableId+' tbody tr')).filter(r => r.style.display !== 'none');
  const dates = new Set(visible.map(r => r.cells[0].textContent.trim()));
  el.innerHTML = `
    <div class="stat-card"><div class="num">${{visible.length}}</div><div class="label">强势个股数量</div></div>
    <div class="stat-card"><div class="num">${{dates.size}}</div><div class="label">交易天数</div></div>
  `;
}}

function updateStats(arg) {{
  filterTable(arg);
}}

document.addEventListener('DOMContentLoaded', () => {{
  // 设置数据周期
  const dateCells = Array.from(document.querySelectorAll('#tab-行业板块 table td:first-child'));
  const allDates = dateCells.map(c => c.textContent.trim()).filter(d => d && d.length === 8);
  const latest = allDates[0] || '';
  const oldest = allDates[allDates.length - 1] || '';
  const periodEl = document.getElementById('data-period');
  if (periodEl && latest && oldest)
    periodEl.textContent = `数据周期：${{oldest}} \u2192 ${{latest}} · ${{allDates.length}} 个交易日`;
  updateStats(4); updateStats(5);
}});
</script>
</body>
</html>
'''

out_path = 'data/每日复盘看板.html'
with open(out_path, 'w', encoding='utf-8') as f:
    f.write(html)
print(f'已生成: {out_path} ({len(html)/1024:.1f} KB)')
