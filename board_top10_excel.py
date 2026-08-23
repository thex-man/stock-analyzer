import json
from pathlib import Path
from collections import Counter
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ============ 读取缓存 ============
cache_files = sorted(Path('data/board_history_ths').glob('history_*.json'))
cache_file = cache_files[-1]  # 取最新的缓存文件
with open(cache_file, encoding='utf-8') as f:
    raw = json.load(f)

all_dates = set()
for sym, info in raw.items():
    for row in info.get('data', []):
        if row.get('d'):
            all_dates.add(row['d'])

dates_sorted = sorted(all_dates, reverse=True)
last_10 = dates_sorted[:10]

# ============ 构建每日Top10（行业/概念分开） ============
daily_top10 = {}
for date in last_10:
    hy_results, gn_results = [], []
    for sym, info in raw.items():
        name = info.get('name', sym)
        btype = info.get('type', '')
        for row in info.get('data', []):
            if row.get('d') == date:
                pct = row.get('p', 0)
                if btype == '行业':
                    hy_results.append((pct, name))
                else:
                    gn_results.append((pct, name))
                break
    hy_results.sort(reverse=True)
    gn_results.sort(reverse=True)
    daily_top10[date] = {
        '行业': hy_results[:10],
        '概念': gn_results[:10]
    }

# ============ 统计每个板块出现次数 ============
def get_board_counts(board_type):
    all_names = []
    for date in last_10:
        for pct, name in daily_top10[date][board_type]:
            all_names.append(name)
    return Counter(all_names)

hy_counts = get_board_counts('行业')
gn_counts = get_board_counts('概念')

repeat_hy = {n: c for n, c in hy_counts.items() if c > 2}
repeat_gn = {n: c for n, c in gn_counts.items() if c > 2}
all_repeat_names = set(repeat_hy.keys()) | set(repeat_gn.keys())

# ============ 每个重复板块分配唯一颜色 ============
# 用固定调色盘，保证每次运行颜色一致且差异大
PALETTE = [
    'FF6B6B', 'FF9F43', 'FECA57', '48DBFB', '1DD1A1',
    'A55EEA', 'FD79A8', '00CEC9', '6C5CE7', 'FDCB6E',
    '74B9FF', '55EFC4', 'FF7675', 'D63031', 'E17055',
    '009432', '0652DD', 'FFC312', '1289A7', '12CBC4',
    'EE5A24', 'B53405', '686DE0', '22A6B3', 'F19066',
    'FDA7DF', 'D980FA', 'FAB1A0', '7ED6DF', 'C44569',
    '778BEB', '70A1FF', '7BED9F', 'ECCC68', 'FF6348',
    'A3CB38', '55A3FF', 'FED330', '26DE81', '2BCBBA',
    'EB5D68', '4B7BEC', '45AAF2', 'FCAB10', 'F8B739',
    'E056FD', 'F8A5C2', 'F3A683', '70A1FF', 'A29BFE',
    '55EFC4', 'FDCB6E', 'FAB1A0', 'E74C3C', '3498DB',
    '1ABC9C', '2ECC71', 'F39C12', '9B59B6', '636E72',
    'D1A3FF', 'FF9FF3', 'F8A5C2', '778BEB', 'D980FA',
    '74B9FF', 'A55EEA', 'FD79A8', '00CEC9', 'FECA57',
    'FF6B6B', 'FF9F43', '1DD1A1', '48DBFB', 'E17055',
    'FDCB6E', 'D63031', '009432', 'B53405', '22A6B3',
    '686DE0', 'F19066', 'FDA7DF', 'FF7675', '55EFC4',
    '1289A7', 'FFC312', 'FF6348', '55A3FF', 'A3CB38',
    '6C5CE7', 'FED330', '26DE81', 'FF9F43', 'FCAB10',
    '45AAF2', 'E056FD', '74B9FF', '7BED9F', 'FAB1A0',
]

board_colors = {}
for i, name in enumerate(sorted(all_repeat_names)):
    board_colors[name] = PALETTE[i % len(PALETTE)]

# 白色（只出现1次不上色）
def get_fill(name):
    if name in board_colors:
        return PatternFill(start_color=board_colors[name], end_color=board_colors[name], fill_type="solid")
    return PatternFill(fill_type=None)

# ============ 样式 ============
thin = Side(style='thin', color='DDDDDD')
border = Border(left=thin, right=thin, top=thin, bottom=thin)

def make_sheet(ws, board_type):
    ws.title = f"{board_type}板块"
    counts = hy_counts if board_type == '行业' else gn_counts

    # 标题行
    headers = ["日期"] + [f"第{i}名" for i in range(1, 11)]
    ws.append(headers)

    title_fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
    title_font = Font(bold=True, size=11, color="FFFFFF")
    title_align = Alignment(horizontal="center", vertical="center")
    for col_idx in range(1, 12):
        c = ws.cell(row=1, column=col_idx)
        c.font = title_font
        c.fill = title_fill
        c.alignment = title_align
        c.border = border

    # 数据行（从旧到新）
    for row_idx, date in enumerate(reversed(last_10), 2):
        results = daily_top10[date][board_type]
        row_data = [date] + [f"{name} {pct:+.2f}%" for pct, name in results]
        ws.append(row_data)

        # 日期列
        dc = ws.cell(row=row_idx, column=1)
        dc.font = Font(bold=True)
        dc.alignment = Alignment(horizontal="center", vertical="center")
        dc.fill = PatternFill(start_color="BDD7EE", end_color="BDD7EE", fill_type="solid")
        dc.border = border

        # 10个排名列
        for col_idx in range(2, 12):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.alignment = Alignment(horizontal="left", vertical="center")
            cell.border = border
            cell_text = cell.value or ""
            board_name = cell_text.split(" ")[0]
            cell.fill = get_fill(board_name)

    # 列宽
    ws.column_dimensions['A'].width = 14
    for col_idx in range(2, 12):
        ws.column_dimensions[get_column_letter(col_idx)].width = 20
    ws.freeze_panes = "A2"

wb = Workbook()
make_sheet(wb.active, '行业')
make_sheet(wb.create_sheet("概念板块"), '概念')

# ============ SHEET 3: 色卡图例 ============
ws_leg = wb.create_sheet("色卡图例")
ws_leg.column_dimensions['A'].width = 24
ws_leg.column_dimensions['B'].width = 12
ws_leg.column_dimensions['C'].width = 12
ws_leg.column_dimensions['D'].width = 16

# 标题
for col, text in enumerate(["行业板块", "上榜次数", "概念板块", "上榜次数"], 1):
    c = ws_leg.cell(row=1, column=col, value=text)
    c.font = Font(bold=True, color="FFFFFF", size=11)
    c.fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
    c.alignment = Alignment(horizontal="center", vertical="center")
    c.border = border

hy_sorted = sorted(repeat_hy.items(), key=lambda x: (-x[1], x[0]))
gn_sorted = sorted(repeat_gn.items(), key=lambda x: (-x[1], x[0]))
max_rows = max(len(hy_sorted), len(gn_sorted))

for i in range(max_rows):
    row = i + 2
    if i < len(hy_sorted):
        name, cnt = hy_sorted[i]
        c1 = ws_leg.cell(row=row, column=1, value=name)
        c2 = ws_leg.cell(row=row, column=2, value=f"{cnt}次")
        fill = get_fill(name)
        c1.fill = fill
        c1.alignment = Alignment(horizontal="left", vertical="center")
        c2.alignment = Alignment(horizontal="center", vertical="center")
        c1.border = border
        c2.border = border
    if i < len(gn_sorted):
        name, cnt = gn_sorted[i]
        c1 = ws_leg.cell(row=row, column=3, value=name)
        c2 = ws_leg.cell(row=row, column=4, value=f"{cnt}次")
        fill = get_fill(name)
        c1.fill = fill
        c1.alignment = Alignment(horizontal="left", vertical="center")
        c2.alignment = Alignment(horizontal="center", vertical="center")
        c1.border = border
        c2.border = border

ws_leg.cell(row=1, column=5, value="★ 灰色=出现1-2次，不涂色").font = Font(italic=True, color="888888")

# ============ 保存 ============
out_path = Path("data/板块轮动Top10_v2_行业概念分开.xlsx")
wb.save(out_path)
print(f"已保存: {out_path}")
print(f"\n行业重复上榜: {len(repeat_hy)} 个")
print(f"概念重复上榜: {len(repeat_gn)} 个")
print(f"共分配颜色: {len(board_colors)} 种")
