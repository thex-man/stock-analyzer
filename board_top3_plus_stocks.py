"""
每天强势板块Top3 + 板块内涨幅>8%个股
=====================================
对于近10个交易日，每天取涨幅前3的板块，
从问财查这些板块的成分股（含当日涨跌幅），
筛选出个股当日涨幅>8%的，输出到Excel。

Excel格式：
  Sheet1: 每天一个区块，显示 Top3板块 + 板块内>8%个股
  Sheet2: 汇总——所有强势个股，按日期+涨幅排序
"""
import sys
sys.path.insert(0, '.')
from stock_data_source import wencai
import json
from pathlib import Path
from collections import defaultdict
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ============ 1. 读取板块历史数据，找每天Top3 ============
cache_files = sorted(Path('data/board_history_ths').glob('history_*.json'))
cache_file = cache_files[-1]  # 取最新缓存
with open(cache_file, encoding='utf-8') as f:
    raw = json.load(f)

all_dates = set()
for sym, info in raw.items():
    for row in info.get('data', []):
        if row.get('d'):
            all_dates.add(row['d'])

dates_sorted = sorted(all_dates, reverse=True)
last_10 = dates_sorted[:10]  # 最新→最旧

# 每天Top3板块: {date: [(pct, name, btype), ...]}
daily_top3 = {}
for date in last_10:
    day_results = []
    for sym, info in raw.items():
        name = info.get('name', sym)
        btype = info.get('type', '')
        for row in info.get('data', []):
            if row.get('d') == date:
                pct = row.get('p', 0)
                day_results.append((pct, name, btype))
                break
    day_results.sort(reverse=True)
    daily_top3[date] = day_results[:3]

print("【每日涨幅前三板块】")
for date in last_10:
    tops = daily_top3[date]
    print(f"\n{date}")
    for pct, name, btype in tops:
        print(f"  {pct:+.2f}%  {name} [{btype}]")

# ============ 2. 查每个Top3板块的成分股（带涨跌幅） ============
# 格式化日期：20260820 → 2026-08-20
def fmt_date(d):
    return f"{d[:4]}-{d[4:6]}-{d[6:]}"

# 缓存成分股数据，避免重复请求
member_cache = {}

def get_board_members_with_change(board_name, date_str):
    """查板块成分股及当日涨跌幅"""
    cache_key = (board_name, date_str)
    if cache_key in member_cache:
        return member_cache[cache_key]

    # 格式化日期用于问财
    ds = fmt_date(date_str)
    # 查询语句：板块名 + 日期 + 排序
    query = f'{board_name}板块 {ds} 按涨跌幅排序'

    try:
        result = wencai(query)
        df = result.get('datas')
        if df is None or (hasattr(df, 'empty') and df.empty):
            # Fallback：不加日期
            result2 = wencai(f'{board_name}板块成分股')
            df2 = result2.get('datas')
            if df2 is not None and not (hasattr(df2, 'empty') and df2.empty):
                df = df2

        if df is not None and not (hasattr(df, 'empty') and df.empty):
            stocks = []
            for _, row in df.iterrows():
                try:
                    change_col = '最新涨跌幅'
                    pct = row.get(change_col, None)
                    if pct is None:
                        pct = row.get('涨跌幅:前复权', None)
                    if pct is None:
                        continue
                    pct = float(pct)
                    if pct > 8.0:  # 只取>8%
                        code = str(row.get('股票代码', '')).strip()
                        name = str(row.get('股票简称', '')).strip()
                        if code and name:
                            stocks.append({
                                'code': code,
                                'name': name,
                                'change': round(pct, 2)
                            })
                except (ValueError, TypeError):
                    continue
            stocks.sort(key=lambda x: -x['change'])
            member_cache[cache_key] = stocks
        else:
            member_cache[cache_key] = []
    except Exception as e:
        print(f"    [WARN] {board_name} {ds} 查询失败: {e}")
        member_cache[cache_key] = []

    return member_cache[cache_key]

# ============ 3. 收集所有结果 ============
print("\n\n【抓取板块成分股（>8%个股）】")
all_results = []  # (date, board_pct, board_name, btype, stocks)

for date in last_10:
    ds_fmt = fmt_date(date)
    print(f"\n{date} ({ds_fmt})")
    for pct, board_name, btype in daily_top3[date]:
        stocks = get_board_members_with_change(board_name, date)
        if stocks:
            print(f"  {board_name}: {len(stocks)} 只涨幅>8%")
            for s in stocks:
                print(f"    {s['name']} {s['code']}  {s['change']:+.2f}%")
        else:
            print(f"  {board_name}: 无>8%个股")
        all_results.append((date, pct, board_name, btype, stocks))

# ============ 4. 输出到Excel ============
wb = Workbook()

# 调色板（每天不同底色）
DAY_COLORS = [
    'D6E4F0', 'EBF5FB', 'D5F5E3', 'FEF9E7', 'FDEDEC',
    'F4ECF7', 'E8F8F5', 'FDF2E9', 'F0F3FF', 'F9EBEA',
]

thin = Side(style='thin', color='DDDDDD')
border = Border(left=thin, right=thin, top=thin, bottom=thin)

# ----- Sheet1: 每天一个区块 -----
ws1 = wb.active
ws1.title = "每日Top3板块+强势个股"

ws1.column_dimensions['A'].width = 14
ws1.column_dimensions['B'].width = 10
ws1.column_dimensions['C'].width = 16
ws1.column_dimensions['D'].width = 12
ws1.column_dimensions['E'].width = 20
ws1.column_dimensions['F'].width = 14

row = 1

for day_idx, date in enumerate(reversed(last_10)):
    day_color = DAY_COLORS[day_idx % len(DAY_COLORS)]

    # 收集这一天3个板块的>8%个股
    day_boards = [(pct, name, btype, stocks)
                  for (d, pct, name, btype, stocks) in all_results if d == date]

    # 如果这一天没有任何>8%的个股，也要显示板块信息
    total_stocks = sum(len(s) for _, _, _, _, s in day_boards)

    # 日期标题行
    date_title = f"{fmt_date(date)}  涨幅前3板块"
    cell = ws1.cell(row=row, column=1, value=date_title)
    cell.font = Font(bold=True, size=12, color="FFFFFF")
    cell.fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
    cell.alignment = Alignment(horizontal="left", vertical="center")
    ws1.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
    row += 1

    # 板块表头
    for col, hdr in enumerate(["板块", "板块涨幅", "类型", "强势个股代码", "强势个股", "个股涨幅"], 1):
        c = ws1.cell(row=row, column=col, value=hdr)
        c.font = Font(bold=True, size=10, color="FFFFFF")
        c.fill = PatternFill(start_color="5B9BD5", end_color="5B9BD5", fill_type="solid")
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = border
    row += 1

    # 3个板块行
    for board_pct, board_name, btype, stocks in day_boards:
        if stocks:
            # 第一只股票填在同行
            c1 = ws1.cell(row=row, column=1, value=board_name)
            c2 = ws1.cell(row=row, column=2, value=f"{board_pct:+.2f}%")
            c3 = ws1.cell(row=row, column=3, value=btype)
            c4 = ws1.cell(row=row, column=4, value=stocks[0]['code'])
            c5 = ws1.cell(row=row, column=5, value=stocks[0]['name'])
            c6 = ws1.cell(row=row, column=6, value=f"{stocks[0]['change']:+.2f}%")
            for c in [c1, c2, c3, c4, c5, c6]:
                c.fill = PatternFill(start_color=day_color, end_color=day_color, fill_type="solid")
                c.border = border
                c.alignment = Alignment(horizontal="center", vertical="center")
            c2.alignment = Alignment(horizontal="right", vertical="center")
            c6.alignment = Alignment(horizontal="right", vertical="center")
            c6.font = Font(bold=True, color="C00000")

            # 其余股票填在后续行（合并板块/涨幅/类型列）
            for i in range(1, len(stocks)):
                row += 1
                # 合并前3列
                ws1.merge_cells(start_row=row, start_column=1, end_row=row, end_column=3)
                c4 = ws1.cell(row=row, column=4, value=stocks[i]['code'])
                c5 = ws1.cell(row=row, column=5, value=stocks[i]['name'])
                c6 = ws1.cell(row=row, column=6, value=f"{stocks[i]['change']:+.2f}%")
                for c in [c4, c5, c6]:
                    c.fill = PatternFill(start_color=day_color, end_color=day_color, fill_type="solid")
                    c.border = border
                    c.alignment = Alignment(horizontal="center", vertical="center")
                c6.alignment = Alignment(horizontal="right", vertical="center")
                c6.font = Font(bold=True, color="C00000")
        else:
            # 没有>8%个股的板块，也显示一行
            c1 = ws1.cell(row=row, column=1, value=board_name)
            c2 = ws1.cell(row=row, column=2, value=f"{board_pct:+.2f}%")
            c3 = ws1.cell(row=row, column=3, value=btype)
            c4 = ws1.cell(row=row, column=4, value="—")
            c5 = ws1.cell(row=row, column=5, value="无涨幅>8%个股")
            c6 = ws1.cell(row=row, column=6, value="—")
            for c in [c1, c2, c3, c4, c5, c6]:
                c.fill = PatternFill(start_color=day_color, end_color=day_color, fill_type="solid")
                c.border = border
                c.alignment = Alignment(horizontal="center", vertical="center")
            c2.alignment = Alignment(horizontal="right", vertical="center")
        row += 1

    row += 1  # 空一行分隔每天

ws1.freeze_panes = "A2"

# ----- Sheet2: 汇总所有强势个股 -----
ws2 = wb.create_sheet("强势个股汇总")

headers2 = ["日期", "所属板块", "板块涨幅", "个股代码", "个股简称", "个股涨幅"]
for col, hdr in enumerate(headers2, 1):
    c = ws2.cell(row=1, column=col, value=hdr)
    c.font = Font(bold=True, size=11, color="FFFFFF")
    c.fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
    c.alignment = Alignment(horizontal="center", vertical="center")
    c.border = border

# 收集所有强势个股并排序（按日期倒序+涨幅倒序）
all_stocks_flat = []
for date, board_pct, board_name, btype, stocks in all_results:
    for s in stocks:
        all_stocks_flat.append({
            'date': date,
            'board': board_name,
            'board_pct': board_pct,
            'code': s['code'],
            'name': s['name'],
            'stock_pct': s['change']
        })

all_stocks_flat.sort(key=lambda x: (-x['date'], -x['stock_pct']))

for row_idx, s in enumerate(all_stocks_flat, 2):
    values = [
        fmt_date(s['date']),
        s['board'],
        f"{s['board_pct']:+.2f}%",
        s['code'],
        s['name'],
        f"{s['stock_pct']:+.2f}%"
    ]
    day_idx = list(reversed(last_10)).index(s['date'])
    row_color = DAY_COLORS[day_idx % len(DAY_COLORS)]
    for col_idx, val in enumerate(values, 1):
        c = ws2.cell(row=row_idx, column=col_idx, value=val)
        c.fill = PatternFill(start_color=row_color, end_color=row_color, fill_type="solid")
        c.border = border
        c.alignment = Alignment(horizontal="center", vertical="center")
        if col_idx == 6:
            c.font = Font(bold=True, color="C00000")

ws2.column_dimensions['A'].width = 14
ws2.column_dimensions['B'].width = 18
ws2.column_dimensions['C'].width = 10
ws2.column_dimensions['D'].width = 16
ws2.column_dimensions['E'].width = 14
ws2.column_dimensions['F'].width = 12
ws2.freeze_panes = "A2"

# ----- Sheet3: 统计每个个股出现的次数 -----
ws3 = wb.create_sheet("强势个股出现次数")
from collections import Counter

name_count = Counter([s['name'] for s in all_stocks_flat])
top_names = sorted(name_count.items(), key=lambda x: -x[1])

ws3.cell(row=1, column=1, value="个股").font = Font(bold=True)
ws3.cell(row=1, column=2, value="出现次数").font = Font(bold=True)
ws3.cell(row=1, column=3, value="最近一次涨幅").font = Font(bold=True)
ws3.cell(row=1, column=4, value="历史涨幅").font = Font(bold=True)

for col in range(1, 5):
    ws3.cell(row=1, column=col).fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
    ws3.cell(row=1, column=col).font = Font(bold=True, color="FFFFFF")
    ws3.cell(row=1, column=col).alignment = Alignment(horizontal="center", vertical="center")
    ws3.cell(row=1, column=col).border = border

for r, (name, cnt) in enumerate(top_names, 2):
    s_all = [s for s in all_stocks_flat if s['name'] == name]
    changes = [s['stock_pct'] for s in s_all]
    ws3.cell(row=r, column=1, value=name).alignment = Alignment(horizontal="left", vertical="center")
    ws3.cell(row=r, column=2, value=f"{cnt}次").alignment = Alignment(horizontal="center", vertical="center")
    ws3.cell(row=r, column=3, value=f"{changes[0]:+.2f}%").alignment = Alignment(horizontal="right", vertical="center")
    ws3.cell(row=r, column=4, value=" / ".join([f"{c:+.1f}%" for c in changes])).alignment = Alignment(horizontal="left", vertical="center")
    if cnt >= 3:
        fill_color = "FFE0E0"
    elif cnt == 2:
        fill_color = "FFFACD"
    else:
        fill_color = "FFFFFF"
    for col in range(1, 5):
        ws3.cell(row=r, column=col).fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")
        ws3.cell(row=r, column=col).border = border

ws3.column_dimensions['A'].width = 14
ws3.column_dimensions['B'].width = 10
ws3.column_dimensions['C'].width = 14
ws3.column_dimensions['D'].width = 30

# ============ 保存 ============
out_path = Path("data/每日Top3板块强势个股.xlsx")
wb.save(out_path)
print(f"\n已保存: {out_path}")
print(f"强势个股总数: {len(all_stocks_flat)}")
