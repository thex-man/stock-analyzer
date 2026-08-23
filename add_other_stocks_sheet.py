"""
Sheet5: 每日涨幅>6%但不在当日Top3板块内的个股
=============================================
用 pywencai 直接查询"当日涨幅超过6%"的股票，再剔除当日 Top3 板块成分股。

优势：
  - 一次请求拿到所有>6%涨幅的股票，无需遍历全A股（11099只）
  - 速度快（秒级），不受 baostock 限速影响

依赖：
  - pywencai（pip install pywencai）
  - iwencai 服务可达（部分网络环境会被 captcha 拦截，需手动复制 cookie）
"""
import sys
sys.path.insert(0, '.')
import json
from pathlib import Path
from collections import defaultdict
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import time

try:
    import pywencai
except ImportError:
    print("缺少 pywencai，请运行: pip install pywencai")
    sys.exit(1)

# ============ 1. 读取板块历史 ============
cache_file = Path('data/board_history_ths/history_20260728_20260820.json')
if not cache_file.exists():
    print(f"历史缓存不存在: {cache_file}")
    sys.exit(1)

with open(cache_file, encoding='utf-8') as f:
    raw = json.load(f)

# 收集所有日期
all_dates = set()
for sym, info in raw.items():
    for row in info.get('data', []):
        if row.get('d'):
            all_dates.add(row['d'])
dates_sorted = sorted(all_dates, reverse=True)
last_10 = dates_sorted[:10]
print(f"历史日期范围: {last_10[-1]} ~ {last_10[0]}")

# 每天 Top3 板块（行业 + 概念合并）
def get_top3_for_date(date):
    """返回 date 当天的 Top3 行业+概念板块集合"""
    day_results = []
    for sym, info in raw.items():
        name = info.get('name', sym)
        btype = info.get('type', '')
        for row in info.get('data', []):
            if row.get('d') == date:
                pct = row.get('p', 0)
                day_results.append((pct, name, btype))
                break
    day_results.sort(reverse=True)
    # 行业取Top3 + 概念取Top3
    by_type = defaultdict(list)
    for pct, name, btype in day_results:
        by_type[btype].append((pct, name))
    top3 = set()
    for t in by_type:
        for _, name in by_type[t][:3]:
            top3.add(name)
    return top3

# 累计 Top3 板块（10日合集，用于查成分股）
all_top3_boards = set()
for date in last_10:
    all_top3_boards |= get_top3_for_date(date)
print(f"近10日Top3板块合计: {len(all_top3_boards)} 个")

# ============ 2. 获取Top3板块的成分股映射 ============
print("\n查询 Top3 板块的成分股...")
# 用 pywencai 查成分股：<板块名> 成分股
top3_constituents = {}  # {board_name: set(codes)}

for i, board in enumerate(sorted(all_top3_boards)):
    try:
        # pywencai 问句：{板块名} 成分股
        df = pywencai.get(
            question=f"{board} 成份股 股票代码",
            perpage=100,
            retry=2,
            delay=1,
        )
        if df is not None and not df.empty:
            # 找代码列
            code_col = None
            for col in df.columns:
                if '代码' in col or 'code' in col.lower():
                    code_col = col
                    break
            if code_col:
                codes = set()
                for v in df[code_col].dropna().astype(str):
                    v = v.strip()
                    if v.isdigit() and len(v) == 6:
                        codes.add(v)
                    elif '.' in v:  # sh.600000 格式
                        codes.add(v.split('.')[-1])
                top3_constituents[board] = codes
                if i < 3:
                    print(f"  {board}: {len(codes)} 只成分股")
        if (i + 1) % 5 == 0:
            print(f"  进度 {i+1}/{len(all_top3_boards)}")
        time.sleep(0.3)
    except Exception as e:
        if i < 3:
            print(f"  {board}: 查询失败 - {str(e)[:100]}")
        continue

all_top3_codes = set()
for codes in top3_constituents.values():
    all_top3_codes |= codes
print(f"\nTop3 板块成分股合计: {len(all_top3_codes)} 只")

# ============ 3. pywencai 查询"今日涨幅>6%" ============
print("\n查询当日涨幅>6%的股票...")
target_date = last_10[0]  # 最新一天
print(f"目标日期: {target_date}")

try:
    df = pywencai.get(
        question=f"{target_date[:4]}-{target_date[4:6]}-{target_date[6:]} 涨幅超过6% 股票代码 股票名称 涨跌幅 所属概念",
        perpage=500,
        loop=True,
        retry=3,
        delay=2,
    )
except Exception as e:
    print(f"pywencai查询失败: {e}")
    sys.exit(1)

if df is None or df.empty:
    print("当日无涨幅>6%的股票（或 pywencai 不可用）")
    df = None
else:
    print(f"拿到 {len(df)} 只股票")

# ============ 4. 筛选非 Top3 板块成分股 ============
if df is not None:
    # 找出代码列
    code_col = None
    name_col = None
    pct_col = None
    for col in df.columns:
        if '代码' in col or 'code' in col.lower():
            code_col = col
        if '名称' in col or 'name' in col.lower():
            name_col = col
        if '涨跌幅' in col or '涨幅' in col or 'pct' in col.lower():
            pct_col = col

    if code_col is None:
        print(f"找不到代码列, 列名: {df.columns.tolist()}")
        df = None
    else:
        print(f"代码列: {code_col}, 名称列: {name_col}, 涨幅列: {pct_col}")

# ============ 5. 写入 Excel ============
print("\n写入Excel...")
out_path = 'data/板块轮动Top10_v4_含非Top3强势个股.xlsx'
wb = load_workbook('data/板块轮动Top10_v3_含每日Top3强势个股.xlsx')

if '非Top3板块强势个股' in wb.sheetnames:
    del wb['非Top3板块强势个股']
ws = wb.create_sheet('非Top3板块强势个股')

thin = Side(style='thin', color='CCCCCC')
border = Border(left=thin, right=thin, top=thin, bottom=thin)

def fmt_date(d):
    return f"{d[:4]}-{d[4:6]}-{d[6:]}"

col_widths = [14, 14, 18, 14, 30, 14, 14]
col_headers = ['日期', '个股代码', '个股简称', '涨幅%', '所属概念', '板块涨幅', '板块类型']
for i, (w, h) in enumerate(zip(col_widths, col_headers), 1):
    ws.column_dimensions[get_column_letter(i)].width = w
    c = ws.cell(row=1, column=i, value=h)
    c.font = Font(bold=True, color='FFFFFF', size=11)
    c.fill = PatternFill(start_color='2F5496', end_color='2F5496', fill_type='solid')
    c.alignment = Alignment(horizontal='center', vertical='center')
    c.border = border

ws.freeze_panes = 'A2'

row_num = 2
total_stocks = 0

if df is not None:
    # 把 df 转 list 再过滤
    records = df.to_dict('records')
    for rec in records:
        code_raw = str(rec.get(code_col, '')).strip()
        # 统一代码格式（去掉 sh./sz. 前缀）
        if '.' in code_raw:
            code = code_raw.split('.')[-1]
        else:
            code = code_raw

        # 不在 Top3 板块成分股里
        if code in all_top3_codes:
            continue

        name = str(rec.get(name_col, '')).strip() if name_col else ''
        pct_val = rec.get(pct_col, 0) if pct_col else 0
        try:
            pct = float(pct_val)
        except (ValueError, TypeError):
            pct = 0

        # 所属概念列
        concepts = str(rec.get('所属概念', '') or rec.get('概念', '')).strip()

        # 找所属板块当日涨幅
        max_board_pct = 0
        max_board_name = ''
        for board_name in all_top3_boards:
            if board_name in concepts:
                # 找该板块当日涨幅
                for sym, info in raw.items():
                    if info.get('name') == board_name:
                        for row in info.get('data', []):
                            if row.get('d') == target_date:
                                max_board_pct = row.get('p', 0)
                                max_board_name = board_name
                                break
                        break

        for col_i, val in enumerate([
            fmt_date(target_date), code, name,
            f"{pct:+.2f}%", concepts,
            f"{max_board_pct:+.2f}%" if max_board_pct else '',
            max_board_name or '其他',
        ], 1):
            c = ws.cell(row=row_num, column=col_i, value=val)
            c.border = border
            c.alignment = Alignment(horizontal='center', vertical='center')
            if col_i == 4:  # 涨幅列
                c.font = Font(bold=True, color='C00000')
                c.alignment = Alignment(horizontal='right', vertical='center')
        row_num += 1
        total_stocks += 1

    print(f"  {fmt_date(target_date)}: 过滤后 {total_stocks} 只非Top3板块>6%个股")

if total_stocks == 0:
    # 写一行说明
    c_date = ws.cell(row=2, column=1, value=fmt_date(target_date))
    c_date.font = Font(bold=True)
    c_note = ws.cell(row=2, column=3, value='无>6%个股（不在Top3板块），或 pywencai 数据源不可用')
    for col_i in range(1, 8):
        ws.cell(row=2, column=col_i).border = border
    c_date.alignment = Alignment(horizontal='center', vertical='center')

wb.save(out_path)
print(f"\n已保存: {out_path}")
print(f"Sheet5 共 {total_stocks} 条")
