# -*- coding: utf-8 -*-
"""
生成 MACD 5日>10% 最近 10 个交易日滚动表
================================================
数据源：Sheet6_v2「MACD强势个股_v2」（sheet6_macd_roll_10d.py 生成的滚动截面）
输出：Sheet6_v2「MACD5日_最近10日」+ HTML tab 友好格式

列：信号日 / 代码 / 名称 / MACD / 5日涨幅% / 信号日→今日% / 缠论分数 / 当前价
排序：信号日降序，每天内按 信号日→今日% 降序

调用：0 次（只读 Sheet6_v2）
"""
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from pathlib import Path

EXCEL = r'D:\stock\tool\stock\data\板块轮动Top10_v4_含非Top3强势个股.xlsx'
NEW_SHEET = 'MACD5日_最近10日'
N_DAYS = 10


def main():
    wb = openpyxl.load_workbook(EXCEL)

    # 1. 读 Sheet6_v2
    src_name = 'MACD强势个股_v2'
    if src_name not in wb.sheetnames:
        print(f'[ERR] {src_name} 不存在')
        return
    src = wb[src_name]
    rows = []
    for r in range(2, src.max_row + 1):
        rows.append([src.cell(row=r, column=c).value for c in range(1, 16)])
    print(f'[*] 从 {src_name} 读 {len(rows)} 行')

    # 2. 取最近 N 个交易日
    days = sorted({r[0] for r in rows if r[0]}, reverse=True)[:N_DAYS]
    print(f'[*] 最近 {N_DAYS} 个交易日: {days[::-1]}')

    # 3. 过滤 + 排序
    filtered = [r for r in rows if r[0] in set(days)]
    # 排序：信号日降序（主键），同一天内 信号日→今日% 降序（次键）
    filtered.sort(key=lambda r: r[7] if r[7] is not None else -999, reverse=True)  # 先按涨跌幅降序
    filtered.sort(key=lambda r: r[0], reverse=True)  # 再按信号日降序（Python sort 稳定）
    print(f'[*] 过滤后 {len(filtered)} 行')

    # 4. 写入新 sheet
    if NEW_SHEET in wb.sheetnames:
        del wb[NEW_SHEET]
    ws = wb.create_sheet(NEW_SHEET)

    # 标题
    title_text = f'MACD 5日>10% 最近 {N_DAYS} 个交易日（{days[-1]} ~ {days[0]}，共 {len(filtered)} 条信号）'
    ws.cell(row=1, column=1, value=title_text)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=9)
    ws.cell(row=1, column=1).font = Font(bold=True, size=12, color='FFFFFF')
    ws.cell(row=1, column=1).fill = PatternFill(start_color='1e2d4a', end_color='1e2d4a', fill_type='solid')
    ws.cell(row=1, column=1).alignment = Alignment(horizontal='center', vertical='center')

    # 表头
    headers = ['信号日', '代码', '名称', '当前价', 'MACD', '5日涨幅%', '信号日→今日%', '缠论分数', '备注']
    for c, h in enumerate(headers, 1):
        cell = ws.cell(row=2, column=c, value=h)
        cell.font = Font(bold=True, color='FFFFFF')
        cell.fill = PatternFill(start_color='2F5496', end_color='2F5496', fill_type='solid')
        cell.alignment = Alignment(horizontal='center', vertical='center')

    # 数据（按信号日分组，组与组之间空一行）
    thin = Side(style='thin', color='CCCCCC')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    last_day = None
    r_out = 3
    for r in filtered:
        d, _, code, name, price, macd, g5d, g2t, score = r[:9]
        note = r[14] or ''
        # 分组空行
        if d != last_day and last_day is not None:
            r_out += 1
        last_day = d
        # 写行
        row_data = [d, code, name, price, macd, g5d, g2t, score, note]
        for c, v in enumerate(row_data, 1):
            cell = ws.cell(row=r_out, column=c, value=v)
            cell.border = border
            cell.alignment = Alignment(horizontal='center', vertical='center')
            if c in (4, 5, 6, 7, 8):
                if isinstance(v, (int, float)):
                    cell.number_format = '0.00'
            # 信号日→今日% 涂色
            if c == 7 and isinstance(v, (int, float)):
                if v >= 20:
                    cell.fill = PatternFill(start_color='FFD700', end_color='FFD700', fill_type='solid')
                elif v >= 10:
                    cell.fill = PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid')
                elif v < 0:
                    cell.fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
        r_out += 1

    # 列宽
    widths = [12, 10, 14, 10, 10, 14, 16, 12, 30]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # 冻结
    ws.freeze_panes = 'A3'

    # 排序：把新 sheet 放到 Sheet6 之后
    desired_after = 'MACD强势个股'
    if desired_after in wb.sheetnames:
        idx = wb.sheetnames.index(desired_after) + 1
        sheets = wb._sheets
        new_sheet_obj = wb[NEW_SHEET]
        sheets.remove(new_sheet_obj)
        sheets.insert(idx, new_sheet_obj)

    wb.save(EXCEL)
    print(f'\n[SAVE] {NEW_SHEET} ({len(filtered)} 行)')


if __name__ == '__main__':
    main()
