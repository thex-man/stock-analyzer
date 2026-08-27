"""
Sheet5: 每日涨幅>6%但不在当日Top3板块内的个股
==============================================
v5 修复（2026-08-26）：
  - 处理全部10天（恢复原始版本逻辑）
  - 列名：涨跌幅:前复权 → 涨跌幅（已修）
  - 日期转换：缓存日期（系统日期）→ Excel显示日期（实际交易日）
  - 从v3 Excel读取Top3强势股代码并排除
"""
import sys
sys.path.insert(0, '.')
import json
from pathlib import Path
from datetime import datetime, time as dtime
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

CONCEPT_DATA_DIR = r'D:\stock\tool\stock\concept_data'

# ========== 1. 日期转换 ==========
now = datetime.now()
cutoff = dtime(9, 15)
current_date_str = now.strftime('%Y%m%d')  # 20260826

if now.time() < cutoff:
    import datetime as dt
    trade_ts = now.replace(hour=0, minute=0, second=0, microsecond=0).timestamp() - 86400
    trade_date_str = dt.datetime.fromtimestamp(trade_ts).strftime('%Y%m%d')
else:
    trade_date_str = current_date_str

def cache_date_to_excel_date(cache_date_str):
    """缓存key（系统日期）→ Excel显示日期（实际交易日）"""
    if now.time() < cutoff and cache_date_str == current_date_str:
        import datetime as dt
        trade_ts = now.replace(hour=0, minute=0, second=0, microsecond=0).timestamp() - 86400
        return dt.datetime.fromtimestamp(trade_ts).strftime('%Y-%m-%d')
    else:
        return f"{cache_date_str[:4]}-{cache_date_str[4:6]}-{cache_date_str[6:]}"

def fmt_date(d):
    return f"{d[:4]}-{d[4:6]}-{d[6:]}"

# ============ 2. 加载概念数据缓存 ============
def load_all_concepts():
    print('[*] 加载概念数据...')
    cache = {}
    for fpath in Path(CONCEPT_DATA_DIR).glob('*_concepts.json'):
        code = fpath.stem.replace('_concepts', '')
        try:
            with open(fpath, 'r', encoding='utf-8') as f:
                data = json.load(f)
            concepts = [c['name'] for c in data.get('concepts', [])]
            cache[code] = concepts
        except Exception:
            cache[code] = []
    print(f'[*] 概念数据: {len(cache)} 只')
    return cache

concept_cache = load_all_concepts()

def get_concepts(code):
    clean = code.zfill(6)
    concepts = concept_cache.get(clean, [])
    return '|'.join(concepts) if concepts else ''

# ============ 3. 读取板块历史缓存（获取Top3板块名）============
cache_files = sorted(Path('data/board_history_ths').glob('history_*.json'))
cache_file = cache_files[-1]
with open(cache_file, encoding='utf-8') as f:
    raw = json.load(f)

all_dates = set()
for sym, info in raw.items():
    for row in info.get('data', []):
        if row.get('d'):
            all_dates.add(row['d'])

dates_sorted = sorted(all_dates, reverse=True)
last_10 = dates_sorted[:10]

daily_top3 = {}
for date in last_10:
    day_results = []
    for sym, info in raw.items():
        name = info.get('name', sym)
        for row in info.get('data', []):
            if row.get('d') == date:
                pct = row.get('p', 0)
                day_results.append((pct, name))
                break
    day_results.sort(reverse=True)
    daily_top3[date] = {name for _, name in day_results[:3]}

board_pct_map = {}
for date in last_10:
    board_pct_map[date] = {}
    for sym, info in raw.items():
        for row in info.get('data', []):
            if row.get('d') == date:
                board_pct_map[date][info.get('name', sym)] = row.get('p', 0)
                break

# ============ 4. 从v3 Excel读取Top3强势个股代码 ============
def load_top3_codes_from_v3(v3_path, target_date_str):
    """读取v3 Excel中某日期的Top3强势股代码集合
    target_date_str: 已经是显示格式 'YYYY-MM-DD'（不是8位缓存key）"""
    # target_date_str 传入时就是 '2026-08-25' 格式，不需要再 fmt_date
    date_match = target_date_str  # 直接是 'YYYY-MM-DD' 格式
    wb = load_workbook(v3_path, data_only=True)
    ws = wb['每日Top3强势个股']
    codes = set()
    collecting = False
    for row in ws.iter_rows(values_only=True):
        date_val = row[0]
        stock_cell = row[4]
        # Excel日期可能是 datetime 对象，格式化为 YYYY-MM-DD
        if date_val is not None:
            try:
                date_str = date_val.strftime('%Y-%m-%d')
            except AttributeError:
                date_str = str(date_val)[:10]
        else:
            date_str = ''
        if date_val is not None and date_str == date_match:
            collecting = True
            if stock_cell:
                parts = str(stock_cell).split()
                if len(parts) >= 2:
                    codes.add(parts[1].strip())
        elif collecting:
            if date_val is not None:
                break
            if stock_cell:
                parts = str(stock_cell).split()
                if len(parts) >= 2:
                    codes.add(parts[1].strip())
    return codes

# ============ 5. 问财查询涨幅>6%（修列名）============
def _save_snapshot(df):
    """把问财返回列中的快照字段写入 stock_snapshot_daily（幂等：先删后插）"""
    import duckdb
    from datetime import datetime as _dt

    def _num(v):
        """'5.26 亿'/'1,515.18'/'30.56' -> float"""
        if v is None:
            return None
        s = str(v).replace(',', '').strip()
        mult = 1.0
        if s.endswith('亿'):
            mult, s = 1e8, s[:-1]
        elif s.endswith('万'):
            mult, s = 1e4, s[:-1]
        try:
            return round(float(s) * mult, 4)
        except ValueError:
            return None

    # ⚠️ 问财列错位问题：仅 最新价/涨跌幅 已验证可靠，其余字段置 NULL
    # 待用参考源交叉验证后放开（TODO 2026-08-28）
    rows = []
    seen = set()
    ts = _dt.now()
    d = trade_date_str[:4] + '-' + trade_date_str[4:6] + '-' + trade_date_str[6:]
    for _, row in df.iterrows():
        code = str(row.get('股票代码', '')).strip()
        if not code or code in seen:
            continue
        seen.add(code)
        rows.append((code, d, _num(row.get('最新价')), _num(row.get('涨跌幅')),
                     None, None, None, None, None, ts))
    if not rows:
        return
    con = duckdb.connect(r'D:\stock\tool\stock\data\stock.duckdb')
    con.execute("""CREATE TABLE IF NOT EXISTS stock_snapshot_daily (
        stock_code VARCHAR, date DATE, close DOUBLE, pct DOUBLE,
        float_mv DOUBLE, pe DOUBLE, turnover DOUBLE, volume_ratio DOUBLE,
        amount DOUBLE, fetch_time TIMESTAMP,
        PRIMARY KEY (stock_code, date))""")
    con.execute('DELETE FROM stock_snapshot_daily WHERE date=?', [d])
    con.executemany('INSERT INTO stock_snapshot_daily VALUES (?,?,?,?,?,?,?,?,?,?)', rows)
    con.close()
    print(f'    [SNAPSHOT] {d} 入库 {len(rows)} 只')


def fetch_gainers_wenqi():
    """用问财查涨幅>6%个股（今日数据，不带日期）"""
    from stock_data_source import wencai

    query = '涨幅超过6%'
    try:
        result = wencai(query, perpage=200)
    except Exception as e:
        print(f"    [WARN] 问财失败: {e}")
        return []

    df = result.get('datas')
    if df is None or (hasattr(df, 'empty') and df.empty):
        return []

    # v2.1: 顺手入库每日快照（流通市值/PE/换手率/量比，0 额外调用）
    try:
        _save_snapshot(df)
    except Exception as e:
        print(f'    [WARN] 快照入库失败(不影响主流程): {e}')

    gainers = []
    for _, row in df.iterrows():
        try:
            pct = row.get('涨跌幅')  # 修复：不是涨跌幅:前复权
            if pct is None:
                continue
            pct = float(pct)
            if pct > 6.0:
                code = str(row.get('股票代码', '')).strip()
                name = str(row.get('股票简称', '')).strip()
                board_raw = str(row.get('所属板块', '')).strip()
                board = board_raw.split('-')[-1] if board_raw else ''
                gainers.append((code, name, round(pct, 2), board))
        except (ValueError, TypeError):
            continue

    gainers.sort(key=lambda x: -x[2])
    return gainers

# ============ 6. 主循环（全部10天）============
v3_path = 'data/板块轮动Top10_v3_含每日Top3强势个股.xlsx'
v4_path = 'data/板块轮动Top10_v4_含非Top3强势个股.xlsx'

print(f'只处理今日（{trade_date_str}）')
print(f'9:15前: {"是" if now.time() < cutoff else "否"}')
print('=' * 50)

# 问财查今日涨幅>6%
print(f'问财查询 今日涨幅>6%...')
today_gainers = fetch_gainers_wenqi()
print(f'今日>6%: {len(today_gainers)} 只')

# 关闭 Selenium
try:
    from stock_data_source import _close_selenium_driver
    _close_selenium_driver()
except:
    pass

# 今日在缓存中的 key（凌晨跑时缓存 key 是上一交易日 = trade_date_str）
cache_today = trade_date_str if trade_date_str in daily_top3 else current_date_str
if cache_today not in daily_top3:
    print(f'今日 {cache_today} 不在缓存中，跳过')
    daily_other_stocks = {}
else:
    excel_date = cache_date_to_excel_date(cache_today)
    top3_codes = load_top3_codes_from_v3(v3_path, excel_date)
    top3_boards = daily_top3.get(cache_today, set())
    print(f'Top3板块: {top3_boards}')
    print(f'Top3强势股: {len(top3_codes)}只 {top3_codes}')

    # 排除 Top3 强势个股
    non_top3 = [(code, name, pct, board)
                for code, name, pct, board in today_gainers
                if code not in top3_codes]
    non_top3.sort(key=lambda x: -x[2])
    daily_other_stocks = {cache_today: (excel_date, non_top3)}
    print(f'非Top3: {len(non_top3)} 只')

# ============ 7. 写入 Excel ============
print('\n写入 Excel...')

# 修复：必须加载 v4 而不是 v3，否则会清空 v4 的 MACD v2 sheets！
wb_out = load_workbook(v4_path)

if '非Top3板块强势个股' in wb_out.sheetnames:
    del wb_out['非Top3板块强势个股']

ws = wb_out.create_sheet('非Top3板块强势个股')

DAY_COLORS = ['D6E4F0', 'EBF5FB', 'D5F5E3', 'FEF9E7', 'FDEDEC',
              'F4ECF7', 'E8F8F5', 'FDF2E9', 'F0F3FF', 'F9EBEA']

thin = Side(style='thin', color='CCCCCC')
border = Border(left=thin, right=thin, top=thin, bottom=thin)

col_widths = [14, 12, 16, 40, 12, 12]
col_headers = ['日期', '个股代码', '个股简称', '所属概念', '板块涨幅', '个股涨幅']

for i, (w, h) in enumerate(zip(col_widths, col_headers), 1):
    ws.column_dimensions[get_column_letter(i)].width = w
    c = ws.cell(row=1, column=i, value=h)
    c.font = Font(bold=True, color='FFFFFF', size=11)
    c.fill = PatternFill(start_color='2F5496', end_color='2F5496', fill_type='solid')
    c.alignment = Alignment(horizontal='center', vertical='center')
    c.border = border

ws.freeze_panes = 'A2'

row_num = 2
total_stocks = 0

for day_idx, date in enumerate(reversed(last_10)):
    day_color = DAY_COLORS[day_idx % len(DAY_COLORS)]
    excel_date = cache_date_to_excel_date(date)

    # 只处理今日（其他日期问财无历史数据）
    if date in daily_other_stocks:
        _, items = daily_other_stocks[date]
    else:
        items = []

    if items:
        for code, name, pct, board in items:
            board_pct_val = 0
            for bname, bpct in board_pct_map.get(date, {}).items():
                if bname in board or board in bname:
                    board_pct_val = bpct
                    break

            concepts = get_concepts(code) or board  # 优先本地概念缓存，否则用问财返回的所属板块
            row_data = [excel_date, code, name, concepts,
                        f'{board_pct_val:+.2f}%', f'{pct:+.2f}%']
            for col_i, val in enumerate(row_data, 1):
                c = ws.cell(row=row_num, column=col_i, value=val)
                c.fill = PatternFill(start_color=day_color, end_color=day_color, fill_type='solid')
                c.border = border
                c.alignment = Alignment(horizontal='center', vertical='center')
                if col_i in (5, 6):
                    c.alignment = Alignment(horizontal='right', vertical='center')
                    c.font = Font(bold=True, color='C00000')
                if col_i == 4 and val:
                    c.font = Font(color='1F497D', size=9)
                    c.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
            row_num += 1
            total_stocks += 1
    else:
        c_date = ws.cell(row=row_num, column=1, value=excel_date)
        c_date.font = Font(bold=True)
        ws.cell(row=row_num, column=3, value='无>6%个股（不在Top3板块）')
        for col_i in range(1, 7):
            ws.cell(row=row_num, column=col_i).fill = PatternFill(start_color=day_color, end_color=day_color, fill_type='solid')
            ws.cell(row=row_num, column=col_i).border = border
        c_date.alignment = Alignment(horizontal='center', vertical='center')
        row_num += 1

out_path = v4_path
wb_out.save(out_path)
print(f'已保存: {out_path}')
days_with_data = len([d for d in daily_other_stocks])
print(f'Sheet5 共 {total_stocks} 条（{days_with_data} 天有数据，其他 {len(last_10)-days_with_data} 天无数据）')
