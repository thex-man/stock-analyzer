# -*- coding: utf-8 -*-
"""
Sheet6: 创业板 MACD>0 且近5日涨幅>10%
用 stock_data_source.wencai 查数据
缠论结构打分排序
"""
import sys, os
sys.path.insert(0, r'D:\stock\tool\stock')
from stock_data_source import wencai, get_kline
import pandas as pd
import numpy as np
import openpyxl
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ===== 1. 用问财查创业板 MACD>0 且近5日涨幅>10% =====
print('===== Step1: 查询问财 =====')
result = wencai('创业板 MACD大于0 近5日涨幅大于10%', perpage=200)
if 'error' in result:
    print('Error:', result['error'])
    sys.exit(1)

df = result['datas']
print(f'问财返回 {len(df)} 只')

# 提取股票代码
stocks = []
for _, row in df.iterrows():
    code_raw = str(row.get('股票代码', ''))
    # 统一格式为 sz300xxx
    code6 = code_raw.replace('.SZ', '').replace('.SH', '').replace('sz', '').replace('sh', '').strip()
    if not code6.startswith('300'):
        continue
    name = str(row.get('股票简称', code6))
    macd_val = row.get('macd(macd值)', 0)
    gain_raw = row.get('区间涨跌幅:前复权', row.get('最新涨跌幅', 0))
    try:
        gain = float(str(gain_raw).replace('%', ''))
    except:
        gain = 0

    stocks.append({
        'code': code6,
        'name': name,
        'macd': float(macd_val) if macd_val else 0,
        'gain_5d': gain,
    })

print(f'提取到 {len(stocks)} 只创业板股票')
print('前5只:', [(s['code'], s['name'], s['macd'], s['gain_5d']) for s in stocks[:5]])

# ===== 2. 获取K线数据 =====
print('\n===== Step2: 获取K线 =====')
START = '20260601'
END = '20260821'

def safe_kline(code6):
    """获取单只K线，失败返回None"""
    code_fmt = 'sz' + code6
    try:
        df_k = get_kline(code_fmt, start=START, end=END, adjust='qfq')
        if df_k is None or len(df_k) < 30:
            return None
        # 确保列名是标准的
        df_k.columns = [c.strip() for c in df_k.columns]
        return df_k
    except Exception as e:
        return None

# ===== 3. 缠论打分函数 =====
def chan_score(df_k):
    """
    缠论结构打分（简化版）
    评分维度：
    - 趋势（20日斜率）: +2/-2
    - 中枢位置（当前价在中枢上方/下方/内部）: +2/-2/~0
    - 底分型: +1.5
    - 底背驰（MACD柱面积放大）: +1.5
    - 放量配合: +0.5
    - 顶分型/顶背驰: 扣分
    """
    closes = df_k['close'].values.astype(float)
    highs = df_k['high'].values.astype(float)
    lows = df_k['low'].values.astype(float)
    volumes = df_k['volume'].values.astype(float) if 'volume' in df_k.columns else np.zeros(len(closes))
    n = len(closes)
    score = 0.0
    details = {}

    # 趋势（20日斜率）
    if n >= 20:
        slope = (closes[-1] / closes[-20] - 1) * 100
        if slope > 8:
            score += 2; trend = f'强势上升({slope:+.1f}%)'
        elif slope > 3:
            score += 1; trend = f'缓慢上升({slope:+.1f}%)'
        elif slope < -8:
            score -= 3; trend = f'下降({slope:+.1f}%)'
        elif slope < -3:
            score -= 1; trend = f'缓慢下降({slope:+.1f}%)'
        else:
            trend = f'横盘({slope:+.1f}%)'
        details['trend'] = trend
    else:
        details['trend'] = '数据不足'

    # 中枢（最近20日）
    look = min(20, n)
    zg = float(highs[-look:].max())
    zd = float(lows[-look:].min())
    z_width = zg - zd if zg > zd else 1
    current = float(closes[-1])

    if current > zg:
        score += 2
        position = '中枢上方强势'
        leave = (current - zg) / z_width
        if leave > 1.5: score += 1
    elif current < zd:
        score -= 2; position = '中枢下方弱势'
    else:
        position = '中枢内部震荡'
    details['zg'] = round(zg, 2)
    details['zd'] = round(zd, 2)
    details['position'] = position

    # 分型（最近5根K线）
    if n >= 5:
        c = closes
        # 简化底分型: 中间一根最低
        is_bottom = (c[-3] < c[-2] and c[-3] < c[-4] and c[-2] > c[-1] and c[-2] > c[-5])
        is_top = (c[-3] > c[-2] and c[-3] > c[-4] and c[-2] < c[-1] and c[-2] < c[-5])
        if is_bottom:
            score += 1.5; details['fx'] = '底分型'
        elif is_top:
            score -= 1; details['fx'] = '顶分型'
        else:
            details['fx'] = '无分型'
    else:
        details['fx'] = '数据不足'

    # MACD背驰（MACD柱面积比较）
    try:
        ema12_s = pd.Series(closes).ewm(span=12, adjust=False).mean()
        ema26_s = pd.Series(closes).ewm(span=26, adjust=False).mean()
        macd_s = ema12_s - ema26_s
        sig_s = pd.Series(macd_s).ewm(span=9, adjust=False).mean()
        hist_s = macd_s - sig_s
        recent_area = float(hist_s.iloc[-5:].sum())
        prev_area = float(hist_s.iloc[-10:-5].sum()) if len(hist_s) >= 10 else 0
        if recent_area > 0 and prev_area > 0 and recent_area > prev_area * 1.3:
            score += 1.5; details['bcie'] = '底背驰(+1.5)'
        elif recent_area < 0 and prev_area < 0 and abs(recent_area) > abs(prev_area) * 1.3:
            score -= 1.5; details['bcie'] = '顶背驰(-1.5)'
        else:
            details['bcie'] = '无背驰'
    except Exception as e:
        details['bcie'] = 'N/A'

    # 成交量配合
    if n >= 20:
        vol5 = float(volumes[-5:].mean())
        vol20 = float(volumes[-20:].mean())
        if vol5 > vol20 * 1.3 and score > 0:
            score += 0.5; details['vol'] = '放量配合'
        elif vol5 < vol20 * 0.7:
            details['vol'] = '缩量'
        else:
            details['vol'] = '量能正常'
    else:
        details['vol'] = '数据不足'

    details['total_score'] = round(score, 1)
    return score, details

# ===== 4. 遍历获取K线 + 缠论打分 =====
print('\n===== Step3: 缠论分析 =====')
final_results = []
for i, st in enumerate(stocks):
    code6 = st['code']
    df_k = safe_kline(code6)
    if df_k is None:
        print(f'  K线获取失败: {code6} {st["name"]}')
        continue

    closes = df_k['close'].values
    if len(closes) < 6:
        continue

    try:
        score, details = chan_score(df_k)
    except Exception as e:
        print(f'  打分失败: {code6} {st["name"]} {e}')
        continue

    try:
        close_price = float(df_k['close'].iloc[-1])
    except:
        close_price = 0

    # MACD值
    ema12_s = pd.Series(closes).ewm(span=12, adjust=False).mean()
    ema26_s = pd.Series(closes).ewm(span=26, adjust=False).mean()
    macd_s = ema12_s - ema26_s
    macd_val = float(macd_s.iloc[-1])

    st['close'] = round(close_price, 2)
    st['macd'] = round(macd_val, 4)
    st['score'] = round(score, 1)
    st['details'] = details
    final_results.append(st)

    if (i + 1) % 10 == 0:
        print(f'  已分析 {i+1}/{len(stocks)} 只')

print(f'\n成功分析 {len(final_results)} 只')

# ===== 5. 按缠论分数排序 =====
final_results.sort(key=lambda x: -x['score'])

print('\n===== 最终排名 =====')
for i, r in enumerate(final_results):
    d = r['details']
    print(f"{i+1:2d}. {r['code']} {r['name']:8s} 分数:{r['score']:+.1f} "
          f"| 5日涨幅:{r['gain_5d']:+.1f}% | {d.get('position','')} "
          f"| {d.get('trend','')} | {d.get('fx','')} | {d.get('bcie','')}")

# ===== 6. 写入 Excel Sheet6 =====
print('\n===== Step4: 写入Excel =====')
wb = load_workbook(r'D:\stock\tool\stock\data\板块轮动Top10_v4_含非Top3强势个股.xlsx')
ws6_name = 'MACD强势个股'
if ws6_name in wb.sheetnames:
    del wb[ws6_name]

ws6 = wb.create_sheet(ws6_name)

thin = Side(style='thin', color='CCCCCC')
border = Border(left=thin, right=thin, top=thin, bottom=thin)

headers = ['排名', '代码', '名称', '最新价', 'MACD', '5日涨幅%', '缠论分数',
           '位置', '趋势', '分型', '背驰', '量能', '备注']
col_widths = [6, 10, 14, 10, 10, 12, 10, 18, 22, 12, 14, 12, 34]

for i, (h, w) in enumerate(zip(headers, col_widths), 1):
    ws6.column_dimensions[get_column_letter(i)].width = w
    c = ws6.cell(row=1, column=i, value=h)
    c.font = Font(bold=True, color='FFFFFF', size=11)
    c.fill = PatternFill(start_color='2F5496', end_color='2F5496', fill_type='solid')
    c.alignment = Alignment(horizontal='center', vertical='center')
    c.border = border

ws6.freeze_panes = 'A2'

RANK_COLORS = ['C0392B', 'E74C3C', 'E67E22', 'F39C12', '27AE60',
               '2ECC71', '3498DB', '8E44AD', '16A085', '7F8C8D']

for row_i, r in enumerate(final_results, 2):
    d = r['details']
    rank = row_i - 2
    rank_color = RANK_COLORS[min(rank - 1, len(RANK_COLORS) - 1)]
    score_val = r['score']

    row_data = [
        rank,
        r['code'],
        r['name'],
        r['close'],
        r['macd'],
        r['gain_5d'],
        score_val,
        d.get('position', ''),
        d.get('trend', ''),
        d.get('fx', ''),
        d.get('bcie', ''),
        d.get('vol', ''),
        f"ZG:{d.get('zg','')} ZD:{d.get('zd','')}",
    ]
    for col_i, val in enumerate(row_data, 1):
        c = ws6.cell(row=row_i, column=col_i, value=val)
        c.border = border
        c.alignment = Alignment(horizontal='center', vertical='center')
        if col_i == 1:
            c.fill = PatternFill(start_color=rank_color, end_color=rank_color, fill_type='solid')
            c.font = Font(bold=True, color='FFFFFF')
        elif col_i == 7:
            if score_val >= 3:
                c.fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
                c.font = Font(bold=True, color='006100')
            elif score_val >= 1:
                c.fill = PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid')
                c.font = Font(bold=True, color='9C5700')
            elif score_val <= -1:
                c.fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
                c.font = Font(bold=True, color='9C0006')
        elif col_i == 6:
            if float(val) > 20: c.font = Font(bold=True, color='FF0000')
            elif float(val) > 15: c.font = Font(bold=True, color='FF4500')

wb.save(r'D:\stock\tool\stock\data\板块轮动Top10_v4_含非Top3强势个股.xlsx')
print(f'Sheet6 已写入，共 {len(final_results)} 只')
