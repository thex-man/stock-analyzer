"""
在已有Excel中新增Sheet：每天强势板块Top3 + 板块内当日涨幅>8%的个股
=========================================================================
v9 修复（2026-08-26）：
  - 只查2天：今日（0825）+ 昨日（0824）
  - 历史日期（0820等）：问财返回今日数据，写入无意义 → 显示"—"
  - 列名：涨跌幅:前复权 → 涨跌幅
  - 日期转换：缓存日期（系统日期）→ Excel显示（实际交易日）
  - 9:15前今日=0825，9:15后今日=0826
"""
import sys
sys.path.insert(0, '.')
import json
from pathlib import Path
import time
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime, time as dtime

# ========== 1. 日期转换 ==========
now = datetime.now()
cutoff = dtime(9, 15)
current_date_str = now.strftime('%Y%m%d')  # 20260826

if now.time() < cutoff:
    import datetime as dt
    trade_ts = now.replace(hour=0, minute=0, second=0, microsecond=0).timestamp() - 86400
    trade_date_str = dt.datetime.fromtimestamp(trade_ts).strftime('%Y%m%d')
else:
    trade_date_str = current_date_str

def cache_date_to_excel_date(cache_date_str):
    if now.time() < cutoff and cache_date_str == current_date_str:
        import datetime as dt
        trade_ts = now.replace(hour=0, minute=0, second=0, microsecond=0).timestamp() - 86400
        return dt.datetime.fromtimestamp(trade_ts).strftime('%Y-%m-%d')
    else:
        return f"{cache_date_str[:4]}-{cache_date_str[4:6]}-{cache_date_str[6:]}"

def fmt_date(d):
    return f"{d[:4]}-{d[4:6]}-{d[6:]}"

# ========== 2. 读取板块历史缓存，找每天Top3 ==========
cache_files = sorted(Path('data/board_history_ths').glob('history_*.json'))
cache_file = cache_files[-1]
with open(cache_file, encoding='utf-8') as f:
    raw = json.load(f)

all_dates = set()
for sym, info in raw.items():
    for row in info.get('data', []):
        if row.get('d'):
            all_dates.add(row['d'])

dates_sorted = sorted(all_dates, reverse=True)
last_10 = dates_sorted[:10]

daily_top3 = {}
for date in last_10:
    day_results = []
    for sym, info in raw.items():
        name = info.get('name', sym)
        btype = info.get('type', '')
        for row in info.get('data', []):
            if row.get('d') == date:
                pct = row.get('p', 0)
                day_results.append((pct, name, btype))
                break
    day_results.sort(reverse=True)
    daily_top3[date] = day_results[:3]

# ========== 3. 问财查询（修列名）============
def get_board_stocks_wenqi(board_name, date_str, board_type):
    """用问财查板块成分股，筛选>8%。返回 [{code, name, change}]"""
    from stock_data_source import wencai

    ds = fmt_date(date_str)
    query = f'{board_name} {ds} 涨跌幅排序'

    try:
        result = wencai(query, perpage=200)
    except Exception:
        return []

    df = result.get('datas')
    if df is None or (hasattr(df, 'empty') and df.empty):
        return []

    stocks = []
    for _, row in df.iterrows():
        try:
            pct = row.get('涨跌幅')  # 修复：不是涨跌幅:前复权
            if pct is None:
                continue
            pct = float(pct)
            if pct > 8.0:
                code = str(row.get('股票代码', '')).strip()
                name = str(row.get('股票简称', '')).strip()
                if code and name:
                    stocks.append({'code': code, 'name': name, 'change': round(pct, 2)})
        except (ValueError, TypeError):
            continue
    stocks.sort(key=lambda x: -x['change'])
    return stocks

# ========== 4. 只处理今日+昨日，其他标记跳过 ==========
# 今日 = cache_today（系统日期20260826），昨日 = last_10[1]（20260824）
cache_today = current_date_str
target_dates = [cache_today, last_10[1]]  # [20260826, 20260824]

print(f"系统时间: {now.strftime('%Y-%m-%d %H:%M')}")
print(f"9:15前: {'是' if now.time() < cutoff else '否'}")
print(f"实际交易日: {trade_date_str}")
print(f"只处理: {target_dates}（其他 {len(last_10)-2} 天跳过）")
print("=" * 60)

all_rows = []  # (cache_date, excel_date, board_pct, board_name, btype, stocks)

for date in target_dates:
    excel_date = cache_date_to_excel_date(date)
    print(f"\n【{date}】（缓存）→ Excel显示 {excel_date}")
    if date not in daily_top3:
        print(f"  缓存中无此日期，跳过")
        continue

    for board_pct, board_name, btype in daily_top3[date]:
        # 只有今日（cache_today）用问财查；昨日跳过（问财返回今日数据无意义）
        if date == cache_today:
            stocks = get_board_stocks_wenqi(board_name, trade_date_str, btype)
        else:
            stocks = []
            print(f"  [跳过] 历史日期问财返回今日数据")

        if stocks:
            print(f"  {board_name}  {board_pct:+.2f}%  → {len(stocks)}只>8%  TOP1:{stocks[0]['name']}")
        else:
            print(f"  {board_name}  {board_pct:+.2f}%  → 无>8%")
        all_rows.append((date, excel_date, board_pct, board_name, btype, stocks))
    time.sleep(1)

# 关闭 Selenium
try:
    from stock_data_source import _close_selenium_driver
    _close_selenium_driver()
except:
    pass

# ========== 5. 写入Excel（新Sheet） ==========
wb = load_workbook('data/板块轮动Top10_v2_行业概念分开.xlsx')

if '每日Top3强势个股' in wb.sheetnames:
    del wb['每日Top3强势个股']

ws = wb.create_sheet('每日Top3强势个股')

thin = Side(style='thin', color='CCCCCC')
border = Border(left=thin, right=thin, top=thin, bottom=thin)

col_widths = [14, 16, 10, 18, 14, 12]
col_headers = ['日期', '板块', '板块涨幅', '类型', '强势个股', '个股涨幅']
for i, (w, h) in enumerate(zip(col_widths, col_headers), 1):
    ws.column_dimensions[get_column_letter(i)].width = w
    c = ws.cell(row=1, column=i, value=h)
    c.font = Font(bold=True, color='FFFFFF', size=11)
    c.fill = PatternFill(start_color='2F5496', end_color='2F5496', fill_type='solid')
    c.alignment = Alignment(horizontal='center', vertical='center')
    c.border = border

ws.freeze_panes = 'A2'

row_num = 2
for date in target_dates:
    day_boards = [(bp, bn, bt, ss, ed) for (d, ed, bp, bn, bt, ss) in all_rows if d == date]
    if not day_boards:
        continue
    date_start_row = row_num

    for board_pct, board_name, btype, stocks, excel_date in day_boards:
        if stocks:
            c_date = ws.cell(row=row_num, column=1, value=excel_date)
            c_board = ws.cell(row=row_num, column=2, value=board_name)
            c_bpct = ws.cell(row=row_num, column=3, value=f"{board_pct:+.2f}%")
            c_btype = ws.cell(row=row_num, column=4, value=btype)
            c_stock = ws.cell(row=row_num, column=5, value=f"{stocks[0]['name']} {stocks[0]['code']}")
            c_spct = ws.cell(row=row_num, column=6, value=f"{stocks[0]['change']:+.2f}%")
            for c in [c_date, c_board, c_bpct, c_btype, c_stock, c_spct]:
                c.border = border
                c.alignment = Alignment(horizontal='center', vertical='center')
            c_bpct.alignment = Alignment(horizontal='right', vertical='center')
            c_spct.alignment = Alignment(horizontal='right', vertical='center')
            c_spct.font = Font(bold=True, color='C00000')

            for j in range(1, len(stocks)):
                row_num += 1
                ws.cell(row=row_num, column=1, value='')
                ws.merge_cells(start_row=row_num, start_column=1, end_row=row_num, end_column=4)
                c_stock = ws.cell(row=row_num, column=5, value=f"{stocks[j]['name']} {stocks[j]['code']}")
                c_spct = ws.cell(row=row_num, column=6, value=f"{stocks[j]['change']:+.2f}%")
                for col_i in range(1, 5):
                    ws.cell(row=row_num, column=col_i).border = border
                for c in [c_stock, c_spct]:
                    c.border = border
                    c.alignment = Alignment(horizontal='center', vertical='center')
                c_spct.alignment = Alignment(horizontal='right', vertical='center')
                c_spct.font = Font(bold=True, color='C00000')
        else:
            c_date = ws.cell(row=row_num, column=1, value=excel_date)
            c_board = ws.cell(row=row_num, column=2, value=board_name)
            c_bpct = ws.cell(row=row_num, column=3, value=f"{board_pct:+.2f}%")
            c_btype = ws.cell(row=row_num, column=4, value=btype)
            c_stock = ws.cell(row=row_num, column=5, value='—')
            c_spct = ws.cell(row=row_num, column=6, value='—')
            for c in [c_date, c_board, c_bpct, c_btype, c_stock, c_spct]:
                c.border = border
                c.alignment = Alignment(horizontal='center', vertical='center')
            c_bpct.alignment = Alignment(horizontal='right', vertical='center')

        row_num += 1

    if date_start_row < row_num - 1:
        ws.merge_cells(start_row=date_start_row, start_column=1, end_row=row_num - 1, end_column=1)
        merged = ws.cell(row=date_start_row, column=1)
        merged.alignment = Alignment(horizontal='center', vertical='center')
        merged.font = Font(bold=True)

    row_num += 1

out_path = 'data/板块轮动Top10_v3_含每日Top3强势个股.xlsx'
wb.save(out_path)
print(f"\n已保存: {out_path}")

total = sum(len(ss) for _, _, _, _, _, ss in all_rows)
print(f"共 {len(target_dates)} 天，{total} 条强势个股（其他 {len(last_10)-len(target_dates)} 天无数据）")
